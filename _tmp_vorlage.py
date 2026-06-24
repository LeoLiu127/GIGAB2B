import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook(r'D:\ziniao browser\BAT-US\Euro\PLANTER-de.xlsm', data_only=True, keep_vba=False)
ws = wb['Vorlage']

print('Max row: %d, Max col: %d' % (ws.max_row, ws.max_column))
# Print rows 1-10
for r in range(1, 11):
    parts = []
    for cell in ws[r]:
        if cell.value is not None:
            v = str(cell.value)
            parts.append('C%d=%s' % (cell.column, repr(v[:80])))
    if parts:
        print('Row %d: ' % r + ' || '.join(parts[:20]))
