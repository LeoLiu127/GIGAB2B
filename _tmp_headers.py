import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook(r'D:\ziniao browser\BAT-US\Euro\PLANTER-de.xlsm', data_only=True, keep_vba=False)
ws = wb['Vorlage']

# Row 4 has actual headers, row 5 has internal field names
headers = {}
fieldnames = {}
for cell in ws[4]:
    if cell.value is not None:
        headers[cell.column] = str(cell.value)
for cell in ws[5]:
    if cell.value is not None:
        fieldnames[cell.column] = str(cell.value)

print('=== HEADERS (Row 4) ===')
for col, name in sorted(headers.items()):
    print('Col %d: %s' % (col, name[:200]))

print()
print('=== FIELD NAMES (Row 5) ===')
for col, name in sorted(fieldnames.items()):
    print('Col %d: %s' % (col, name[:200]))

print()
print('=== SAMPLE DATA (Row 7) ===')
for cell in ws[7]:
    if cell.value is not None:
        print('Col %d: %s' % (cell.column, repr(str(cell.value)[:200])))
