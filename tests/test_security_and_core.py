from __future__ import annotations

import io
import os
import tempfile
import time
import unittest
from unittest.mock import patch

import openpyxl

import app as app_module


def workbook_bytes() -> bytes:
    stream = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Vorlage")
    sheet.cell(row=1, column=271, value="template-marker")
    sheet.cell(row=7, column=1, value="data-row-marker")
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


class SecurityAndCoreTests(unittest.TestCase):
    def setUp(self) -> None:
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()

    def tearDown(self) -> None:
        app_module._LOGIN_FAILURES.clear()

    def test_template_upload_rejects_non_spreadsheet_and_cannot_overwrite_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            with patch.object(app_module, "TEMPLATE_DIR", tmp):
                response = self.client.post(
                    "/api/upload-template",
                    data={"file": (io.BytesIO(b"print('owned')"), "app.py")},
                    content_type="multipart/form-data",
                )

        self.assertEqual(response.status_code, 400)
        self.assertIn("xls", response.get_json()["error"].lower())

    def test_chinese_template_name_keeps_a_real_extension_and_uses_unique_storage_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            with patch.object(app_module, "TEMPLATE_UPLOAD_DIR", tmp):
                response = self.client.post(
                    "/api/upload-template",
                    data={"file": (io.BytesIO(workbook_bytes()), "英国模板.xlsm")},
                    content_type="multipart/form-data",
                )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["filename"].endswith(".xlsm"))
        self.assertNotEqual(payload["filename"], "xlsm")
        self.assertEqual(payload["detected_market"], "UK")

    def test_private_and_loopback_image_urls_are_rejected(self):
        for url in (
            "http://127.0.0.1:5182/api/health",
            "http://169.254.169.254/latest/meta-data/",
            "http://10.0.0.4/internal.png",
        ):
            with self.subTest(url=url):
                with self.assertRaises(ValueError):
                    app_module._validate_remote_image_url(url)

    def test_giga_image_validation_accepts_rotated_signature_for_same_cdn_path(self):
        requested = "https://cdn.example/products/image-1.jpg?x-ct=200&x-cs=new-signature"
        allowed = {"https://cdn.example/products/image-1.jpg?x-ct=100&x-cs=old-signature"}

        with patch.object(app_module, "_validate_remote_image_url", side_effect=lambda url: url):
            validated = app_module._validate_giga_reference_url(requested, allowed)

        self.assertEqual(validated, requested)

    def test_giga_image_validation_rejects_different_cdn_path_even_with_similar_signature(self):
        requested = "https://cdn.example/products/other.jpg?x-cs=same"
        allowed = {"https://cdn.example/products/image-1.jpg?x-cs=same"}

        with self.assertRaises(ValueError):
            app_module._validate_giga_reference_url(requested, allowed)

    def test_market_detection_prioritizes_taxfree_and_token_boundaries(self):
        self.assertEqual(app_module._detect_market_from_template("PLANTER-de-taxfree.xlsm"), "DE_TAXFREE")
        self.assertEqual(app_module._detect_market_from_template("garden-template-uk.xlsm"), "UK")
        self.assertEqual(app_module._detect_market_from_template("法国模板.xlsx"), "FR")

    def test_reference_images_keep_main_and_detail_groups_separate(self):
        item = {
            "sku": "SKU-1",
            "productName": "Sideboard",
            "mainImageUrl": "https://cdn.example/main.jpg",
            "imageUrls": [f"https://cdn.example/detail-{i}.jpg" for i in range(1, 9)],
        }

        view = app_module._assemble_variant_view(item, is_main=True)

        self.assertEqual(view["mainImageUrls"], ["https://cdn.example/main.jpg"])
        self.assertEqual(
            view["detailImageUrls"],
            [f"https://cdn.example/detail-{i}.jpg" for i in range(1, 7)],
        )
        self.assertEqual(
            view["imageUrls"],
            ["https://cdn.example/main.jpg", *[f"https://cdn.example/detail-{i}.jpg" for i in range(1, 9)]],
        )
        self.assertEqual(view["main_image_count"], 1)
        self.assertEqual(view["detail_image_count"], 6)
        self.assertEqual(view["image_count"], 9)

    def test_listing_reference_images_keep_each_active_variant_color_isolated(self):
        main = {
            "sku": "MAIN",
            "productName": "Main",
            "mainImageUrl": "https://cdn.example/main.jpg",
            "imageUrls": [f"https://cdn.example/main-detail-{i}.jpg" for i in range(1, 8)],
        }
        variants = [
            {
                "sku": f"VAR-{i}",
                "productName": f"Variant {i}",
                "mainImageUrl": f"https://cdn.example/variant-{i}.jpg",
                "imageUrls": [f"https://cdn.example/variant-{i}-detail.jpg"],
            }
            for i in range(1, 12)
        ]

        active = app_module._assemble_variant_view(main, is_main=True)
        grouped = app_module._with_listing_reference_images(active, [main, *variants])

        self.assertEqual(grouped["mainImageUrls"], ["https://cdn.example/main.jpg"])
        self.assertEqual(len(grouped["detailImageUrls"]), 6)
        self.assertEqual(
            grouped["imageUrls"],
            ["https://cdn.example/main.jpg", *[f"https://cdn.example/main-detail-{i}.jpg" for i in range(1, 8)]],
        )
        self.assertFalse(any("variant-" in url for url in grouped["imageUrls"]))
        self.assertEqual(grouped["main_image_count"], 1)
        self.assertEqual(grouped["detail_image_count"], 6)

    def test_reference_image_candidates_are_not_truncated_before_dimension_classification(self):
        item = {
            "sku": "SKU-1",
            "mainImageUrl": "https://cdn.example/main.jpg",
            "imageUrls": [f"https://cdn.example/image-{i}.jpg" for i in range(1, 15)],
        }

        fields = app_module._build_reference_image_fields(item)

        self.assertEqual(len(fields["imageUrls"]), 15)
        self.assertEqual(fields["imageUrls"][-1], "https://cdn.example/image-14.jpg")

    def test_listing_grouping_preserves_all_active_candidates_for_later_dimension_classification(self):
        raw = {
            "sku": "SKU-1",
            "mainImageUrl": "https://cdn.example/main.jpg",
            "imageUrls": [f"https://cdn.example/image-{i}.jpg" for i in range(1, 15)],
        }
        active = app_module._assemble_variant_view(raw, is_main=True)

        grouped = app_module._with_listing_reference_images(active, [active])

        self.assertEqual(len(grouped["imageUrls"]), 15)
        self.assertEqual(grouped["imageUrls"][-1], "https://cdn.example/image-14.jpg")

    def test_dimension_classification_promotes_square_field_images_and_caps_groups(self):
        records = [
            {"index": 0, "originalUrl": "https://cdn.example/declared-main.jpg", "width": 1200, "height": 1200},
            *[
                {"index": i, "originalUrl": f"https://cdn.example/square-{i}.jpg", "width": 1500, "height": 1500}
                for i in range(1, 11)
            ],
            *[
                {"index": 11 + i, "originalUrl": f"https://cdn.example/detail-{i}.jpg", "width": 1600, "height": 900}
                for i in range(8)
            ],
        ]

        grouped = app_module._classify_reference_image_records(records, declared_main_urls={"https://cdn.example/declared-main.jpg"})

        self.assertEqual(len(grouped["main"]), 9)
        self.assertEqual(len(grouped["detail"]), 6)
        self.assertTrue(all(image["group"] == "main" for image in grouped["main"]))
        self.assertTrue(all(image["group"] == "detail" for image in grouped["detail"]))
        self.assertIn("https://cdn.example/square-1.jpg", [image["originalUrl"] for image in grouped["main"]])
        self.assertNotIn("https://cdn.example/square-10.jpg", [image["originalUrl"] for image in grouped["detail"]])

    def test_unknown_dimensions_fall_back_to_declared_source_group(self):
        records = [
            {"index": 0, "originalUrl": "https://cdn.example/main.jpg", "width": None, "height": None},
            {"index": 1, "originalUrl": "https://cdn.example/detail.jpg", "width": None, "height": None},
        ]

        grouped = app_module._classify_reference_image_records(records, declared_main_urls={"https://cdn.example/main.jpg"})

        self.assertEqual([image["originalUrl"] for image in grouped["main"]], ["https://cdn.example/main.jpg"])
        self.assertEqual([image["originalUrl"] for image in grouped["detail"]], ["https://cdn.example/detail.jpg"])

    def test_image_dimensions_reads_png_header_without_extra_dependencies(self):
        png = b"\x89PNG\r\n\x1a\n" + b"\x00\x00\x00\rIHDR" + (1600).to_bytes(4, "big") + (900).to_bytes(4, "big") + b"\x08\x02\x00\x00\x00"

        self.assertEqual(app_module._image_dimensions(png, "image/png"), (1600, 900))

    def test_fetch_images_classifies_the_exact_listing_candidates_requested_by_frontend(self):
        urls = [
            "https://cdn.example/main.jpg",
            "https://cdn.example/square-from-image-urls.jpg",
            "https://cdn.example/detail.jpg",
        ]

        def proxy(url: str) -> dict:
            if url.endswith("detail.jpg"):
                return {"dataUrl": f"data:{url}", "width": 1600, "height": 900}
            return {"dataUrl": f"data:{url}", "width": 1200, "height": 1200}

        with patch.object(app_module, "_allowed_giga_reference_urls", return_value=set(urls)), patch.object(
            app_module, "_proxy_image_with_metadata", side_effect=proxy
        ):
            response = self.client.post(
                "/api/fetch-images",
                json={"sku": "SKU-1", "market": "UK", "image_urls": urls, "declared_main_urls": [urls[0]]},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["raw_image_count"], 3)
        self.assertEqual(payload["main_image_count"], 2)
        self.assertEqual(payload["detail_image_count"], 1)
        self.assertEqual([image["index"] for image in payload["images"]], [0, 1, 2])
        self.assertEqual(payload["images"][1]["group"], "main")

    def test_fetch_images_rejects_frontend_candidate_not_returned_by_giga(self):
        with patch.object(app_module, "_allowed_giga_reference_urls", return_value={"https://cdn.example/allowed.jpg"}):
            response = self.client.post(
                "/api/fetch-images",
                json={"sku": "SKU-1", "market": "UK", "image_urls": ["https://evil.example/not-giga.jpg"]},
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("不属于当前 GIGA", response.get_json()["error"])

    def test_chair_prompt_does_not_inject_planter_keywords(self):
        product = {
            "sku": "CHAIR-1",
            "productName": "Velvet Chaise Lounge Chair",
            "category": "CHAIR",
            "mainMaterial": "Velvet",
            "attributes": {"Main Color": "Blue"},
        }

        prompt = app_module._build_copy_prompt(product, "UK")

        self.assertIn("Velvet Chaise Lounge Chair", prompt)
        self.assertNotIn("raised garden bed", prompt.lower())
        self.assertNotIn("galvanized steel", prompt.lower())

    def test_copy_prompt_includes_marketplace_title_bullet_and_ai_shopping_rules(self):
        product = {
            "sku": "SOFA-1",
            "productName": "Convertible Sofa Bed",
            "category": "Sofa Bed",
            "mainMaterial": "Chenille",
            "attributes": {"Main Color": "Dark Green"},
        }

        prompt = app_module._build_copy_prompt(
            product,
            "UK",
            prompt_extra="Make it warmer for apartment renters.",
            keywords=["small space sleeper sofa"],
        )

        self.assertIn("DEFAULT MARKETPLACE LISTING RULES", prompt)
        self.assertIn("capitalize the first letter of major words", prompt)
        self.assertIn("minor function words lowercase", prompt)
        self.assertIn("brand name first when the brand is real", prompt)
        self.assertIn("core product keyword", prompt)
        self.assertIn("Exact title blueprint", prompt)
        self.assertIn("Do not start the title with size, color, material, finish, or style", prompt)
        self.assertIn("Sideboard Cabinet", prompt)
        self.assertIn("Bad: Dark Oak Grain Sideboard", prompt)
        self.assertIn("scenario words", prompt)
        self.assertIn("strongest and most distinctive selling points first", prompt)
        self.assertIn("AI shopping assistants", prompt)
        self.assertIn("USER ADDITIONAL REQUIREMENTS TO MERGE", prompt)
        self.assertIn("Make it warmer for apartment renters.", prompt)

    def test_copy_prompt_priority_allows_user_preferences_but_not_platform_hard_rule_overrides(self):
        prompt = app_module._build_copy_prompt(
            {"sku": "SKU-1", "productName": "Sideboard Cabinet", "attributes": {}},
            "UK",
            prompt_extra="Put the room scenario before the dimensions.",
        )

        self.assertIn("PLATFORM HARD RULES AND PRODUCT FACTS HAVE HIGHEST PRIORITY", prompt)
        self.assertIn("override built-in default preferences", prompt)
        self.assertNotIn("If there is any conflict, the marketplace rules win", prompt)

    def test_copy_sanitizer_removes_numbering_and_special_bullet_markers(self):
        parsed = {
            "title": "Test Product",
            "bullets": [
                "1. **DURABLE FRAME**: Built for daily use",
                "2) EASY ASSEMBLY Installs quickly",
                "* SPACE SAVING Folds flat",
                "• COMFORTABLE Chenille surface",
                "— VERSATILE For living rooms",
            ],
            "description": "",
            "search_terms": "1. sofa bed, 2. sleeper couch",
        }

        cleaned = app_module._sanitize_copy(parsed)

        self.assertEqual(cleaned["bullets"][0], "DURABLE FRAME: Built for daily use")
        self.assertEqual(cleaned["bullets"][1], "EASY ASSEMBLY Installs quickly")
        self.assertEqual(cleaned["bullets"][2], "SPACE SAVING Folds flat")
        self.assertEqual(cleaned["bullets"][3], "COMFORTABLE Chenille surface")
        self.assertEqual(cleaned["bullets"][4], "VERSATILE For living rooms")
        self.assertEqual(cleaned["search_terms"], "sofa bed, 2. sleeper couch")

    def test_title_order_repair_moves_core_product_before_finish_and_size(self):
        product = {
            "productName": "150x40x80cm dark oak grain sideboard, 3 drawers, large sideboard, kitchen cabinet",
            "category": "Sideboard",
            "attributes": {"Main Color": "Dark Oak"},
        }
        title = (
            "Dark Oak Grain Sideboard 150x40x80cm - 3 Drawers, Adjustable Shelves, "
            "Mid-Century Modern Cabinet for Kitchen, Living Room, Hallway"
        )

        repaired = app_module._repair_title_order(title, product)

        self.assertTrue(repaired.startswith("Sideboard Cabinet, "))
        self.assertIn("150x40x80cm", repaired)
        self.assertIn("3 Drawers", repaired)
        self.assertTrue(repaired.endswith("Dark Oak Grain"))
        self.assertFalse(repaired.startswith("Dark Oak Grain"))

    def test_ai_response_applies_title_order_repair_before_returning_to_ui(self):
        product = {
            "sku": "CABINET-1",
            "productName": "150x40x80cm dark oak grain sideboard, 3 drawers, large sideboard, kitchen cabinet",
            "category": "Sideboard",
        }
        gen = {
            "ok": True,
            "attempts": 1,
            "raw": {
                "choices": [{
                    "message": {
                        "content": (
                            "### Product Title\n"
                            "Dark Oak Grain Sideboard 150x40x80cm - 3 Drawers, Adjustable Shelves, "
                            "Mid-Century Modern Cabinet for Kitchen, Living Room, Hallway\n\n"
                            "### Five Bullet Points\n"
                            "1. STYLISH STORAGE Designed for dining rooms and hallways with practical shelves.\n\n"
                            "### Product Description\n"
                            "A practical sideboard cabinet.\n\n"
                            "### Search Terms\n"
                            "sideboard cabinet, storage cabinet"
                        )
                    }
                }]
            },
        }

        parsed = app_module._parse_ai_response(gen, product, "UK")

        self.assertTrue(parsed["title"].startswith("Sideboard Cabinet, "))
        self.assertFalse(parsed["title"].startswith("Dark Oak Grain"))

    def test_structured_non_planter_category_cannot_be_overridden_by_title(self):
        product = {"category": "CHAIR", "productName": "Chair beside a flower pot"}
        self.assertFalse(app_module._is_planter_product(product))

    def test_planter_excel_writer_fails_closed_for_unsupported_category(self):
        with tempfile.TemporaryDirectory() as templates, tempfile.TemporaryDirectory() as outputs:
            template_path = os.path.join(templates, "PLANTER-uk.xlsm")
            workbook = openpyxl.Workbook()
            workbook.create_sheet("Vorlage")
            workbook.save(template_path)
            workbook.close()
            product = {
                "sku": "CHAIR-1",
                "productName": "Velvet Chaise Lounge Chair",
                "category": "CHAIR",
            }
            with patch.object(app_module, "TEMPLATE_DIR", templates), patch.object(
                app_module, "EXCEL_OUTPUT_DIR", outputs
            ):
                with self.assertRaisesRegex(ValueError, "PLANTER"):
                    app_module.fill_excel(product, {}, "UK", "PLANTER-uk.xlsm")

    def test_generated_image_filename_is_unique_and_matches_png_content(self):
        png = b"\x89PNG\r\n\x1a\n" + b"payload"
        encoded = __import__("base64").b64encode(png).decode("ascii")
        with tempfile.TemporaryDirectory() as tmp:
            with patch.object(app_module, "OUTPUT_DIR", tmp):
                first = app_module._save_base64_to_outputs("SKU-1", "main", encoded)
                second = app_module._save_base64_to_outputs("SKU-1", "main", encoded)

        self.assertTrue(first["filename"].endswith(".png"))
        self.assertNotEqual(first["filename"], second["filename"])

    def test_generated_image_public_url_uses_current_request_origin(self):
        with app_module.app.test_request_context(base_url="http://127.0.0.1:5173/"):
            public = app_module._public_output_url("/outputs/SKU-1/main.png")

        self.assertEqual(public, "http://127.0.0.1:5173/outputs/SKU-1/main.png")

    def test_output_cleanup_removes_old_images_and_keeps_recent_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            sku_dir = os.path.join(tmp, "SKU-1")
            os.makedirs(sku_dir)
            old_path = os.path.join(sku_dir, "old.png")
            recent_path = os.path.join(sku_dir, "recent.png")
            with open(old_path, "wb") as output:
                output.write(b"old")
            with open(recent_path, "wb") as output:
                output.write(b"recent")
            now = time.time()
            os.utime(old_path, (now - 15 * 86400, now - 15 * 86400))
            os.utime(recent_path, (now, now))
            with patch.object(app_module, "OUTPUT_DIR", tmp):
                removed = app_module._cleanup_old_outputs(max_age_days=7, now=now)

            self.assertEqual(removed, 1)
            self.assertFalse(os.path.exists(old_path))
            self.assertTrue(os.path.exists(recent_path))

    def test_generation_prompt_contains_amazon_image_compliance_rules(self):
        copy = {
            "title": "Convertible Sofa Bed",
            "bullets": ["SPACE SAVING Converts from couch to sleeper"],
            "description": "Compact sleeper sofa for apartments",
            "search_terms": "sofa bed, sleeper couch",
        }
        product = {"mainColor": "Dark Green", "mainMaterial": "Chenille", "size": "77 inch"}

        main_prompt, _, _ = app_module._build_generation_prompt("main", "1600x1600", copy, product, "show a cozy room")
        sub_prompt, _, _ = app_module._build_generation_prompt("sub", "1600x1600", copy, product, "")

        self.assertIn("AMAZON IMAGE COMPLIANCE RULES", main_prompt)
        self.assertIn("pure white background RGB 255, 255, 255", main_prompt)
        self.assertIn("product must fill about 85% of the image frame", main_prompt)
        self.assertIn("No text, logos, borders, color blocks, watermarks", main_prompt)
        self.assertIn("No Amazon logo", main_prompt)
        self.assertIn("No buyer reviews, star ratings", sub_prompt)
        self.assertIn("USER ADDITIONAL REQUIREMENTS TO MERGE", main_prompt)

    def test_generation_prompt_priority_distinguishes_hard_compliance_from_creative_defaults(self):
        prompt, _, _ = app_module._build_generation_prompt(
            "sub",
            "1600x1600",
            {"title": "Sideboard", "bullets": []},
            {"mainColor": "Oak", "mainMaterial": "Wood"},
            "Use a warm dining room and emphasize storage capacity.",
        )

        self.assertIn("AMAZON HARD COMPLIANCE AND PRODUCT IDENTITY HAVE HIGHEST PRIORITY", prompt)
        self.assertIn("override built-in creative defaults", prompt)
        self.assertNotIn("If there is any conflict, the Amazon image compliance rules win", prompt)

    def test_uploaded_template_path_must_stay_inside_allowed_directories(self):
        with tempfile.TemporaryDirectory() as templates, tempfile.TemporaryDirectory() as uploads:
            with patch.object(app_module, "TEMPLATE_DIR", templates), patch.object(
                app_module, "TEMPLATE_UPLOAD_DIR", uploads
            ):
                outside = os.path.abspath(os.path.join(templates, "..", "secret.xlsm"))
                with self.assertRaises(ValueError):
                    app_module._resolve_template("UK", outside)

    def test_upload_rejects_workbook_without_required_template_structure(self):
        stream = io.BytesIO()
        workbook = openpyxl.Workbook()
        workbook.save(stream)
        workbook.close()
        with tempfile.TemporaryDirectory() as tmp:
            with patch.object(app_module, "TEMPLATE_UPLOAD_DIR", tmp):
                response = self.client.post(
                    "/api/upload-template",
                    data={"file": (io.BytesIO(stream.getvalue()), "uk-template.xlsx")},
                    content_type="multipart/form-data",
                )
        self.assertEqual(response.status_code, 400)
        self.assertIn("Vorlage", response.get_json()["error"])

    def test_giga_reference_url_must_match_server_fetched_product_urls(self):
        allowed = {"https://cdn.gigab2b.example/product-1.jpg"}
        with patch.object(app_module, "_validate_remote_image_url", side_effect=lambda url: url):
            self.assertEqual(
                app_module._validate_giga_reference_url(next(iter(allowed)), allowed),
                next(iter(allowed)),
            )
            with self.assertRaises(ValueError):
                app_module._validate_giga_reference_url("https://attacker.example/metadata", allowed)

    def test_api_is_open_by_default_while_authentication_is_bypassed(self):
        status = self.client.get("/api/auth/status")
        markets = self.client.get("/api/markets")

        self.assertEqual(status.status_code, 200)
        self.assertEqual(status.get_json(), {"required": False, "authenticated": True})
        self.assertEqual(markets.status_code, 200)

    def test_api_requires_login_when_access_password_is_configured(self):
        with patch.object(app_module, "AUTH_ENABLED", True), patch.object(
            app_module, "ACCESS_PASSWORD", "correct-horse-battery-staple"
        ):
            anonymous = self.client.get("/api/markets")
            wrong = self.client.post("/api/auth/login", json={"password": "wrong"})
            accepted = self.client.post(
                "/api/auth/login", json={"password": "correct-horse-battery-staple"}
            )
            authenticated = self.client.get("/api/markets")

        self.assertEqual(anonymous.status_code, 401)
        self.assertEqual(wrong.status_code, 401)
        self.assertEqual(accepted.status_code, 200)
        self.assertEqual(authenticated.status_code, 200)

    def test_login_throttles_repeated_wrong_passwords(self):
        app_module._LOGIN_FAILURES.clear()
        with patch.object(app_module, "AUTH_ENABLED", True), patch.object(
            app_module, "ACCESS_PASSWORD", "correct-horse-battery-staple"
        ):
            responses = [
                self.client.post("/api/auth/login", json={"password": "wrong"})
                for _ in range(6)
            ]

        self.assertTrue(all(response.status_code == 401 for response in responses[:5]))
        self.assertEqual(responses[5].status_code, 429)

    def test_excel_output_has_authenticated_download_route(self):
        with tempfile.TemporaryDirectory() as tmp:
            filename = "SKU-UK-abc123.xlsm"
            with open(os.path.join(tmp, filename), "wb") as output:
                output.write(b"excel")
            with patch.object(app_module, "EXCEL_OUTPUT_DIR", tmp):
                response = self.client.get(f"/api/downloads/{filename}")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.data, b"excel")
                self.assertIn("attachment", response.headers.get("Content-Disposition", ""))
                response.close()

    def test_generic_output_route_does_not_serve_excel_or_templates(self):
        with tempfile.TemporaryDirectory() as tmp:
            os.makedirs(os.path.join(tmp, "excel"), exist_ok=True)
            with open(os.path.join(tmp, "excel", "secret.xlsm"), "wb") as output:
                output.write(b"excel")
            with patch.object(app_module, "OUTPUT_DIR", tmp):
                response = self.client.get("/outputs/excel/secret.xlsm")
                self.assertEqual(response.status_code, 404)
                response.close()

    def test_generate_image_accepts_only_server_verified_giga_reference(self):
        png = b"\x89PNG\r\n\x1a\n" + b"payload"
        encoded = __import__("base64").b64encode(png).decode("ascii")
        allowed_url = "https://example.com/product-1.png"
        with tempfile.TemporaryDirectory() as tmp, patch.dict(
            app_module.LAOZHANG_CONFIG, {"api_key": "test-key"}
        ), patch.object(
            app_module, "giga_fetch_product", return_value={"imageUrls": [allowed_url]}
        ), patch.object(
            app_module, "_validate_remote_image_url", side_effect=lambda url: url
        ), patch.object(
            app_module, "_proxy_image", return_value=f"data:image/png;base64,{encoded}"
        ), patch.object(
            app_module,
            "_generate_image_local",
            return_value={"ok": True, "data": {"images": [{"base64": encoded}]}},
        ), patch.object(app_module, "OUTPUT_DIR", tmp):
            response = self.client.post(
                "/api/generate-image",
                json={
                    "sku": "SKU-1",
                    "market": "UK",
                    "slot": "main",
                    "size": "1600x1600",
                    "reference_images": [{"source": "giga", "index": 0, "url": allowed_url}],
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["filename"].endswith(".png"))

    def test_generate_image_rejects_unverified_or_malformed_references(self):
        allowed_url = "https://example.com/product-1.png"
        with patch.dict(app_module.LAOZHANG_CONFIG, {"api_key": "test-key"}), patch.object(
            app_module, "giga_fetch_product", return_value={"imageUrls": [allowed_url]}
        ), patch.object(app_module, "_validate_remote_image_url", side_effect=lambda url: url):
            unverified = self.client.post(
                "/api/generate-image",
                json={
                    "sku": "SKU-1",
                    "market": "UK",
                    "reference_images": [
                        {"source": "giga", "index": 0, "url": "https://attacker.example/internal"}
                    ],
                },
            )
            malformed = self.client.post(
                "/api/generate-image",
                json={"sku": "SKU-1", "market": "UK", "reference_images": ["not-an-object"]},
            )

        self.assertEqual(unverified.status_code, 400)
        self.assertEqual(malformed.status_code, 400)


if __name__ == "__main__":
    unittest.main()
