from __future__ import annotations

import unittest
import tempfile
import io
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

import app as app_module
from template_filler.models import SkuRow, TemplateField, TemplateProfile
from template_filler.mapping import build_fill_plan
from template_filler.policy import POLICY_ACTION_REQUIRED, PolicyRule, TemplatePolicy
from template_filler.variants import ListingProducts, VariantRow, expand_variant_rows
from template_filler.writer import write_filled_workbook


def _field(base_name: str, column: int, allowed_values: tuple[str, ...] = ()) -> TemplateField:
    return TemplateField(
        field_id=f"{base_name}#1.value",
        base_name=base_name,
        column=column,
        column_letter="A",
        label=base_name,
        requirement="optional",
        is_dropdown=bool(allowed_values),
        allowed_values=allowed_values,
    )


def _profile() -> TemplateProfile:
    return TemplateProfile(
        source_path=None,  # type: ignore[arg-type]
        platform="amazon",
        market="UK",
        marketplace_id="A1F83G8C2ARO7P",
        language_tag="en_GB",
        category="CHAIR",
        sheet_name="Template",
        label_row=4,
        attribute_row=5,
        data_row=7,
        fields=(
            _field("contribution_sku", 1),
            _field("parentage_level", 2, ("Parent", "Child")),
            _field("child_parent_sku_relationship", 3),
            _field("variation_theme", 4, ("COLOR/MATERIAL", "COLOR/SIZE")),
            _field("color", 5),
            _field("material", 6),
        ),
        sku_rows=(SkuRow(7, "A"),),
    )


class VariantExpansionTests(unittest.TestCase):
    def test_template_listing_fetch_reuses_main_collector_and_skips_unavailable_associations(self):
        original_listing = app_module.giga_fetch_listing
        calls = []
        try:
            def fake_listing(sku, market, include_variants=True):
                calls.append((sku, market, include_variants))
                return {
                    "parent_sku": sku,
                    "main": {"sku": "A", "productName": "A"},
                    "variants": [],
                    "raw_products": [
                        {"sku": "A", "productName": "A"},
                        {"sku": "B", "productName": "B"},
                    ],
                    "requested_skus": ["A", "B", "C", "D"],
                    "skipped_skus": ["C", "D"],
                    "truncated": False,
                    "fetch_error": None,
                    "warning": None,
                }

            app_module.giga_fetch_listing = fake_listing
            listing = app_module.giga_fetch_listing_products("A", "UK")
        finally:
            app_module.giga_fetch_listing = original_listing

        self.assertEqual(calls, [("A", "UK", True)])
        self.assertEqual(listing["requested_skus"], ["A", "B"])
        self.assertEqual([item["sku"] for item in listing["products"]], ["A", "B"])
        self.assertEqual(listing["missing_skus"], [])
        self.assertIn("C", listing["warning"])
        self.assertIn("D", listing["warning"])

    def test_main_listing_collector_exposes_raw_products_and_skipped_skus(self):
        original_product = app_module.giga_fetch_product
        original_bulk = app_module.giga_fetch_products_bulk
        try:
            app_module.giga_fetch_product = lambda sku, market: {
                "sku": "A", "productName": "A", "associateProductList": ["B", "C"]
            }
            app_module.giga_fetch_products_bulk = lambda skus, market: [{"sku": "A", "productName": "A"}, {"sku": "B", "productName": "B"}]
            listing = app_module.giga_fetch_listing("A", "UK")
        finally:
            app_module.giga_fetch_product = original_product
            app_module.giga_fetch_products_bulk = original_bulk

        self.assertEqual([item["sku"] for item in listing["raw_products"]], ["A", "B"])
        self.assertEqual(listing["requested_skus"], ["A", "B", "C"])
        self.assertEqual(listing["skipped_skus"], ["C"])
        self.assertFalse(listing["truncated"])
        self.assertIsNone(listing["fetch_error"])

    def test_expands_a_complete_listing_into_parent_and_children_with_unique_allowed_theme(self):
        products = {
            "A": {"sku": "A", "mainColor": "Red", "mainMaterial": "Wood"},
            "B": {"sku": "B", "mainColor": "Blue", "mainMaterial": "Steel"},
        }
        listing = ListingProducts(seed_sku="A", main_sku="A", requested_skus=("A", "B"), products=products)

        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing)

        self.assertEqual([row.role for row in expansion.rows], ["parent", "child", "child"])
        self.assertEqual([row.sku for row in expansion.rows], ["A-PARENT", "A", "B"])
        self.assertEqual([row.parent_sku for row in expansion.rows], ["", "A-PARENT", "A-PARENT"])
        self.assertEqual({row.variation_theme for row in expansion.rows}, {"COLOR/MATERIAL"})
        self.assertFalse(expansion.issues)
        self.assertEqual(expansion.summary["groups_expanded"], 1)

    def test_keeps_seed_row_and_blocks_group_when_no_unique_theme_can_be_inferred(self):
        products = {
            "A": {"sku": "A", "mainColor": "Red", "mainMaterial": "Wood"},
            "B": {"sku": "B", "mainColor": "Red", "mainMaterial": "Steel"},
        }
        listing = ListingProducts(seed_sku="A", main_sku="A", requested_skus=("A", "B"), products=products)

        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing)

        self.assertEqual([(row.role, row.sku) for row in expansion.rows], [("seed", "A")])
        self.assertEqual([issue.status for issue in expansion.issues], ["variant_theme_unresolved"])
        self.assertEqual(expansion.summary["groups_blocked"], 1)

    def test_blocks_group_when_operator_theme_conflicts_with_inferred_theme(self):
        products = {
            "A": {"sku": "A", "mainColor": "Red", "mainMaterial": "Wood"},
            "B": {"sku": "B", "mainColor": "Blue", "mainMaterial": "Steel"},
        }
        listing = ListingProducts(seed_sku="A", main_sku="A", requested_skus=("A", "B"), products=products)

        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing, manual_themes={7: "COLOR/SIZE"})

        self.assertEqual([(row.role, row.sku) for row in expansion.rows], [("seed", "A")])
        self.assertEqual([issue.status for issue in expansion.issues], ["variant_manual_theme_conflict"])

    def test_blocks_listing_with_missing_raw_variant_data_without_partial_expansion(self):
        listing = ListingProducts(
            seed_sku="A", main_sku="A", requested_skus=("A", "B"),
            products={"A": {"sku": "A", "mainColor": "Red"}}, missing_skus=("B",),
        )

        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing)

        self.assertEqual([(row.role, row.sku) for row in expansion.rows], [("seed", "A")])
        self.assertEqual([issue.status for issue in expansion.issues], ["variant_fetch_incomplete"])

    def test_keeps_single_sku_listing_as_non_blocking_seed_row(self):
        listing = ListingProducts(
            seed_sku="A", main_sku="A", requested_skus=("A",),
            products={"A": {"sku": "A", "mainColor": "Red", "mainMaterial": "Wood"}},
        )

        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing)

        self.assertEqual([(row.role, row.sku) for row in expansion.rows], [("seed", "A")])
        self.assertFalse(expansion.issues)
        self.assertEqual(expansion.groups[0].status, "no_variants")

    def test_parent_preserves_seed_values_but_child_uses_its_own_data_and_child_only_policy(self):
        profile = _profile()
        fields = profile.fields + (
            _field("item_name", 5),
            _field("custom_child_value", 6),
        )
        profile = TemplateProfile(**{**profile.__dict__, "fields": fields})
        policy = TemplatePolicy(
            "amazon", "uk", "chair", 1, "test", {},
            {fields[-1].field_id: PolicyRule(POLICY_ACTION_REQUIRED, scope="child")},
        )
        rows = [
            VariantRow(7, 7, "A-PARENT", "parent", variation_theme="COLOR/MATERIAL", product={"sku": "A", "productName": "Main title"}),
            VariantRow(7, 8, "A", "child", "A-PARENT", "COLOR/MATERIAL", {"sku": "A", "productName": "Child title"}),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "seed.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Template"
            sheet.cell(7, 5, "Operator parent title")
            workbook.save(source)
            workbook.close()
            plan = build_fill_plan(profile, source, {"A": rows[-1].product}, policy=policy, rows=rows)

        self.assertNotIn((7, 5), plan.changes)
        self.assertEqual(plan.changes[(8, 5)], "Child title")
        self.assertEqual(plan.changes[(7, 1)], "A-PARENT")
        self.assertEqual(plan.changes[(8, 1)], "A")
        self.assertEqual(plan.changes[(7, 2)], "Parent")
        self.assertEqual(plan.changes[(8, 2)], "Child")
        self.assertEqual(plan.changes[(8, 3)], "A-PARENT")
        self.assertFalse(any(issue.status == "business_required" and issue.row == 7 for issue in plan.issues))
        self.assertTrue(any(issue.status == "business_required" and issue.row == 8 for issue in plan.issues))

    def test_policy_default_scope_applies_only_to_its_selected_variant_role(self):
        profile = _profile()
        title = _field("item_name", 5)
        profile = TemplateProfile(**{**profile.__dict__, "fields": profile.fields + (title,)})
        policy = TemplatePolicy("amazon", "uk", "chair", 1, "test", {}, {title.field_id: PolicyRule("default", "Parent title", "parent")})
        rows = [
            VariantRow(7, 7, "A-PARENT", "parent", product={"sku": "A", "productName": "GIGA title"}),
            VariantRow(7, 8, "A", "child", "A-PARENT", product={"sku": "A", "productName": "GIGA title"}),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "seed.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Template"
            workbook.save(source)
            workbook.close()
            plan = build_fill_plan(profile, source, {"A": rows[-1].product}, policy=policy, rows=rows)

        self.assertEqual(plan.changes[(7, title.column)], "Parent title")
        self.assertEqual(plan.changes[(8, title.column)], "GIGA title")

    def test_color_alias_used_for_theme_inference_is_written_to_child(self):
        products = {
            "A": {"sku": "A", "attributes": {"Colour": "Red", "Material": "Wood"}},
            "B": {"sku": "B", "attributes": {"Colour": "Blue", "Material": "Steel"}},
        }
        listing = ListingProducts(seed_sku="A", main_sku="A", requested_skus=("A", "B"), products=products)
        expansion = expand_variant_rows(_profile(), lambda _sku, _market: listing)
        self.assertEqual(expansion.groups[0].status, "expanded")
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "seed.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Template"
            workbook.save(source)
            workbook.close()
            plan = build_fill_plan(_profile(), source, products, rows=expansion.rows)

        color = _profile().field_by_base_name("color")
        self.assertEqual(plan.changes[(8, color.column)], "Red")

    def test_maps_every_inferred_variation_component_to_child_template_fields(self):
        profile = _profile()
        theme = _field("variation_theme", 4, ("FABRIC_TYPE",))
        fabric = _field("fabric_type", 5)
        profile = TemplateProfile(**{**profile.__dict__, "fields": profile.fields[:3] + (theme, fabric)})
        products = {
            "A": {"sku": "A", "attributes": {"Fabric Type": "Linen"}},
            "B": {"sku": "B", "attributes": {"Fabric Type": "Velvet"}},
        }
        listing = ListingProducts(seed_sku="A", main_sku="A", requested_skus=("A", "B"), products=products)
        expansion = expand_variant_rows(profile, lambda _sku, _market: listing)
        self.assertEqual(expansion.groups[0].variation_theme, "FABRIC_TYPE")
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "seed.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Template"
            workbook.save(source)
            workbook.close()
            plan = build_fill_plan(profile, source, products, rows=expansion.rows)

        self.assertEqual(plan.changes[(8, fabric.column)], "Linen")
        self.assertEqual(plan.changes[(9, fabric.column)], "Velvet")

    def test_writer_clones_template_row_for_child_and_extends_validation_without_copying_manual_values(self):
        profile = _profile()
        profile = TemplateProfile(**{**profile.__dict__, "fields": profile.fields + (_field("item_name", 5),)})
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "seed.xlsx"
            output = Path(temp_dir) / "filled.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Template"
            sheet.cell(7, 1, "A")
            sheet.cell(7, 3, "OLD-PARENT")
            sheet.cell(7, 5, "Manual parent title")
            validation = DataValidation(type="list", formula1='"Parent,Child"')
            sheet.add_data_validation(validation)
            validation.add("B7")
            full_height_validation = DataValidation(type="list", formula1='"Parent,Child"')
            sheet.add_data_validation(full_height_validation)
            full_height_validation.add("C7:C1048576")
            sheet.conditional_formatting.add("B7", CellIsRule(operator="equal", formula=["\"Child\""], fill=openpyxl.styles.PatternFill("solid", fgColor="FF0000")))
            sheet.merge_cells("A9:B9")
            sheet.auto_filter.ref = "A7:E9"
            sheet.cell(9, 6, "=A7")
            workbook.save(source)
            workbook.close()

            rows = [
                VariantRow(7, 7, "A-PARENT", "parent", variation_theme="COLOR/MATERIAL"),
                VariantRow(7, 8, "A", "child", "A-PARENT", "COLOR/MATERIAL"),
            ]
            write_filled_workbook(
                source, output, profile,
                {(7, 1): "A-PARENT", (7, 2): "Parent", (8, 1): "A", (8, 2): "Child"},
                row_materializations=rows,
            )

            filled = openpyxl.load_workbook(output)
            sheet = filled["Template"]
            self.assertEqual(sheet.cell(7, 1).value, "A-PARENT")
            self.assertEqual(sheet.cell(8, 1).value, "A")
            self.assertIsNone(sheet.cell(7, 3).value)
            self.assertIsNone(sheet.cell(8, 5).value)
            validation_range = str(sheet.data_validations.dataValidation[0].sqref)
            self.assertIn("B7", validation_range)
            self.assertIn("B8", validation_range)
            self.assertIn("C7:C1048576", str(sheet.data_validations.dataValidation[1].sqref))
            self.assertIn("A10:B10", {str(item) for item in sheet.merged_cells.ranges})
            self.assertIn("B7:B8", {str(item.sqref) for item in sheet.conditional_formatting})
            self.assertEqual(sheet.auto_filter.ref, "A7:E10")
            self.assertEqual(sheet.cell(10, 6).value, "=A7")
            filled.close()

    def test_fill_endpoint_expands_listing_by_default_and_returns_variant_report(self):
        source = Path(r"F:\AI Projects\GIGAB2B\input\CHAIR-UK-1SKU.xlsm")
        if not source.is_file():
            self.skipTest("real Amazon template fixture is not available")
        products = {
            "W5807S00002": {"sku": "W5807S00002", "productName": "Red chair", "mainColor": "Red", "mainMaterial": "Acacia", "skuAvailable": True},
            "W5807S00003": {"sku": "W5807S00003", "productName": "Blue chair", "mainColor": "Blue", "mainMaterial": "Alloy Steel", "skuAvailable": False},
        }
        listing = {
            "seed_sku": "W5807S00002", "main": products["W5807S00002"],
            "requested_skus": list(products), "products": list(products.values()),
            "missing_skus": [], "over_limit": False, "warning": None,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            uploads = Path(temp_dir) / "templates"
            outputs = Path(temp_dir) / "outputs"
            uploads.mkdir()
            outputs.mkdir()
            original_config = {key: app_module.app.config.get(key) for key in (
                "TEMPLATE_FILLER_TEMPLATE_DIR", "TEMPLATE_FILLER_OUTPUT_DIR", "TEMPLATE_FILLER_POLICY_DB",
                "TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS",
            )}
            app_module.app.config.update(
                TEMPLATE_FILLER_TEMPLATE_DIR=str(uploads), TEMPLATE_FILLER_OUTPUT_DIR=str(outputs),
                TEMPLATE_FILLER_POLICY_DB=str(Path(temp_dir) / "policies.sqlite3"),
                TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS=lambda sku, market: listing,
            )
            client = app_module.app.test_client()
            try:
                analyzed = client.post("/api/template-filler/analyze", data={"file": (io.BytesIO(source.read_bytes()), source.name)}, content_type="multipart/form-data")
                self.assertEqual(analyzed.status_code, 200, analyzed.get_data(as_text=True))
                filled = client.post("/api/template-filler/fill", json={"template_id": analyzed.get_json()["template_id"]})
                self.assertEqual(filled.status_code, 200, filled.get_data(as_text=True))
                payload = filled.get_json()
            finally:
                for key, value in original_config.items():
                    if value is None:
                        app_module.app.config.pop(key, None)
                    else:
                        app_module.app.config[key] = value

        self.assertEqual(payload["variant_summary"]["groups_expanded"], 1)
        self.assertEqual(payload["variant_summary"]["parents_added"], 1)
        self.assertEqual(payload["variant_summary"]["children_added"], 2)
        self.assertEqual(payload["variant_groups"][0]["parent_sku"], "W5807S00002-PARENT")

    def test_fill_endpoint_reports_listing_fetch_failure_as_blocked_group(self):
        source = Path(r"F:\AI Projects\GIGAB2B\input\CHAIR-UK-1SKU.xlsm")
        if not source.is_file():
            self.skipTest("real Amazon template fixture is not available")
        with tempfile.TemporaryDirectory() as temp_dir:
            uploads = Path(temp_dir) / "templates"
            outputs = Path(temp_dir) / "outputs"
            uploads.mkdir()
            outputs.mkdir()
            original_config = {key: app_module.app.config.get(key) for key in (
                "TEMPLATE_FILLER_TEMPLATE_DIR", "TEMPLATE_FILLER_OUTPUT_DIR", "TEMPLATE_FILLER_POLICY_DB",
                "TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS",
            )}
            app_module.app.config.update(
                TEMPLATE_FILLER_TEMPLATE_DIR=str(uploads), TEMPLATE_FILLER_OUTPUT_DIR=str(outputs),
                TEMPLATE_FILLER_POLICY_DB=str(Path(temp_dir) / "policies.sqlite3"),
                TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS=lambda _sku, _market: (_ for _ in ()).throw(RuntimeError("network unavailable")),
            )
            client = app_module.app.test_client()
            try:
                analyzed = client.post("/api/template-filler/analyze", data={"file": (io.BytesIO(source.read_bytes()), source.name)}, content_type="multipart/form-data")
                self.assertEqual(analyzed.status_code, 200, analyzed.get_data(as_text=True))
                filled = client.post("/api/template-filler/fill", json={"template_id": analyzed.get_json()["template_id"]})
                self.assertEqual(filled.status_code, 200, filled.get_data(as_text=True))
                payload = filled.get_json()
            finally:
                for key, value in original_config.items():
                    if value is None:
                        app_module.app.config.pop(key, None)
                    else:
                        app_module.app.config[key] = value

        self.assertFalse(payload["summary"]["upload_ready"])
        self.assertTrue(any(item["status"] == "variant_fetch_incomplete" for item in payload["issues"]))


if __name__ == "__main__":
    unittest.main()
