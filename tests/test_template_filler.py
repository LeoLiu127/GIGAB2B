from __future__ import annotations

import os
import hashlib
import io
import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

import openpyxl

import app as app_module
from template_filler.mapping import build_fill_plan
from template_filler.parser import parse_amazon_template
from template_filler.writer import write_filled_workbook


FIXTURE_DIR = Path(os.environ.get("GIGAB2B_TEMPLATE_FIXTURE_DIR", "input"))


class AmazonTemplateParserTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not FIXTURE_DIR.exists():
            raise unittest.SkipTest("real Amazon template fixtures are not available")

    def test_cabinet_template_discovers_schema_required_fields_and_sku_row(self):
        profile = parse_amazon_template(FIXTURE_DIR / "CABINET-UK-1SKU.xlsm")

        self.assertEqual(profile.platform, "amazon")
        self.assertEqual(profile.market, "UK")
        self.assertEqual(profile.marketplace_id, "A1F83G8C2ARO7P")
        self.assertEqual(profile.language_tag, "en_GB")
        self.assertEqual(profile.category, "CABINET")
        self.assertEqual(profile.sheet_name, "Template")
        self.assertEqual(profile.attribute_row, 5)
        self.assertEqual(profile.data_row, 7)
        self.assertEqual(len(profile.fields), 354)
        self.assertEqual([(row.row, row.sku) for row in profile.sku_rows], [(7, "N890P39984041W")])

        required = [field for field in profile.fields if field.requirement == "required"]
        self.assertEqual(len(required), 8)
        self.assertIn("brand", {field.base_name for field in required})

        product_type = profile.field_by_base_name("product_type")
        self.assertTrue(product_type.is_dropdown)
        self.assertIn("CABINET", product_type.allowed_values)

    def test_chair_template_skips_prefill_notice_and_starts_at_real_sku_row(self):
        profile = parse_amazon_template(FIXTURE_DIR / "CHAIR-UK-1SKU.xlsm")

        self.assertEqual(profile.category, "CHAIR")
        self.assertEqual(profile.data_row, 8)
        self.assertEqual(len(profile.fields), 403)
        self.assertEqual([(row.row, row.sku) for row in profile.sku_rows], [(8, "W5807S00002")])

    def test_chair_template_distinguishes_required_conditional_and_dropdown_fields(self):
        profile = parse_amazon_template(FIXTURE_DIR / "CHAIR-UK-1SKU.xlsm")

        brand = profile.field_by_base_name("brand")
        product_id_value = profile.field_by_base_name("amzn1.volt.ca.product_id_value")
        material = profile.field_by_base_name("material")

        self.assertEqual(brand.requirement, "required")
        self.assertEqual(product_id_value.requirement, "conditionally_required")
        self.assertEqual(material.requirement, "conditionally_required")
        self.assertTrue(material.is_dropdown)
        self.assertIn("Velvet", material.allowed_values)
        self.assertNotIn("Chenille", material.allowed_values)


class AmazonTemplateFillTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not FIXTURE_DIR.exists():
            raise unittest.SkipTest("real Amazon template fixtures are not available")

    def test_chair_plan_maps_only_semantically_safe_giga_fields_and_reports_missing_required(self):
        source = FIXTURE_DIR / "CHAIR-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        product = {
            "sku": "W5807S00002",
            "productName": "Oversized Chaise Lounge Sofa",
            "mpn": "UK-991-1BK",
            "mainColor": "Black",
            "mainMaterial": "Chenille",
            "assembledLength": "184.50",
            "assembledWidth": "145.00",
            "assembledHeight": "76.50",
            "assembledWeight": "39.70",
            "assembledWeightUnit": "kg",
            "placeOfOrigin": "",
            "upc": "",
            "description": "",
            "characteristics": ["Pocket storage", "Deep seat", "Soft fabric", "No assembly"],
            "brandInfo": {"brandName": None},
        }

        plan = build_fill_plan(profile, source, {"W5807S00002": product})

        def planned(base_name: str):
            field = profile.field_by_base_name(base_name)
            return plan.changes.get((8, field.column))

        self.assertEqual(planned("product_type"), "CHAIR")
        self.assertEqual(planned("item_name"), "Oversized Chaise Lounge Sofa")
        self.assertEqual(planned("model_number"), "UK-991-1BK")
        self.assertEqual(planned("color"), "Black")
        self.assertIsNone(planned("material"), "Chenille is not an allowed Material dropdown value")

        statuses = {(issue.status, issue.field_id) for issue in plan.issues}
        self.assertNotIn(("missing_required", profile.field_by_base_name("brand").field_id), statuses)
        self.assertNotIn(("missing_required", profile.field_by_base_name("amzn1.volt.ca.product_id_type").field_id), statuses)
        self.assertNotIn(("missing_required", profile.field_by_base_name("country_of_origin").field_id), statuses)
        self.assertTrue(any(issue.status == "dropdown_required" and issue.field_id.startswith("material[") for issue in plan.issues))
        self.assertEqual(
            sum(issue.status == "dropdown_required" and issue.field_id.startswith("material[") for issue in plan.issues),
            1,
        )
        browse_nodes = profile.field_by_base_name("recommended_browse_nodes")
        self.assertIn(("manual_attention", browse_nodes.field_id), statuses)
        self.assertNotIn(("missing_required", browse_nodes.field_id), statuses)

    def test_cabinet_plan_applies_approved_uk_defaults_and_tracks_their_provenance(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        product = {
            "sku": "N890P39984041W",
            "productName": "Sideboard",
            "brandInfo": {"brandName": "A GIGA supplier brand that must not be used"},
            "upc": "012345678905",
            "skuAvailable": True,
        }

        plan = build_fill_plan(profile, source, {"N890P39984041W": product})

        def planned(base_name: str, suffix: str = ""):
            field = next(field for field in profile.fields if field.base_name == base_name and suffix in field.field_id)
            return plan.changes.get((7, field.column))

        self.assertEqual(planned("brand"), "GENERIC")
        self.assertEqual(planned("amzn1.volt.ca.product_id_type"), "GTIN Exempt")
        self.assertIsNone(planned("amzn1.volt.ca.product_id_value"))
        self.assertEqual(planned("condition_type"), "New")
        self.assertEqual(planned("country_of_origin"), "China")
        self.assertEqual(planned("batteries_required"), "No")
        self.assertEqual(planned("batteries_included"), "No")
        self.assertEqual(planned("fulfillment_availability", ".quantity"), 5)
        self.assertEqual(planned("supplier_declared_dg_hz_regulation"), "Not Applicable")

        defaults = {item.field_id: item for item in plan.filled_fields if item.source == "business_default"}
        self.assertEqual(defaults[profile.field_by_base_name("brand").field_id].value, "GENERIC")
        self.assertEqual(defaults[profile.field_by_base_name("amzn1.volt.ca.product_id_type").field_id].value, "GTIN Exempt")
        self.assertEqual(defaults[profile.field_by_base_name("supplier_declared_dg_hz_regulation").field_id].value, "Not Applicable")

    def test_cabinet_plan_uses_zero_quantity_when_giga_marks_sku_unavailable(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        plan = build_fill_plan(profile, source, {"N890P39984041W": {"sku": "N890P39984041W", "skuAvailable": False}})

        quantity = next(
            field for field in profile.fields
            if field.base_name == "fulfillment_availability" and field.field_id.endswith(".quantity")
        )
        self.assertEqual(plan.changes[(7, quantity.column)], 0)

    def test_cabinet_plan_keeps_manual_fields_empty_and_suppresses_unconfigured_conditional_noise(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        plan = build_fill_plan(profile, source, {"N890P39984041W": {"sku": "N890P39984041W"}})

        manual = {(issue.status, issue.field_id) for issue in plan.issues}
        self.assertTrue(any(status == "manual_attention" and field_id.startswith("recommended_browse_nodes") for status, field_id in manual))
        self.assertTrue(any(status == "manual_attention" and field_id.startswith("manufacturer") for status, field_id in manual))
        self.assertFalse(any(issue.status == "conditional_attention" for issue in plan.issues))

    def test_chair_plan_reports_operator_required_listing_fields_and_applies_default_channel(self):
        source = FIXTURE_DIR / "CHAIR-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        plan = build_fill_plan(profile, source, {"W5807S00002": {"sku": "W5807S00002"}})

        required_fields = {
            "number_of_items",
            "is_assembly_required",
            "size",
            "unit_count",
            "included_components",
            "is_fragile",
            "list_price",
            "merchant_shipping_group",
        }
        reported = {issue.field_id for issue in plan.issues if issue.status == "business_required"}
        for field in profile.fields:
            if field.base_name in required_fields and (
                field.base_name != "included_components" or field.field_id.endswith("#1.value")
            ):
                self.assertIn(field.field_id, reported)

        channel = next(
            field for field in profile.fields
            if field.base_name == "fulfillment_availability" and field.field_id.endswith(".fulfillment_channel_code")
        )
        self.assertEqual(plan.changes[(8, channel.column)], "DEFAULT")
        self.assertFalse(any(issue.status == "business_required" and issue.field_id == channel.field_id for issue in plan.issues))

        cabinet_source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        cabinet_profile = parse_amazon_template(cabinet_source)
        cabinet_plan = build_fill_plan(cabinet_profile, cabinet_source, {"N890P39984041W": {"sku": "N890P39984041W"}})
        self.assertFalse(any(issue.status == "business_required" for issue in cabinet_plan.issues))

    def test_fill_plan_preserves_existing_operator_values(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        workbook = openpyxl.load_workbook(source, keep_vba=True)
        title_field = profile.field_by_base_name("item_name")
        workbook[profile.sheet_name].cell(7, title_field.column, "Operator approved title")
        with tempfile.TemporaryDirectory() as temp_dir:
            prepared = Path(temp_dir) / "prepared.xlsm"
            workbook.save(prepared)
            workbook.close()
            prepared_profile = parse_amazon_template(prepared)
            plan = build_fill_plan(
                prepared_profile,
                prepared,
                {"N890P39984041W": {"sku": "N890P39984041W", "productName": "Raw GIGA title"}},
            )

        self.assertNotIn((7, title_field.column), plan.changes)
        self.assertTrue(any(issue.status == "preserved" and issue.field_id == title_field.field_id for issue in plan.issues))

    def test_writer_changes_only_target_sheet_package_part_and_keeps_template_controls(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        profile = parse_amazon_template(source)
        title = profile.field_by_base_name("item_name")
        changes = {(7, title.column): "Filled cabinet title"}

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "filled.xlsm"
            write_filled_workbook(source, output, profile, changes)

            original_workbook = openpyxl.load_workbook(source, keep_vba=True, data_only=False)
            original_validation_count = len(original_workbook["Template"].data_validations.dataValidation)
            original_validation_formulas = [
                item.formula1 for item in original_workbook["Template"].data_validations.dataValidation
            ]
            original_workbook.close()

            workbook = openpyxl.load_workbook(output, keep_vba=True, data_only=False)
            self.assertEqual(workbook["Template"].cell(7, title.column).value, "Filled cabinet title")
            self.assertEqual(workbook["Template"].cell(5, title.column).value, title.field_id)
            self.assertEqual(len(workbook["Template"].data_validations.dataValidation), original_validation_count)
            self.assertEqual(
                [item.formula1 for item in workbook["Template"].data_validations.dataValidation],
                original_validation_formulas,
            )
            self.assertEqual(workbook["Dropdown Lists"].sheet_state, "hidden")
            workbook.close()

            with ZipFile(source) as before, ZipFile(output) as after:
                self.assertEqual(set(before.namelist()), set(after.namelist()))
                changed_parts = []
                for name in before.namelist():
                    before_hash = hashlib.sha256(before.read(name)).digest()
                    after_hash = hashlib.sha256(after.read(name)).digest()
                    if before_hash != after_hash:
                        changed_parts.append(name)
                self.assertEqual(changed_parts, ["xl/worksheets/sheet5.xml"])


class TemplateFillerApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not FIXTURE_DIR.exists():
            raise unittest.SkipTest("real Amazon template fixtures are not available")

    def test_analyze_then_fill_returns_workbook_and_missing_field_report(self):
        source = FIXTURE_DIR / "CABINET-UK-1SKU.xlsm"
        product = {
            "sku": "N890P39984041W",
            "productName": "Sideboard 140cm with storage",
            "mpn": "N890P39984041W",
            "mainColor": "Brown+White",
            "mainMaterial": "MDF",
            "assembledLength": "140.00",
            "assembledWidth": "35.00",
            "assembledHeight": "95.00",
            "assembledWeight": "51.00",
            "assembledWeightUnit": "kg",
            "characteristics": ["Ample storage", "Farmhouse design", "Easy assembly"],
            "description": "",
            "brandInfo": {"brandName": None},
            "placeOfOrigin": "",
            "upc": "N890P39984041W",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            uploads = Path(temp_dir) / "templates"
            outputs = Path(temp_dir) / "excel"
            uploads.mkdir()
            outputs.mkdir()
            original_excel_dir = app_module.EXCEL_OUTPUT_DIR
            original_config = {
                key: app_module.app.config.get(key)
                for key in (
                    "TEMPLATE_FILLER_TEMPLATE_DIR",
                    "TEMPLATE_FILLER_OUTPUT_DIR",
                    "TEMPLATE_FILLER_FETCH_PRODUCTS",
                )
            }
            app_module.EXCEL_OUTPUT_DIR = str(outputs)
            app_module.app.config.update(
                TEMPLATE_FILLER_TEMPLATE_DIR=str(uploads),
                TEMPLATE_FILLER_OUTPUT_DIR=str(outputs),
                TEMPLATE_FILLER_FETCH_PRODUCTS=lambda skus, market: [product],
            )
            client = app_module.app.test_client()
            try:
                with source.open("rb") as stream:
                    analyzed = client.post(
                        "/api/template-filler/analyze",
                        data={"file": (io.BytesIO(stream.read()), source.name)},
                        content_type="multipart/form-data",
                    )
                self.assertEqual(analyzed.status_code, 200, analyzed.get_data(as_text=True))
                analysis = analyzed.get_json()
                self.assertEqual(analysis["template"]["category"], "CABINET")
                self.assertEqual(analysis["sku_rows"], [{"row": 7, "sku": "N890P39984041W"}])
                self.assertGreater(analysis["summary"]["dropdown_fields"], 0)

                filled = client.post(
                    "/api/template-filler/fill",
                    json={"template_id": analysis["template_id"]},
                )
                self.assertEqual(filled.status_code, 200, filled.get_data(as_text=True))
                payload = filled.get_json()
                self.assertGreater(payload["summary"]["fields_filled"], 0)
                self.assertEqual(payload["summary"]["missing_required"], 0)
                self.assertGreater(payload["summary"]["dropdown_required"], 0)
                self.assertEqual(payload["summary"]["business_required"], 0)
                self.assertFalse(payload["summary"]["upload_ready"])
                self.assertIn("filled_fields", payload)
                self.assertTrue(payload["filled_fields"])
                self.assertTrue((outputs / payload["output_file"]).is_file())
                self.assertTrue((outputs / payload["report_file"]).is_file())

                downloaded = client.get(f"/api/downloads/{payload['output_file']}")
                self.assertEqual(downloaded.status_code, 200)
                report = client.get(f"/api/template-filler/reports/{payload['report_file']}")
                self.assertEqual(report.status_code, 200)
                downloaded.close()
                report.close()
            finally:
                app_module.EXCEL_OUTPUT_DIR = original_excel_dir
                for key, value in original_config.items():
                    if value is None:
                        app_module.app.config.pop(key, None)
                    else:
                        app_module.app.config[key] = value


if __name__ == "__main__":
    unittest.main()
