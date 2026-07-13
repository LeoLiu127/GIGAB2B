from __future__ import annotations

import base64
import os
import re
from pathlib import Path
from urllib.parse import unquote
from zipfile import ZipFile

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries

from .models import SkuRow, TemplateField, TemplateProfile


MARKETPLACE_TO_MARKET = {
    "A1F83G8C2ARO7P": "UK",
    "A1PA6795UKMFR9": "DE_TAX",
    "ATVPDKIKX0DER": "US",
    "A13V1IB3VIYZZH": "FR",
}


def _setting(text: str, name: str) -> str:
    match = re.search(rf"(?:^|&){re.escape(name)}=([^&]*)", text)
    return unquote(match.group(1)) if match else ""


def _decode_category(settings: str) -> str:
    encoded = _setting(settings, "ptds")
    if not encoded:
        return ""
    try:
        return base64.b64decode(encoded).decode("utf-8").strip().upper()
    except (ValueError, UnicodeDecodeError):
        return ""


def _base_name(field_id: str) -> str:
    clean = field_id.lstrip(":")
    if clean.startswith("amzn1.volt.ca.product_id_"):
        return clean
    return re.split(r"\[|#", clean, maxsplit=1)[0]


def _requirement(value: object) -> str:
    normalized = re.sub(r"\s+", " ", str(value or "").strip().lower())
    return {
        "required": "required",
        "conditionally required": "conditionally_required",
        "recommended": "recommended",
        "optional": "optional",
    }.get(normalized, "unknown")


def _find_template_sheet(workbook):
    for name in ("Template", "Vorlage"):
        if name in workbook.sheetnames:
            return workbook[name]
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, 10) + 1):
            if any(cell.value == "contribution_sku#1.value" for cell in sheet[row]):
                return sheet
    raise ValueError("无法找到 Amazon Template/Vorlage 数据工作表")


def _find_definitions(workbook):
    for sheet in workbook.worksheets:
        if sheet.max_row < 2 or sheet.max_column < 6:
            continue
        header = str(sheet.cell(2, 2).value or "").strip().lower()
        if header in {"field name", "feldname"}:
            return sheet
    raise ValueError("模板缺少可识别的 Data Definitions/Datendefinitionen 工作表")


def _definitions(sheet) -> dict[str, tuple[str, str]]:
    result: dict[str, tuple[str, str]] = {}
    for row in range(3, sheet.max_row + 1):
        field_id = str(sheet.cell(row, 2).value or "").strip()
        if not field_id:
            continue
        result[field_id] = (
            str(sheet.cell(row, 3).value or "").strip(),
            _requirement(sheet.cell(row, 6).value),
        )
    return result


def _valid_sku(value: object) -> str:
    text = str(value or "").strip()
    if not text or text == "ABC123" or len(text) > 100:
        return ""
    lowered = text.lower()
    if "prefilled attributes" in lowered or lowered.startswith("✅"):
        return ""
    if any(char in text for char in "\r\n"):
        return ""
    return text


def _sku_rows(sheet, attribute_row: int, sku_column: int, product_type_column: int) -> tuple[SkuRow, ...]:
    rows: list[SkuRow] = []
    for row in range(attribute_row + 2, sheet.max_row + 1):
        sku = _valid_sku(sheet.cell(row, sku_column).value)
        if sku:
            rows.append(
                SkuRow(
                    row=row,
                    sku=sku,
                    product_type=str(sheet.cell(row, product_type_column).value or "").strip(),
                )
            )
    return tuple(rows)


def _validation_for_column(sheet, column: int, data_row: int):
    if not sheet.data_validations:
        return None
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        for cell_range in validation.ranges.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
            if min_col <= column <= max_col and min_row <= data_row <= max_row:
                return validation
    return None


def _defined_name_values(workbook, name: str) -> tuple[str, ...]:
    target = workbook.defined_names.get(name)
    if target is None:
        lower_name = name.lower()
        target = next((item for item in workbook.defined_names.values() if item.name.lower() == lower_name), None)
    if target is None:
        return ()
    values: list[str] = []
    try:
        destinations = list(target.destinations)
    except (TypeError, AttributeError):
        return ()
    for sheet_name, coordinates in destinations:
        if sheet_name not in workbook.sheetnames:
            continue
        selected = workbook[sheet_name][coordinates]
        rows = ((selected,),) if isinstance(selected, Cell) else selected
        for row in rows:
            for cell in row:
                value = str(cell.value or "").strip()
                if value and value not in values:
                    values.append(value)
    return tuple(values)


def _defined_name_key(field_id: str) -> str:
    return re.sub(r"[\[\]#=:]", "", field_id)


def _allowed_values(workbook, formula: str | None, field_id: str, category: str) -> tuple[str, ...]:
    if not formula:
        return ()
    clean_formula = formula.lstrip("=")
    if not clean_formula.upper().startswith("INDIRECT"):
        return _defined_name_values(workbook, clean_formula)

    wanted = f"{category}{_defined_name_key(field_id)}".lower()
    matched = next((item.name for item in workbook.defined_names.values() if item.name.lower() == wanted), "")
    return _defined_name_values(workbook, matched) if matched else ()


def parse_amazon_template(path: str | os.PathLike[str]) -> TemplateProfile:
    source_path = Path(path).resolve()
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Amazon 模板仅支持 .xlsx / .xlsm")
    if not source_path.is_file():
        raise FileNotFoundError(source_path)

    workbook = openpyxl.load_workbook(
        source_path,
        read_only=False,
        data_only=False,
        keep_vba=source_path.suffix.lower() == ".xlsm",
        keep_links=True,
    )
    try:
        sheet = _find_template_sheet(workbook)
        settings = str(sheet.cell(1, 1).value or "")
        attribute_row = int(_setting(settings, "attributeRow") or 5)
        label_row = int(_setting(settings, "labelRow") or attribute_row - 1)

        raw_fields = [
            (cell.column, str(cell.value).strip())
            for cell in sheet[attribute_row]
            if cell.value not in (None, "")
        ]
        if not raw_fields:
            raise ValueError(f"{sheet.title} 第 {attribute_row} 行没有 Amazon 字段 ID")

        by_id = {field_id: column for column, field_id in raw_fields}
        try:
            sku_column = by_id["contribution_sku#1.value"]
            product_type_column = by_id["product_type#1.value"]
        except KeyError as exc:
            raise ValueError(f"模板缺少关键字段: {exc.args[0]}") from exc

        rows = _sku_rows(sheet, attribute_row, sku_column, product_type_column)
        configured_data_row = int(_setting(settings, "dataRow") or 0)
        data_row = rows[0].row if rows else configured_data_row or attribute_row + 2

        marketplace_id = _setting(settings, "primaryMarketplaceId").rsplit(".", 1)[-1]
        if not marketplace_id:
            match = re.search(r"marketplace_id=([^\]]+)", " ".join(by_id))
            marketplace_id = match.group(1) if match else ""
        language_tag = _setting(settings, "contentLanguageTag")
        if not language_tag:
            match = re.search(r"language_tag=([^\]]+)", " ".join(by_id))
            language_tag = match.group(1) if match else ""
        category = _decode_category(settings)
        if not category:
            if "AttributePTDMAP" in workbook.sheetnames:
                category = str(workbook["AttributePTDMAP"].cell(1, 2).value or "").strip().upper()
        if not category:
            names = [item.name for item in workbook.defined_names.values()]
            match = next((re.match(r"^([A-Z][A-Z0-9_]*?)parentage_level", name) for name in names if "parentage_level" in name), None)
            category = match.group(1) if match else ""

        definitions = _definitions(_find_definitions(workbook))
        fields: list[TemplateField] = []
        for column, field_id in raw_fields:
            definition_label, requirement = definitions.get(field_id, ("", "unknown"))
            validation = _validation_for_column(sheet, column, data_row)
            formula = validation.formula1 if validation is not None else None
            fields.append(
                TemplateField(
                    field_id=field_id,
                    base_name=_base_name(field_id),
                    column=column,
                    column_letter=get_column_letter(column),
                    label=definition_label or str(sheet.cell(label_row, column).value or "").strip(),
                    requirement=requirement,
                    is_dropdown=validation is not None,
                    validation_formula=formula,
                    allowed_values=_allowed_values(workbook, formula, field_id, category),
                )
            )

        has_vba = False
        if source_path.suffix.lower() == ".xlsm":
            with ZipFile(source_path) as archive:
                has_vba = "xl/vbaProject.bin" in archive.namelist()

        return TemplateProfile(
            source_path=source_path,
            platform="amazon",
            market=MARKETPLACE_TO_MARKET.get(marketplace_id, "UNKNOWN"),
            marketplace_id=marketplace_id,
            language_tag=language_tag,
            category=category,
            sheet_name=sheet.title,
            label_row=label_row,
            attribute_row=attribute_row,
            data_row=data_row,
            fields=tuple(fields),
            sku_rows=rows,
            has_vba=has_vba,
        )
    finally:
        workbook.close()
