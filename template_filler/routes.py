from __future__ import annotations

import json
import os
import re
import uuid
from collections import Counter
from pathlib import Path

import openpyxl

from flask import Blueprint, current_app, jsonify, request, send_from_directory
from werkzeug.utils import secure_filename

from .mapping import build_fill_plan
from .parser import parse_amazon_template
from .policy import PolicyStore
from .variants import ListingProducts, VariantExpansion, expand_variant_rows
from .writer import write_filled_workbook


template_filler_bp = Blueprint("template_filler", __name__, url_prefix="/api/template-filler")


def _directories() -> tuple[Path, Path]:
    templates = Path(current_app.config["TEMPLATE_FILLER_TEMPLATE_DIR"]).resolve()
    outputs = Path(current_app.config["TEMPLATE_FILLER_OUTPUT_DIR"]).resolve()
    templates.mkdir(parents=True, exist_ok=True)
    outputs.mkdir(parents=True, exist_ok=True)
    return templates, outputs


def _policy_store() -> PolicyStore:
    return PolicyStore(current_app.config["TEMPLATE_FILLER_POLICY_DB"])


def _policy_state(profile) -> tuple[object | None, str, list[dict]]:
    store = _policy_store()
    store.ensure_initial_policy(profile)
    policy, drift = store.lookup(profile)
    if policy is None:
        return None, "unconfigured", []
    return policy, "drift_detected" if drift else "active", drift


def _policy_payload(profile) -> dict:
    policy, status, drift = _policy_state(profile)
    return {
        "policy_status": status,
        "policy_drift": drift,
        "policy": policy.to_dict() if policy else None,
        "policy_required": sum(rule.action == "required" for rule in policy.rules.values()) if policy else 0,
    }


def _field_payload(field) -> dict:
    return {
        "field_id": field.field_id,
        "label": field.label,
        "requirement": field.requirement,
        "column": field.column_letter,
        "is_dropdown": field.is_dropdown,
        "allowed_values": list(field.allowed_values[:100]),
    }


def _profile_payload(profile) -> dict:
    return {
        "platform": profile.platform,
        "market": profile.market,
        "marketplace_id": profile.marketplace_id,
        "language_tag": profile.language_tag,
        "category": profile.category,
        "sheet_name": profile.sheet_name,
        "attribute_row": profile.attribute_row,
        "data_row": profile.data_row,
        "field_count": len(profile.fields),
        "has_vba": profile.has_vba,
    }


def _manual_variation_themes(profile, source: Path) -> dict[int, str]:
    field = next((item for item in profile.fields if item.base_name == "variation_theme"), None)
    if field is None:
        return {}
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=False, keep_vba=True)
    try:
        sheet = workbook[profile.sheet_name]
        return {
            row.row: str(sheet.cell(row.row, field.column).value or "").strip()
            for row in profile.sku_rows
            if sheet.cell(row.row, field.column).value not in (None, "")
        }
    finally:
        workbook.close()


def _resolve_template(template_id: str, templates: Path) -> tuple[Path, dict]:
    if not re.fullmatch(r"[0-9a-f]{32}", str(template_id or "")):
        raise ValueError("template_id 无效")
    matches = [path for suffix in (".xlsx", ".xlsm") if (path := templates / f"{template_id}{suffix}").is_file()]
    if len(matches) != 1:
        raise FileNotFoundError("模板不存在或已过期")
    metadata_path = templates / f"{template_id}.json"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8")) if metadata_path.is_file() else {}
    return matches[0], metadata


@template_filler_bp.post("/analyze")
def analyze_template():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "请选择 Amazon XLSX/XLSM 模板"}), 400
    original_filename = os.path.basename(upload.filename)
    extension = Path(original_filename).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        return jsonify({"error": "模板仅支持 .xlsx / .xlsm"}), 400

    templates, _ = _directories()
    template_id = uuid.uuid4().hex
    destination = templates / f"{template_id}{extension}"
    temporary = templates / f"{template_id}.uploading{extension}"
    try:
        upload.save(temporary)
        profile = parse_amazon_template(temporary)
        if not profile.sku_rows:
            raise ValueError("SKU 列没有可处理的 GIGA Item code")
        if len(profile.sku_rows) > 200:
            raise ValueError("MVP 每个模板最多处理 200 行 SKU")
        if profile.market == "UNKNOWN":
            raise ValueError(f"暂不支持 marketplace: {profile.marketplace_id or 'unknown'}")
        os.replace(temporary, destination)
        (templates / f"{template_id}.json").write_text(
            json.dumps({"original_filename": original_filename}, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        destination.unlink(missing_ok=True)
        return jsonify({"error": f"模板解析失败: {exc}"}), 400

    requirement_counts = Counter(field.requirement for field in profile.fields)
    policy_details = _policy_payload(profile)
    template_fields = [_field_payload(field) for field in profile.fields]
    return jsonify(
        {
            "success": True,
            "template_id": template_id,
            "original_filename": original_filename,
            "template": _profile_payload(profile),
            "sku_rows": [{"row": row.row, "sku": row.sku} for row in profile.sku_rows],
            "summary": {
                "sku_count": len(profile.sku_rows),
                "required_fields": requirement_counts["required"],
                "conditional_fields": requirement_counts["conditionally_required"],
                "dropdown_fields": sum(field.is_dropdown for field in profile.fields),
                "policy_required": policy_details["policy_required"],
            },
            "fields": template_fields,
            "policy_status": policy_details["policy_status"],
            "policy_drift": policy_details["policy_drift"],
            "policy": policy_details["policy"],
        }
    )


@template_filler_bp.get("/policies/<template_id>")
def get_template_policy(template_id: str):
    templates, _ = _directories()
    try:
        source, _ = _resolve_template(template_id, templates)
        profile = parse_amazon_template(source)
        return jsonify({"success": True, **_policy_payload(profile), "fields": [_field_payload(field) for field in profile.fields]})
    except Exception as exc:
        return jsonify({"error": f"策略读取失败: {exc}"}), 400


@template_filler_bp.post("/policies/<template_id>")
def save_template_policy(template_id: str):
    templates, _ = _directories()
    body = request.get_json(silent=True) or {}
    raw_rules = body.get("rules")
    if not isinstance(raw_rules, list):
        return jsonify({"error": "rules 必须是规则数组"}), 400
    try:
        source, _ = _resolve_template(template_id, templates)
        profile = parse_amazon_template(source)
        rules = {}
        for item in raw_rules:
            if not isinstance(item, dict) or not str(item.get("field_id") or "").strip():
                raise ValueError("每条规则必须包含 field_id")
            field_id = str(item["field_id"]).strip()
            if field_id in rules:
                raise ValueError(f"策略字段重复: {field_id}")
            rules[field_id] = {"action": item.get("action"), "value": item.get("value"), "scope": item.get("scope")}
        policy = _policy_store().save(profile, rules)
        return jsonify({"success": True, "policy_status": "active", "policy_drift": [], "policy": policy.to_dict()})
    except Exception as exc:
        return jsonify({"error": f"策略保存失败: {exc}"}), 400


@template_filler_bp.post("/fill")
def fill_template():
    body = request.get_json(silent=True) or {}
    templates, outputs = _directories()
    try:
        source, metadata = _resolve_template(str(body.get("template_id") or ""), templates)
        profile = parse_amazon_template(source)
        policy, policy_status, policy_drift = _policy_state(profile)
        expand_variants = body.get("expand_variants") is not False
        expansion: VariantExpansion | None = None
        if expand_variants:
            fetch_listing = current_app.config["TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS"]

            def listing_products(sku: str, market: str) -> ListingProducts:
                try:
                    raw = fetch_listing(sku, market)
                except Exception as exc:
                    return ListingProducts(
                        seed_sku=sku,
                        main_sku=sku,
                        requested_skus=(sku,),
                        products={},
                        missing_skus=(sku,),
                        warning=f"GIGA 变体详情请求失败: {exc}",
                    )
                if isinstance(raw, ListingProducts):
                    return raw
                main = raw.get("main") or {}
                products = {
                    str(item.get("sku") or "").strip(): item
                    for item in raw.get("products") or []
                    if isinstance(item, dict) and item.get("sku")
                }
                requested = tuple(str(item).strip() for item in raw.get("requested_skus") or [] if str(item).strip())
                return ListingProducts(
                    seed_sku=str(raw.get("seed_sku") or sku),
                    main_sku=str(main.get("sku") or sku),
                    requested_skus=requested or (sku,),
                    products=products,
                    missing_skus=tuple(str(item).strip() for item in raw.get("missing_skus") or [] if str(item).strip()),
                    over_limit=bool(raw.get("over_limit")),
                    warning=str(raw.get("warning") or "") or None,
                )

            expansion = expand_variant_rows(profile, listing_products, manual_themes=_manual_variation_themes(profile, source))
            products_by_sku = {row.sku: row.product for row in expansion.rows if row.product is not None}
            plan = build_fill_plan(
                profile, source, products_by_sku, policy=policy, policy_configured=policy is not None, rows=expansion.rows,
            )
            plan.issues.extend(expansion.issues)
        else:
            fetch_products = current_app.config["TEMPLATE_FILLER_FETCH_PRODUCTS"]
            products = fetch_products([row.sku for row in profile.sku_rows], profile.market)
            products_by_sku = {
                str(item.get("sku") or "").strip(): item
                for item in products
                if isinstance(item, dict) and item.get("sku")
            }
            plan = build_fill_plan(profile, source, products_by_sku, policy=policy, policy_configured=policy is not None)

        original = secure_filename(str(metadata.get("original_filename") or source.name)) or "amazon-template.xlsm"
        stem = Path(original).stem[:80] or "amazon-template"
        token = uuid.uuid4().hex[:10]
        output_name = f"{stem}-filled-{token}{source.suffix.lower()}"
        report_name = f"{stem}-report-{token}.json"
        output_path = outputs / output_name
        report_path = outputs / report_name
        write_filled_workbook(source, output_path, profile, plan.changes, row_materializations=expansion.rows if expansion else None)

        counts = Counter(issue.status for issue in plan.issues)
        report = {
            "template": _profile_payload(profile),
            "source_filename": metadata.get("original_filename") or source.name,
            "output_file": output_name,
            "summary": {
                "rows_processed": plan.rows_processed,
                "fields_filled": plan.fields_filled,
                "missing_required": counts["missing_required"],
                "conditional_attention": counts["conditional_attention"],
                "manual_attention": counts["manual_attention"],
                "business_required": counts["business_required"],
                "policy_required": counts["business_required"],
                "dropdown_required": counts["dropdown_required"],
                "api_not_found": counts["api_not_found"],
                "invalid_existing_value": counts["invalid_existing_value"],
                "variant_groups_blocked": counts["variant_theme_unresolved"] + counts["variant_group_too_large"] + counts["variant_fetch_incomplete"] + counts["variant_manual_theme_conflict"],
                "upload_ready": not any(
                    issue.status in {"missing_required", "business_required", "policy_unconfigured", "dropdown_required", "api_not_found", "invalid_existing_value", "variant_theme_unresolved", "variant_group_too_large", "variant_fetch_incomplete", "variant_manual_theme_conflict"}
                    for issue in plan.issues
                ),
            },
            "filled_fields": [item.to_dict() for item in plan.filled_fields],
            "issues": [issue.to_dict() for issue in plan.issues],
            "policy_status": policy_status,
            "policy_drift": policy_drift,
            "policy": policy.to_dict() if policy else None,
            "variant_summary": expansion.summary if expansion else {"seed_rows": len(profile.sku_rows), "groups_expanded": 0, "groups_blocked": 0, "parents_added": 0, "children_added": 0},
            "variant_groups": [group.to_dict() for group in expansion.groups] if expansion else [],
        }
        temporary_report = report_path.with_suffix(".json.tmp")
        temporary_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temporary_report, report_path)
        return jsonify(
            {
                "success": True,
                "output_file": output_name,
                "report_file": report_name,
                "summary": report["summary"],
                "filled_fields": report["filled_fields"],
                "issues": report["issues"],
                "policy_status": report["policy_status"],
                "policy_drift": report["policy_drift"],
                "policy": report["policy"],
                "variant_summary": report["variant_summary"],
                "variant_groups": report["variant_groups"],
            }
        )
    except Exception as exc:
        return jsonify({"error": f"模板填表失败: {exc}"}), 400


@template_filler_bp.get("/reports/<filename>")
def download_report(filename: str):
    if os.path.basename(filename) != filename or not filename.lower().endswith(".json"):
        return jsonify({"error": "报告文件名非法"}), 400
    _, outputs = _directories()
    return send_from_directory(outputs, filename, as_attachment=True)
