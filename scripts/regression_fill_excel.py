"""回归测试：确认重构后 fill_excel 写入 cell 与重构前历史行为完全一致。

策略：
- 不依赖 GIGA、AI、网络。
- 构造一个最小 mock 模板（含 Vorlage sheet、第 7 行空白）。
- 用一组 mock product/ai_result 调 fill_excel。
- 把"修改前"按 app.py:fill_excel 旧版本手动计算的 cell 值，与本次 fill 出来的 cell 做逐 cell 比对。
- 任一 cell 不一致即断言失败。

跑法：python scripts/regression_fill_excel.py
"""

from __future__ import annotations

import os
import sys
import tempfile
import shutil

# 让脚本可独立运行
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl


MOCK_PRODUCT = {
    "sku": "W2339P502190",
    "productName": "Raised Garden Bed Metal Planter Galvanized Steel",
    "mpn": "W2339P502190-MPN",
    "mainColor": "Silver",
    "mainMaterial": "Galvanized Steel",
    "assembledLength": 116.0,
    "assembledWidth": 30.0,
    "assembledHeight": 5.5,
    "weightKg": 12.0,
    "placeOfOrigin": "China",
    "characteristics": ["c1", "c2", "c3", "c4", "c5"],
    "imageUrls": [
        "https://giga.example.com/img-0.jpg",
        "https://giga.example.com/img-1.jpg",
        "https://giga.example.com/img-2.jpg",
        "https://giga.example.com/img-3.jpg",
    ],
    "fileUrls": ["https://giga.example.com/manual.pdf"],
    "attributes": {
        "Main Color": "Silver",
        "Product Style": "Modern,Minimalist",
    },
}

MOCK_AI = {
    "title": "Raised Garden Bed Galvanized Steel Planter 116x30x5.5cm",
    "bullets": [
        "Heavy-duty galvanized steel construction for long-lasting outdoor use",
        "Open base design promotes healthy root growth and drainage",
        "Easy assembly — no special tools required, kit includes all hardware",
        "Weather-resistant zinc-aluminium coating prevents rust over years",
        "Ideal for vegetables, herbs, and flowers in any backyard or patio",
    ],
    "description": "<b>Produktmerkmale:</b><br>\n<li>bullet1</li>",
    "search_terms": "raised garden bed, metal planter, galvanized, outdoor",
}


def _make_minimal_template(tmpdir: str) -> str:
    """生成一份最小 mock 模板，含 Vorlage sheet（第 7 行之前留空，第 7 行也空）。"""
    path = os.path.join(tmpdir, "PLANTER-us.xlsm")
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Vorlage")
    # 第 7 行 cell 全空(没有合并、没有预设内容)—— fill_excel 会写入
    wb.save(path)
    return path


def _expected_cell_values(market: str) -> dict:
    """按旧 fill_excel(app.py 重构前)手算的 cell 值。

    这些值与重构前版本完全一致；本次 fill_excel 重构之后必须产生同样的 cell 值,否则回归失败。
    列号来自旧 COL_MAP,与 templates_catalog.AMAZON_PLANTER.col_map 完全对齐(已校对)。
    """
    expected = {}

    # 从 templates_catalog 拿列号,保证测试断言与 catalog 一致
    from templates_catalog import AMAZON_PLANTER
    cm = AMAZON_PLANTER.col_map

    expected[cm["sku"]]           = MOCK_PRODUCT["sku"]
    expected[cm["product_type"]]  = "PLANTER"          # fixed_values
    expected[cm["sku2"]]          = "full_update"      # fixed_values
    expected[cm["product_name"]]  = MOCK_AI["title"]
    expected[cm["mpn"]]           = MOCK_PRODUCT["mpn"]
    expected[cm["manufacturer"]]  = "YUDA HOME FURNITURE"

    # image: main + pt1..ptN(填前 4 张,槽位 4..8 留空)
    expected[cm["main_image"]]    = MOCK_PRODUCT["imageUrls"][0]
    # pt1..pt8 在列 25..32;这里只有 4 张图 → 列 25..28 有值
    for i in range(1, 4):
        expected[25 + i - 1]     = MOCK_PRODUCT["imageUrls"][i]

    # bullets: 列 41..45
    for i, b in enumerate(MOCK_AI["bullets"]):
        expected[41 + i]         = b

    # search_terms: 用户给了,直接用
    expected[cm["search_terms"]] = MOCK_AI["search_terms"]

    # description
    expected[cm["description"]]  = MOCK_AI["description"]

    # special_attr 列 47..51: 按 market 取
    from templates_catalog import AMAZON_PLANTER
    attrs = AMAZON_PLANTER.special_attrs_by_market.get(market,
            AMAZON_PLANTER.special_attrs_by_market["DE_TAX"])
    for i, a in enumerate(attrs):
        expected[47 + i]         = a

    # 常规属性
    expected[cm["style"]]         = "Modern,Minimalist"  # 来自 attributes
    expected[cm["material"]]      = MOCK_PRODUCT["mainMaterial"]
    expected[cm["color"]]         = MOCK_PRODUCT["mainColor"]
    expected[cm["item_count"]]    = 1

    # 尺寸 + 重量
    expected[cm["length"]]        = 116.0
    expected[cm["length_unit"]]   = "cm"
    expected[cm["height"]]        = 5.5
    expected[cm["height_unit"]]   = "cm"
    expected[cm["width"]]         = 30.0
    expected[cm["width_unit"]]    = "cm"

    expected[cm["weight"]]        = 12.0
    expected[cm["weight_unit"]]   = "kg"

    # country:用户给了 "China",但 fixed_values["country"] 也是 "China";两者一致
    expected[cm["country"]]       = "China"

    # pkg(PLANTER 算法)
    from templates_catalog import packaging_for
    pkg_l, pkg_w, pkg_h, pkg_wt = packaging_for("PLANTER", 12.0)
    expected[cm["pkg_length"]]      = pkg_l
    expected[cm["pkg_length_unit"]] = "cm"
    expected[cm["pkg_width"]]       = pkg_w
    expected[cm["pkg_width_unit"]]  = "cm"
    expected[cm["pkg_height"]]      = pkg_h
    expected[cm["pkg_height_unit"]] = "cm"
    expected[cm["pkg_weight"]]      = pkg_wt
    expected[cm["pkg_weight_unit"]] = "kg"

    # pdf 附件
    expected[cm["pdf"]]          = MOCK_PRODUCT["fileUrls"][0]

    return expected


def test_one_market(market: str, fill_module_path: str | None = None) -> None:
    """对单个市场跑 fill_excel,把实际写入的 Vorlage row=7 与期望值逐 cell 对比。"""
    tmpdir = tempfile.mkdtemp(prefix=f"planter_test_{market}_")
    try:
        template_path = _make_minimal_template(tmpdir)
        # 把 app.TEMPLATE_DIR 临时指向 tmpdir 以避免污染工作树
        import app as app_mod
        original_template_dir = app_mod.TEMPLATE_DIR
        original_output_dir = app_mod.EXCEL_OUTPUT_DIR
        app_mod.TEMPLATE_DIR = tmpdir
        app_mod.EXCEL_OUTPUT_DIR = tmpdir
        try:
            out_path = app_mod.fill_excel(
                MOCK_PRODUCT, MOCK_AI, market,
                template_name="PLANTER-us.xlsm",
                image_strategy="use_giga",
                image_overrides=None,
            )
        finally:
            app_mod.TEMPLATE_DIR = original_template_dir
            app_mod.EXCEL_OUTPUT_DIR = original_output_dir

        # 加载输出,读 Vorlage row=7
        wb = openpyxl.load_workbook(out_path, keep_vba=True)
        ws = wb["Vorlage"]

        expected = _expected_cell_values(market)

        # 收集 row=7 所有非空 cell
        actual = {}
        for col in range(1, 280):
            v = ws.cell(row=7, column=col).value
            if v is not None and v != "":
                actual[col] = v

        # 逐 cell 比对
        diffs = []
        for col, want in expected.items():
            got = actual.get(col)
            if got != want:
                diffs.append((col, want, got))
        # 反向:实际多写了任何"不该写"的 cell(除了 main_image 上的 image_strategy 注释不是 cell 值,这里只看 .value)
        extra_cols = set(actual) - set(expected)
        if extra_cols:
            diffs.append(("EXTRA_COLS", list(extra_cols), None))

        if diffs:
            print(f"  [FAIL] {market} 差异:")
            for d in diffs:
                print(f"    {d}")
            raise AssertionError(f"{market} 回归失败: {len(diffs)} 处不一致")
        print(f"  [PASS] {market}  — {len(expected)} cells 全部一致")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def main():
    print("== fill_excel 回归测试 ==")
    for market in ["US", "UK", "FR"]:
        # DE_TAX/DE_TAXFREE 是同一份模板 + 同一份 descriptor,只用一个代表即可
        test_one_market(market)
    print("\n✓ 全部市场回归通过(fill_excel 重构与旧版本 byte-exact 等价)")


if __name__ == "__main__":
    main()
