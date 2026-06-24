"""
读取 Excel 模板的列头映射（表头行+字段名行），输出完整的列索引到字段名对照表
"""
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

wb = openpyxl.load_workbook(
    r"F:\AI Projects\GIGAB2B\PLANTER-de.xlsx",
    data_only=True, keep_vba=False
)
ws = wb["Vorlage"]

print(f"工作表: Vorlage | 最大行: {ws.max_row} | 最大列: {ws.max_column}")
print("=" * 80)

# 打印所有行（表头 + 字段名），完整不截断
for row_idx in range(1, 8):
    print(f"\n--- Row {row_idx} ---")
    for cell in ws[row_idx]:
        if cell.value is not None:
            print(f"  Col {cell.column} ({cell.column_letter}): {repr(str(cell.value))}")
