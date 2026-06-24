"""
将 GIGA 产品数据 AI 优化后填入 Amazon 德国模板 (PLANTER-de.xlsm)

工作流：
  GIGA API 取数 → AI 文案优化（MiniMax M3） → 填入 Excel 模板 → 保存

用法：
  python _fill_template.py                              # 默认：DE_TAX 市场，SKU W3372P314940
  python _fill_template.py --sku W3372P314940 --market DE_TAX
  python _fill_template.py --no-ai                      # 跳过 AI 优化，仅填入原始数据
"""

import sys
import os
import json
import argparse
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── GIGA API ──────────────────────────────────────────────────
from giga_config import get_credentials

def hmac_sha256_hex(message: str, key: str) -> str:
    import hmac, hashlib
    return hmac.new(key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).hexdigest()

def generate_nonce(length=10):
    import random, string
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def fetch_giga_product(sku: str, market: str) -> dict:
    """
    通过 GIGA OpenAPI v1 获取单个 SKU 的产品详情。
    成功返回产品字典，失败抛出 RuntimeError。
    """
    from giga_config import BASE_URL
    import time, base64

    client_id, client_secret = get_credentials(market)

    timestamp_ms = int(time.time() * 1000)
    nonce = generate_nonce(10)
    uri = "/b2b-overseas-api/v1/buyer/product/detailInfo/v1"
    msg = f"{client_id}&{uri}&{timestamp_ms}&{nonce}"
    key = f"{client_id}&{client_secret}&{nonce}"
    sign = base64.b64encode(hmac_sha256_hex(msg, key).encode("utf-8")).decode("utf-8")

    headers = {
        "Content-Type": "application/json",
        "client-id": client_id,
        "timestamp": str(timestamp_ms),
        "nonce": nonce,
        "sign": sign,
    }

    import requests
    try:
        resp = requests.post(
            f"{BASE_URL}{uri}",
            json={"skus": [sku]},
            headers=headers,
            timeout=30,
        )
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"GIGA API 请求失败: {e}")

    if resp.status_code != 200:
        raise RuntimeError(f"GIGA API 返回错误状态码 {resp.status_code}: {resp.text[:200]}")

    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError(f"GIGA API 返回非 JSON: HTTP {resp.status_code} {resp.text[:200]}")

    items = data.get("data") or []
    if not items:
        # 业务错误（B20003 等）也抛异常，message 直接来自 API
        msg_text = data.get("msg", "no data")
        sub_msg = data.get("subMsg", "")
        raise RuntimeError(f"GIGA API 返回空数据：{msg_text}（{sub_msg}）")
    return items[0]


# ── AI 优化 ───────────────────────────────────────────────────
def ai_optimize(product: dict, market: str = "DE_TAX") -> dict:
    """
    调用 _ai_optimizer 生成 AI 优化文案。
    若 image-studio server 不可用，返回原始字段（降级策略）。
    """
    try:
        from _ai_optimizer import generate_copy
        return generate_copy(product, market=market)
    except ImportError:
        print("  ⚠️ _ai_optimizer 未找到，使用原始数据")
        return {
            "title": product.get("productName", ""),
            "bullets": product.get("characteristics", [])[:5],
            "description": "",
            "search_terms": "",
        }


# ── Excel 写入 ───────────────────────────────────────────────
TEMPLATE_PATH = r"F:\AI Projects\GIGAB2B\PLANTER-de.xlsm"
TEMPLATE_DIR = os.path.dirname(TEMPLATE_PATH)

MARKET_TEMPLATE_PATHS = {
    "DE_TAX":     os.path.join(TEMPLATE_DIR, "PLANTER-de.xlsm"),
    "DE_TAXFREE": os.path.join(TEMPLATE_DIR, "PLANTER-de.xlsm"),
    "UK":         os.path.join(TEMPLATE_DIR, "PLANTER-uk.xlsm"),
    "US":         os.path.join(TEMPLATE_DIR, "PLANTER-us.xlsm"),
    "FR":         os.path.join(TEMPLATE_DIR, "PLANTER-fr.xlsm"),
}


def fill_template(
    product: dict,
    ai_result: dict,
    row: int = 7,
    out_path: str | None = None,
    image_strategy: str = "use_giga",
    market: str = "DE_TAX",
) -> str:
    """
    将产品数据（原始 + AI 优化）写入 Excel 模板。

    字段映射：
      A  col=1   SKU
      B  col=2   Produkttyp（固定 PLANTER）
      C  col=3   Angebotsaktion（full_update）
      G  col=7   商品名称（AI 优化标题）
      T  col=20  Modellnummer
      U  col=21  Hersteller
      J  col=10  Produkt-ID-Typ（GTIN-Freistellung）
      X  col=24  主图 URL
      Y-AB cols 25-28  其他图片 URL（最多8张）
      AO col=41  要点1（AI 优化）
      AP col=42  要点2
      AQ col=43  要点3
      AR col=44  要点4
      AS col=45  要点5
      AT col=46  Suchbegriffe（AI 优化）
      AU col=47  特殊属性1
      AV col=48  特殊属性2
      AW col=49  特殊属性3
      AX col=50  特殊属性4
      AY col=51  特殊属性5
      AZ col=52  Stil
      BA col=53  Material
      BH col=60  颜色
      BF col=58  Anzahl der Artikel
      BI col=61  Größe（参考 assembled dims）
      CZ col=104 深度
      DA col=105 深度单位
      DB col=106 高度
      DC col=107 高度单位
      DD col=108 宽度
      DE col=109 宽度单位
      DB col=115 物品重量
      DL col=116 重量单位
      HK col=219 原产国
      FS col=175 包装长度
      FT col=176 包装长度单位
      FU col=177 包装宽度
      FV col=178 包装宽度单位
      FW col=179 包装高度
      FX col=180 包装高度单位
      FY col=181 包装重量
      FZ col=182 包装重量单位
      JL col=271 说明书 PDF
      AN col=40  产品描述（AI 优化）
    """
    template_path = MARKET_TEMPLATE_PATHS.get(market, TEMPLATE_PATH)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb["Vorlage"]

    def w(col: int, val):
        ws.cell(row=row, column=col, value=val)

    sku = product.get("sku", "")
    attrs = product.get("attributes", {})
    imgs = product.get("imageUrls") or []
    chars = product.get("characteristics") or []

    # ── 基本信息 ──────────────────────────────────────────────
    w(1,  sku)                                     # A  SKU
    w(2,  "PLANTER")                               # B  Produkttyp
    w(3,  "full_update")                            # C  Angebotsaktion
    w(7,  ai_result.get("title") or product.get("productName", ""))  # G  商品名称
    w(10, "GTIN-Freistellung")                     # J  Produkt-ID-Typ
    # K → 空（GTIN 豁免）
    w(20, product.get("mpn", ""))                  # T  Modellnummer
    w(21, "YUDA HOME FURNITURE")                   # U  Hersteller

    # ── 图片 ─────────────────────────────────────────────────
    if imgs:
        w(24, imgs[0])                             # X  主图
    for i, url in enumerate(imgs[1:9], start=25):  # Y-AB 其他图
        w(i, url)

    # ── AI 优化字段 ───────────────────────────────────────────
    bullets = ai_result.get("bullets", []) or chars[:5]

    w(41, bullets[0] if len(bullets) > 0 else "")  # AO
    w(42, bullets[1] if len(bullets) > 1 else "")  # AP
    w(43, bullets[2] if len(bullets) > 2 else "")  # AQ
    w(44, bullets[3] if len(bullets) > 3 else "")  # AR
    w(45, bullets[4] if len(bullets) > 4 else "")  # AS

    st = ai_result.get("search_terms", "").strip()
    if not st:
        color = attrs.get("Main Color") or product.get("mainColor", "")
        material = product.get("mainMaterial", "")
        st = " ".join([
            "Hochbeet Metall", "Pflanzenbeet", "Gartenbeet", "Stahlblech",
            "Gemüsebeet", "Kräuterbeet", "Rostfrei", "Blumenbeet",
            material, color,
        ])
    w(46, st)                                      # AT  Suchbegriffe

    w(40, ai_result.get("description", ""))        # AN  产品描述

    # ── 特殊属性（来自 characteristics 整理）──────────────────
    special_attrs = [
        "Robustes Stahlblech mit Zink-Aluminium Beschichtung",
        "Wetterfest und Rostschutz",
        "Einfache Montage, Bausatz ohne Boden",
        "Offenes Design für freies Wurzelwachstum",
        "Ideal für Gemüse, Kräuter und Blumen",
    ]
    for i, attr in enumerate(special_attrs, start=47):
        w(i, attr)

    # ── 样式/材质 ─────────────────────────────────────────────
    color_val = attrs.get("Main Color") or product.get("mainColor") or ""
    w(52, attrs.get("Product Style", "Casual,Classic,Farmhouse"))  # AZ  Stil
    w(53, product.get("mainMaterial", "Metal"))        # BA  Material
    w(60, color_val)                                 # BH  颜色
    w(58, 1)                                         # BF  Anzahl der Artikel

    # BI  Größe → 留空（用户手动填）

    # ── 产品尺寸 ──────────────────────────────────────────────
    def safe_float(val, default=0.0):
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    length = safe_float(product.get("assembledLength"))
    width  = safe_float(product.get("assembledWidth"))
    height = safe_float(product.get("assembledHeight"))

    w(104, length)                                  # CZ  深度（长边）
    w(105, "cm")                                    # DA
    w(106, height)                                  # DB  高度
    w(107, "cm")                                    # DC
    w(108, width)                                   # DD  宽度
    w(109, "cm")                                    # DE

    # ── 商品重量 ─────────────────────────────────────────────
    w(115, safe_float(product.get("weightKg"), 0))  # DB  物品重量
    w(116, "kg")                                    # DL

    # ── 原产国 ───────────────────────────────────────────────
    w(219, product.get("placeOfOrigin", "China"))    # HK

    # ── 包装尺寸（平板包装估算）───────────────────────────────
    pkg_l = round(116 / 2, 1)
    pkg_w = round(30  / 2, 1)
    pkg_h = round(5.5 * 4, 1)
    pkg_wt = round(safe_float(product.get("weightKg")) / 2, 1)

    w(175, pkg_l)                                   # FS  Paketlänge
    w(176, "cm")                                    # FT
    w(177, pkg_w)                                   # FU  Paketbreite
    w(178, "cm")                                    # FV
    w(179, pkg_h)                                   # FW  Pakethöhe
    w(180, "cm")                                    # FX
    w(181, pkg_wt)                                  # FY  Paketgewicht
    w(182, "kg")                                    # FZ

    # ── 说明书 PDF ───────────────────────────────────────────
    file_urls = product.get("fileUrls") or []
    if file_urls:
        w(271, file_urls[0])                        # JL

    # ── 图片策略说明（写入单元格注释）────────────────────────────────
    if image_strategy != "use_giga":
        ws.cell(row=row, column=24).comment = (
            f"图片策略: {image_strategy} | "
            "AI 生成图片请在 image-studio 中下载后上传到 Seller Central"
        )

    # ── 保存 ─────────────────────────────────────────────────
    out = out_path or TEMPLATE_PATH
    wb.save(out)
    print(f"\n  ✅ 已保存: {out}")
    return out


# ── 主流程 ────────────────────────────────────────────────────

def main(sku: str, market: str, skip_ai: bool = False):
    print("=" * 60)
    print(f"  GIGAB2B → AI 优化 → Amazon DE 模板")
    print(f"  SKU: {sku}  |  市场: {market}  |  AI: {'跳过' if skip_ai else '启用'}")
    print("=" * 60)

    # 1. 从 GIGA 取数
    print(f"\n[1/3] 从 GIGA API 获取产品数据 ({market})...")
    product = fetch_giga_product(sku, market)
    print(f"  ✅ SKU={sku} 获取成功")
    print(f"     productName: {product.get('productName','')[:80]}...")

    # 2. AI 优化
    if skip_ai:
        print("\n[2/3] ⏭️ 跳过 AI 优化（--no-ai）")
        ai_result = {
            "title": product.get("productName", ""),
            "bullets": product.get("characteristics", [])[:5],
            "description": "",
            "search_terms": "",
        }
    else:
        print("\n[2/3] 调用 AI 文案优化（MiniMax M3 via image-studio）...")
        ai_result = ai_optimize(product, market=args.market)
        if ai_result.get("title"):
            print(f"     AI 标题: {ai_result['title'][:80]}...")
        if ai_result.get("bullets"):
            print(f"     AI 要点: {len(ai_result['bullets'])} 条")
        if ai_result.get("search_terms"):
            print(f"     AI Suchbegriffe: {ai_result['search_terms'][:80]}...")

    # 3. 填入模板
    print("\n[3/3] 写入 Excel 模板...")
    out_path = fill_template(product, ai_result, market=market)

    # 4. 摘要
    print("\n" + "=" * 60)
    print("  填写摘要（第7行）")
    print("=" * 60)

    imgs = product.get("imageUrls") or []
    bullets = ai_result.get("bullets") or product.get("characteristics", [])[:5]

    filled = [
        ("A",    "SKU",              sku),
        ("B",    "Produkttyp",       "PLANTER"),
        ("G",    "商品名称(AI)",      (ai_result.get("title") or product.get("productName",""))[:60] + "..."),
        ("T",    "Modellnummer",     product.get("mpn", "")),
        ("U",    "Hersteller",       "YUDA HOME FURNITURE"),
        ("X",    "主图 URL",         (imgs[0] or "无")[:80] + "..."),
        ("Y-AB", "其他图片",         f"{max(0, len(imgs)-1)} 张（主图除外）"),
        ("AO-AS","要点(AI)",         f"✓ {len(bullets)} 条"),
        ("AT",   "Suchbegriffe(AI)", (ai_result.get("search_terms") or "无")[:60] + "..."),
        ("AN",   "产品描述(AI)",     f"{len(ai_result.get('description') or '')} 字符"),
        ("AZ",   "Stil",             product.get("attributes", {}).get("Product Style", "-")),
        ("BA",   "Material",         product.get("mainMaterial", "-")),
        ("BH",   "Farbe",            (product.get("attributes", {}).get("Main Color") or product.get("mainColor") or "-")),
        ("CZ-DE","产品尺寸",         f"L={product.get('assembledLength','-')} W={product.get('assembledWidth','-')} H={product.get('assembledHeight','-')} cm"),
        ("HK",   "Ursprungsland",    product.get("placeOfOrigin", "-")),
        ("JL",   "说明书 PDF",       (product.get("fileUrls") or ["(无)"])[0][:60]),
    ]
    for col, field, val in filled:
        print(f"  [{col:8}] {field}: {val}")

    print("\n  以下字段仍需手动填写：")
    manual = [
        ("I",    "Markenname（品牌）",      "GIGA 数据为空，需自行填入"),
        ("J-K",  "Produkt-ID-Typ / ID",    "J=GTIN-Freistellung ✓，K=空 ✓"),
        ("L-P",  "Browse Node ID",         "需查 Amazon 德国站点表"),
        ("BI",   "Größe（尺寸）",           "可参考 assembled dimensions"),
        ("EN-ER","Fulfillment / 价格 / 库存", "需填价格策略"),
    ]
    for col, field, note in manual:
        print(f"  [{col}] {field} → {note}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GIGAB2B → AI 优化 → Amazon DE 模板")
    parser.add_argument("--sku",      default="W3372P314940", help="GIGA 产品 SKU")
    parser.add_argument("--market",   default="DE_TAX",       help="市场: DE_TAX / DE_TAXFREE / US / UK")
    parser.add_argument("--no-ai",    action="store_true",   help="跳过 AI 优化，使用原始数据")
    args = parser.parse_args()

    main(args.sku, args.market, skip_ai=args.no_ai)
