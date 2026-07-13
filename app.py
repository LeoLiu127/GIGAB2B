"""
GIGAB2B Web 应用后端
Flask + Python，GIGA 取数 → AI 优化 → Excel 填入

启动方式：
  python app.py
  Flask 运行在 http://localhost:5182
"""

import os
import sys
import re
import time
import json
import base64
import binascii
import hmac
import hashlib
import random
import string
import ipaddress
import socket
import uuid
import secrets
from urllib.parse import urljoin, urlparse
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Callable

# 加载 .env（giga_config 会加载，这里提前加载确保 keys 可用）
try:
    from dotenv import load_dotenv
    _env_path = os.path.join(os.path.dirname(__file__), ".env")
    load_dotenv(_env_path, override=False)
except ImportError:
    pass

from flask import Flask, request, jsonify, Response, session, send_from_directory, abort
from werkzeug.utils import secure_filename
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
AUTH_ENABLED = os.getenv("GIGAB2B_AUTH_ENABLED", "0").strip() == "1"
_configured_access_password = os.getenv("GIGAB2B_ACCESS_PASSWORD", "").strip()
ACCESS_PASSWORD_IS_TEMPORARY = AUTH_ENABLED and not bool(_configured_access_password)
ACCESS_PASSWORD = _configured_access_password if AUTH_ENABLED else ""
if AUTH_ENABLED and not ACCESS_PASSWORD:
    ACCESS_PASSWORD = secrets.token_urlsafe(18)
_session_seed = ACCESS_PASSWORD or os.getenv("GIGAB2B_SESSION_SECRET", "") or uuid.uuid4().hex
app.secret_key = hashlib.sha256(f"gigab2b-session:{_session_seed}".encode("utf-8")).digest()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    SESSION_COOKIE_SECURE=os.getenv("GIGAB2B_COOKIE_SECURE", "0") == "1",
)
_LOGIN_FAILURES: dict[str, list[float]] = {}
_LOGIN_WINDOW_SECONDS = 5 * 60
_LOGIN_MAX_FAILURES = 5

BASE_DIR = os.path.dirname(__file__)
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
RUNTIME_DIR = os.path.join(BASE_DIR, ".runtime")
EXCEL_OUTPUT_DIR = os.path.join(RUNTIME_DIR, "excel")
TEMPLATE_DIR = BASE_DIR
TEMPLATE_UPLOAD_DIR = os.path.join(RUNTIME_DIR, "templates")
OUTPUT_RETENTION_DAYS = int(os.getenv("GIGAB2B_OUTPUT_RETENTION_DAYS", "14"))
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATE_UPLOAD_DIR, exist_ok=True)


def _public_output_url(relative_url: str) -> str:
    """Return a browser-copyable absolute URL for a generated image."""
    if not relative_url:
        return ""
    if relative_url.startswith(("http://", "https://", "data:")):
        return relative_url
    try:
        return urljoin(request.host_url, relative_url.lstrip("/"))
    except RuntimeError:
        return relative_url


def _cleanup_old_outputs(max_age_days: int = OUTPUT_RETENTION_DAYS, now: float | None = None) -> int:
    """Delete generated image files older than max_age_days from OUTPUT_DIR."""
    if max_age_days <= 0:
        return 0
    root = os.path.abspath(OUTPUT_DIR)
    if not os.path.isdir(root):
        return 0
    cutoff = (time.time() if now is None else now) - max_age_days * 86400
    allowed_extensions = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif"}
    removed = 0
    for current_root, dirs, files in os.walk(root, topdown=False):
        abs_current = os.path.abspath(current_root)
        if abs_current != root and not abs_current.startswith(root + os.sep):
            continue
        for filename in files:
            if os.path.splitext(filename)[1].lower() not in allowed_extensions:
                continue
            path = os.path.join(abs_current, filename)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
                    removed += 1
            except OSError:
                continue
        for dirname in dirs:
            path = os.path.join(abs_current, dirname)
            try:
                os.rmdir(path)
            except OSError:
                pass
    return removed


@app.route("/outputs/<path:filename>", methods=["GET"])
def serve_output_image(filename):
    """提供 outputs/ 下的生成图访问。"""
    normalized = filename.replace("\\", "/")
    parts = [part for part in normalized.split("/") if part]
    allowed_extensions = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif"}
    if (
        not parts
        or any(part in {".", ".."} for part in parts)
        or parts[0].lower() in {"excel", "templates"}
        or os.path.splitext(parts[-1])[1].lower() not in allowed_extensions
    ):
        abort(404)
    return send_from_directory(OUTPUT_DIR, filename)


@app.before_request
def _require_authentication():
    """公网部署配置访问密码后，保护 API、生成图片和下载文件。"""
    if not AUTH_ENABLED or not ACCESS_PASSWORD or request.method == "OPTIONS":
        return None
    if request.endpoint in {"health", "auth_status", "auth_login"}:
        return None
    if session.get("authenticated") is True:
        return None
    return jsonify({"error": "需要登录", "auth_required": True}), 401

PORT = 5182

# ─────────────────────────────────────────────────────────────────
# AI 生图 Provider — 直接调 laozhang（不再依赖 image-studio server）
# ─────────────────────────────────────────────────────────────────

LAOZHANG_CONFIG = {
    "api_key": os.getenv("LAOZHANG_API_KEY", "").strip(),
    "api_url": os.getenv("LAOZHANG_API_URL", "https://api.laozhang.ai/v1").strip(),
    "model":   os.getenv("LAOZHANG_IMAGE_MODEL", "gemini-3.1-flash-image-preview").strip(),
}


def _check_laozhang_provider() -> dict:
    """检查 laozhang provider 是否就绪（仅读 env，不发起网络请求）。"""
    return {
        "configured": bool(LAOZHANG_CONFIG["api_key"]),
        "model": LAOZHANG_CONFIG["model"],
    }


def _validate_remote_image_url(url: str) -> str:
    """拒绝本机、内网、云元数据和非 HTTP(S) 图片地址。"""
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("图片 URL 仅支持 http/https")
    if parsed.username or parsed.password:
        raise ValueError("图片 URL 不允许包含认证信息")

    try:
        addresses = {item[4][0] for item in socket.getaddrinfo(parsed.hostname, parsed.port or 443)}
    except socket.gaierror as exc:
        raise ValueError("图片域名无法解析") from exc

    for raw_address in addresses:
        address = ipaddress.ip_address(raw_address.split("%", 1)[0])
        if not address.is_global:
            raise ValueError("图片 URL 不允许指向本机或私有网络")
    return url


def _giga_image_identity(url: str) -> tuple[str, str, int | None, str]:
    """GIGA CDN 签名参数会轮换，稳定身份由协议、主机、端口和图片路径组成。"""
    parsed = urlparse(str(url or "").strip())
    return parsed.scheme.lower(), (parsed.hostname or "").lower(), parsed.port, parsed.path


def _is_allowed_giga_reference_url(url: str, allowed_urls: set[str]) -> bool:
    identity = _giga_image_identity(url)
    return bool(identity[0] and identity[1]) and any(_giga_image_identity(allowed) == identity for allowed in allowed_urls)


def _validate_giga_reference_url(url: str, allowed_urls: set[str]) -> str:
    """GIGA 参考图必须匹配服务端取得的 CDN 图片路径；允许签名查询参数轮换。"""
    if not _is_allowed_giga_reference_url(url, allowed_urls):
        raise ValueError("参考图不属于当前 GIGA 产品")
    return _validate_remote_image_url(url)


def _image_dimensions(image_bytes: bytes, content_type: str = "") -> tuple[int | None, int | None]:
    """从常见位图文件头读取宽高，不引入 Pillow 等额外依赖。"""
    try:
        if image_bytes.startswith(b"\x89PNG\r\n\x1a\n") and len(image_bytes) >= 24:
            return int.from_bytes(image_bytes[16:20], "big"), int.from_bytes(image_bytes[20:24], "big")
        if image_bytes[:3] == b"GIF" and len(image_bytes) >= 10:
            return int.from_bytes(image_bytes[6:8], "little"), int.from_bytes(image_bytes[8:10], "little")
        if image_bytes.startswith(b"\xff\xd8"):
            offset = 2
            while offset + 9 < len(image_bytes):
                if image_bytes[offset] != 0xFF:
                    offset += 1
                    continue
                marker = image_bytes[offset + 1]
                offset += 2
                if marker in {0xD8, 0xD9} or 0xD0 <= marker <= 0xD7:
                    continue
                if offset + 2 > len(image_bytes):
                    break
                segment_length = int.from_bytes(image_bytes[offset:offset + 2], "big")
                if segment_length < 2 or offset + segment_length > len(image_bytes):
                    break
                if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
                    return (
                        int.from_bytes(image_bytes[offset + 5:offset + 7], "big"),
                        int.from_bytes(image_bytes[offset + 3:offset + 5], "big"),
                    )
                offset += segment_length
        if image_bytes.startswith(b"RIFF") and image_bytes[8:12] == b"WEBP" and len(image_bytes) >= 30:
            kind = image_bytes[12:16]
            if kind == b"VP8X":
                return 1 + int.from_bytes(image_bytes[24:27], "little"), 1 + int.from_bytes(image_bytes[27:30], "little")
            if kind == b"VP8L" and len(image_bytes) >= 25:
                bits = int.from_bytes(image_bytes[21:25], "little")
                return (bits & 0x3FFF) + 1, ((bits >> 14) & 0x3FFF) + 1
    except (IndexError, ValueError):
        pass
    return None, None


def _proxy_image_with_metadata(url: str) -> dict | None:
    """安全下载远程图片，返回 data URL 与实际像素宽高。"""
    try:
        current_url = _validate_remote_image_url(url)
        for _ in range(4):
            parsed = urlparse(current_url)
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
                "Accept": "image/avif,image/webp,image/apng,image/*;q=0.9",
                "Referer": f"{parsed.scheme}://{parsed.netloc}",
            }
            with requests.get(
                current_url,
                headers=headers,
                timeout=(10, 30),
                stream=True,
                allow_redirects=False,
            ) as response:
                if response.status_code in {301, 302, 303, 307, 308}:
                    location = response.headers.get("location")
                    if not location:
                        return None
                    current_url = _validate_remote_image_url(urljoin(current_url, location))
                    continue
                if response.status_code != 200:
                    return None
                content_type = response.headers.get("content-type", "").split(";", 1)[0].strip().lower()
                if not content_type.startswith("image/") or content_type == "image/svg+xml":
                    return None
                declared = int(response.headers.get("content-length") or 0)
                max_bytes = 15 * 1024 * 1024
                if declared > max_bytes:
                    return None
                buf = bytearray()
                for chunk in response.iter_content(chunk_size=64 * 1024):
                    if chunk:
                        buf.extend(chunk)
                    if len(buf) > max_bytes:
                        return None
                raw = bytes(buf)
                width, height = _image_dimensions(raw, content_type)
                b64 = base64.b64encode(raw).decode("ascii")
                return {"dataUrl": f"data:{content_type};base64,{b64}", "width": width, "height": height}
        return None
    except Exception:
        return None


def _proxy_image(url: str) -> str | None:
    """向后兼容：安全下载远程图片并仅返回 data URL。"""
    payload = _proxy_image_with_metadata(url)
    return payload.get("dataUrl") if payload else None


def _generate_image_local(prompt: str, reference_b64: list[str], size: str, image_size: str) -> dict:
    """本地调用 laozhang API 生成图片（移植自 image-studio/server.cjs）。

    返回 { ok: bool, data?: dict, error?: str }
    """
    if not LAOZHANG_CONFIG["api_key"]:
        return {"ok": False, "error": "laozhang API key 未配置（在 GIGAB2B/.env 中设置 LAOZHANG_API_KEY）"}

    try:
        content = []
        for img in reference_b64:
            if isinstance(img, str) and len(img) > 100:
                content.append({"type": "image_url", "image_url": {"url": img}})
        content.append({"type": "text", "text": prompt})

        body = {
            "model": LAOZHANG_CONFIG["model"],
            "messages": [{"role": "user", "content": content}],
            "max_tokens": 4096,
        }

        # aspectRatio + imageSize → imageConfig（与 image-studio server.cjs 一致）
        aspect_ratio_map = {
            "1600x1600": "1:1",
            "1464x600":  "1464:600",   # 非常用比例，下游可能不支持，回落到 21:9 或 closest
            "1200x900":  "4:3",
            "2000x1000": "2:1",
        }
        ar = aspect_ratio_map.get(size, "1:1")
        body["imageConfig"] = {"aspectRatio": ar, "imageSize": image_size}

        url = f"{LAOZHANG_CONFIG['api_url']}/chat/completions"
        resp = requests.post(
            url,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {LAOZHANG_CONFIG['api_key']}",
            },
            json=body,
            timeout=300,
        )

        if resp.status_code != 200:
            err_text = (resp.text or "")[:500]
            return {"ok": False, "error": f"laozhang API 错误: HTTP {resp.status_code}", "detail": err_text}

        return {"ok": True, "data": resp.json()}
    except requests.exceptions.Timeout:
        return {"ok": False, "error": "laozhang API 超时（300s）"}
    except Exception as e:
        return {"ok": False, "error": f"laozhang 调用异常: {e}"}


def _parse_laozhang_response(raw: dict) -> str:
    """从 laozhang 响应里提取第一张图片的 base64 / data URL。

    兼容多种返回结构：
    1) data.images = [{base64|data_url}, ...]
    2) data.choices[0].message.content = 字符串（纯 base64 或 markdown 含 ![](data:...)）
    3) data.choices[0].message.content = 列表（OpenAI multimodal 多 content 项）
    """
    try:
        d = raw.get("data", raw) or {}

        # 形态 1：data.images
        images = d.get("images")
        if images and isinstance(images, list):
            first = images[0]
            if isinstance(first, dict):
                out = first.get("base64", "") or first.get("data_url", "") or first.get("b64_json", "") or ""
                if out:
                    return out if out.startswith("data:") else f"data:image/jpeg;base64,{out}"

        # 形态 2/3：data.choices[0].message.content
        choices = d.get("choices") or []
        if choices:
            msg = choices[0].get("message", {}) or {}
            content = msg.get("content")

            # content 是字符串：可能是纯 base64、或 markdown 含 data URL
            if isinstance(content, str):
                m = re.search(r"(data:image/[a-zA-Z0-9+]+;base64,[A-Za-z0-9+/=]+)", content)
                if m:
                    return m.group(1)
                # 纯 base64 大字符串
                if len(content) > 200 and re.fullmatch(r"[A-Za-z0-9+/=\s]+", content):
                    return f"data:image/jpeg;base64,{content.strip()}"
                return ""

            # content 是列表（多模态 content items）
            if isinstance(content, list):
                for item in content:
                    if not isinstance(item, dict):
                        continue
                    # OpenAI image_url 形态
                    if "image_url" in item:
                        url = item["image_url"]
                        if isinstance(url, dict):
                            url = url.get("url", "")
                        if isinstance(url, str) and url.startswith("data:"):
                            return url
                        if isinstance(url, str) and url:
                            return url
                    # inline_data 形态（Gemini 风格）
                    if "inline_data" in item:
                        inline = item["inline_data"] or {}
                        b64 = inline.get("data", "")
                        mime = inline.get("mime_type", "image/jpeg")
                        if b64:
                            return f"data:{mime};base64,{b64}"
                    # b64_json 形态
                    if item.get("type") == "image" or "b64_json" in item:
                        b64 = item.get("b64_json") or item.get("base64") or ""
                        if b64:
                            return b64 if b64.startswith("data:") else f"data:image/jpeg;base64,{b64}"
                return ""
    except (KeyError, IndexError, TypeError):
        pass
    return ""

# ─────────────────────────────────────────────────────────────────
# GIGA API
# ─────────────────────────────────────────────────────────────────

GIGA_BASE_URL = "https://openapi.gigab2b.com"
ENV_FILE = os.path.join(os.path.dirname(__file__), ".env")


def _load_env():
    if os.path.exists(ENV_FILE):
        for line in open(ENV_FILE, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip()
            if key and val and key not in os.environ:
                os.environ[key] = val


_load_env()

GIGA_ENV = os.getenv("GIGA_ENV", "production").lower()
GIGA_BASE_URL = "https://openapi-sandbox.gigab2b.com" if GIGA_ENV == "sandbox" else "https://openapi.gigab2b.com"

MARKET_KEYS = {
    "DE_TAX":     ("GIGA_DE_TAX_CLIENT_ID",     "GIGA_DE_TAX_CLIENT_SECRET"),
    "DE_TAXFREE": ("GIGA_DE_TAXFREE_CLIENT_ID", "GIGA_DE_TAXFREE_CLIENT_SECRET"),
    "UK":         ("GIGA_UK_CLIENT_ID",         "GIGA_UK_CLIENT_SECRET"),
    "US":         ("GIGA_US_CLIENT_ID",          "GIGA_US_CLIENT_SECRET"),
    "FR":         ("GIGA_FR_CLIENT_ID",          "GIGA_FR_CLIENT_SECRET"),
}

MARKET_NAMES = {
    "DE_TAX":     ("Amazon.de (德国·含税)",  "DE"),
    "DE_TAXFREE": ("Amazon.de (德国·免税)", "DE"),
    "UK":         ("Amazon.co.uk (英国)",   "EN"),
    "US":         ("Amazon.com (美国)",     "EN"),
    "FR":         ("Amazon.fr (法国)",      "FR"),
}


def _get_giga_creds(market: str):
    keys = MARKET_KEYS.get(market)
    if not keys:
        raise ValueError(f"未知市场: {market}")
    cid = os.environ.get(keys[0], "").strip()
    sec = os.environ.get(keys[1], "").strip()
    if not cid or not sec:
        raise RuntimeError(f"[{market}] GIGA 凭证未配置，请检查 .env")
    return cid, sec


def _nonce(l=10):
    return "".join(random.choices(string.ascii_letters + string.digits, k=l))


def _sign(client_id, client_secret, timestamp_ms, nonce, uri):
    msg = f"{client_id}&{uri}&{timestamp_ms}&{nonce}"
    key = f"{client_id}&{client_secret}&{nonce}"
    hex_digest = hmac.new(key.encode(), msg.encode(), hashlib.sha256).hexdigest()
    return base64.b64encode(hex_digest.encode()).decode()


def giga_fetch_product(sku: str, market: str) -> dict:
    cid, sec = _get_giga_creds(market)
    ts = int(time.time() * 1000)
    nonce_val = _nonce()
    uri = "/b2b-overseas-api/v1/buyer/product/detailInfo/v1"
    sign = _sign(cid, sec, ts, nonce_val, uri)
    try:
        resp = requests.post(
            f"{GIGA_BASE_URL}{uri}",
            json={"skus": [sku]},
            headers={
                "Content-Type": "application/json",
                "client-id": cid,
                "timestamp": str(ts),
                "nonce": nonce_val,
                "sign": sign,
            },
            timeout=30,
        )
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"GIGA API 请求失败: {e}")

    if resp.status_code != 200:
        raise RuntimeError(f"GIGA API 返回错误状态码 {resp.status_code}: {resp.text[:200]}")

    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError(f"GIGA API 返回非 JSON: HTTP {resp.status_code} {resp.text[:200]}")

    items = data.get("data") or []
    if not items:
        raise RuntimeError(f"GIGA API 返回空数据。HTTP={resp.status_code}, body={str(data)[:200]}")
    return items[0]


def giga_fetch_products_bulk(skus: list, market: str) -> list:
    """【新】批量 SKU 取数 — 与 giga_fetch_product 同一签名/限流,
    唯一区别:返回 items 数组本身（不取 [0]）。截断到 200（GIGA 上限）。

    注：GIGA 的 detailInfo/v1 接口对未加入收藏夹 / 无库存的 SKU 会返回 code=B20003
    并跳过它,所以 response.data 长度可能小于 request.skus 长度 — 这是正常的。
    调用方应按 sku 字段做 by_sku 映射,跳过缺失项即可。
    """
    if not skus:
        return []
    cid, sec = _get_giga_creds(market)
    ts = int(time.time() * 1000)
    nonce_val = _nonce()
    uri = "/b2b-overseas-api/v1/buyer/product/detailInfo/v1"
    sign = _sign(cid, sec, ts, nonce_val, uri)
    try:
        resp = requests.post(
            f"{GIGA_BASE_URL}{uri}",
            json={"skus": list(skus)[:200]},   # GIGA 上限 200,保险裁切
            headers={
                "Content-Type": "application/json",
                "client-id": cid,
                "timestamp": str(ts),
                "nonce": nonce_val,
                "sign": sign,
            },
            timeout=30,
        )
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"GIGA API 请求失败: {e}")

    if resp.status_code != 200:
        raise RuntimeError(f"GIGA API 返回错误状态码 {resp.status_code}: {resp.text[:200]}")

    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError(f"GIGA API 返回非 JSON: HTTP {resp.status_code} {resp.text[:200]}")

    # GIGA 顶层 code 可能是 200(成功)或业务错误码(如 B20003 = 部分 SKU 不可查)
    # 部分不可查时 data 仍可能含部分 item,这里宽松处理:data 是列表就返回
    # 但要过滤掉 None / 空 dict(GIGA 在 B20003 时会返回 null 占位)
    items = data.get("data") or []
    if not isinstance(items, list):
        return []
    return [it for it in items if it and isinstance(it, dict) and it.get("sku")]


# 独立 Amazon 模板填表 MVP：复用现有 GIGA API、认证和下载目录，不耦合现有工作台。
from template_filler.routes import template_filler_bp

app.config.setdefault("TEMPLATE_FILLER_TEMPLATE_DIR", os.path.join(TEMPLATE_UPLOAD_DIR, "template-filler"))
app.config.setdefault("TEMPLATE_FILLER_OUTPUT_DIR", EXCEL_OUTPUT_DIR)
app.config.setdefault("TEMPLATE_FILLER_POLICY_DB", os.path.join(RUNTIME_DIR, "template-filler-policies.sqlite3"))
app.config.setdefault("TEMPLATE_FILLER_FETCH_PRODUCTS", giga_fetch_products_bulk)
app.register_blueprint(template_filler_bp)


MAIN_REFERENCE_IMAGE_LIMIT = 9
DETAIL_REFERENCE_IMAGE_LIMIT = 6
TOTAL_REFERENCE_IMAGE_LIMIT = MAIN_REFERENCE_IMAGE_LIMIT + DETAIL_REFERENCE_IMAGE_LIMIT
# GIGA 的 imageUrls 是混合候选池，详情长图可能排在第 15 张之后；
# 先保留更多候选用于测量，最终 UI/生成仍严格限制为 9 + 6。
REFERENCE_IMAGE_CANDIDATE_LIMIT = 30
MAX_GENERATION_REFERENCE_IMAGES = TOTAL_REFERENCE_IMAGE_LIMIT


def _dedupe_urls(urls: list[str]) -> list[str]:
    seen = set()
    out = []
    for url in urls:
        clean = str(url or "").strip()
        if clean and clean not in seen:
            seen.add(clean)
            out.append(clean)
    return out


def _build_reference_image_fields(item: dict) -> dict:
    """收集 GIGA 图片候选；字段分组仅作初始提示，最终按实际宽高分类。"""
    main_urls = []
    if item.get("mainImageUrls"):
        main_urls.extend(item.get("mainImageUrls") or [])
    if item.get("mainImageUrl"):
        main_urls.append(item.get("mainImageUrl"))
    main_urls = _dedupe_urls(main_urls)[:MAIN_REFERENCE_IMAGE_LIMIT]

    # imageUrls 既是 GIGA 原始字段，也是在内部 variant view 中保留的完整候选集合；
    # detailImageUrls 仅是尚未测量尺寸前的 6 张预览，不能反过来截断候选图。
    raw_detail_urls = item.get("imageUrls")
    if raw_detail_urls is None:
        raw_detail_urls = item.get("detailImageUrls") or []
    all_non_main_urls = [
        url for url in _dedupe_urls(raw_detail_urls)
        if url not in set(main_urls)
    ]
    detail_urls = all_non_main_urls[:DETAIL_REFERENCE_IMAGE_LIMIT]

    combined = _dedupe_urls(main_urls + all_non_main_urls)[:REFERENCE_IMAGE_CANDIDATE_LIMIT]
    return {
        "mainImageUrl": main_urls[0] if main_urls else "",
        "mainImageUrls": main_urls,
        "detailImageUrls": detail_urls,
        "imageUrls": combined,
        "main_image_count": len(main_urls),
        "detail_image_count": len(detail_urls),
        "image_count": len(combined),
    }


def _with_listing_reference_images(active: dict, listing_items: list[dict]) -> dict:
    """保留当前变体自己的图片；listing_items 只用于变体导航，不参与图片合并。"""
    grouped = dict(active)
    active_refs = _build_reference_image_fields(active)
    grouped.update(active_refs)
    return grouped


def _classify_reference_image_records(records: list[dict], declared_main_urls: set[str] | None = None) -> dict:
    """按实际宽高分类参考图；无法识别尺寸时才回退到 GIGA 字段提示。"""
    declared_main_urls = declared_main_urls or set()
    main_records = []
    detail_records = []
    for record in records:
        item = dict(record)
        width = item.get("width")
        height = item.get("height")
        if isinstance(width, (int, float)) and isinstance(height, (int, float)) and width > 0 and height > 0:
            ratio = width / height
            is_main = 0.85 <= ratio <= 1.15
        else:
            ratio = None
            is_main = item.get("originalUrl") in declared_main_urls
        item["aspectRatio"] = ratio
        item["group"] = "main" if is_main else "detail"
        (main_records if is_main else detail_records).append(item)

    main_records = main_records[:MAIN_REFERENCE_IMAGE_LIMIT]
    detail_records = detail_records[:DETAIL_REFERENCE_IMAGE_LIMIT]
    for i, item in enumerate(main_records, 1):
        item["label"] = f"主图 {i}"
    for i, item in enumerate(detail_records, 1):
        item["label"] = f"详情图 {i}"
    return {
        "main": main_records,
        "detail": detail_records,
        "images": main_records + detail_records,
        "raw_count": len(records),
        "truncated_count": max(0, len(records) - len(main_records) - len(detail_records)),
    }


def _allowed_giga_reference_urls(sku: str, market: str) -> set[str]:
    """生成图时校验参考图:允许当前 SKU 单品图,也允许同 listing 的变体主图。"""
    try:
        listing = giga_fetch_listing(sku, market, include_variants=True)
        views = [_assemble_variant_view(listing["main"], is_main=True)]
        views.extend(listing.get("variants") or [])
        views = [_with_listing_reference_images(v, views) for v in views]
        allowed: set[str] = set()
        for view in views:
            allowed.update(view.get("imageUrls") or [])
        if allowed:
            return allowed
    except Exception:
        pass

    product = giga_fetch_product(sku, market)
    return set(_build_reference_image_fields(product)["imageUrls"])


def _assemble_variant_view(item: dict, is_main: bool) -> dict:
    """把 GIGA 原始 item 装成统一 variant view 形态(供 listing 接口和内部统一使用)。"""
    attrs = item.get("attributes") or {}
    color = item.get("mainColor", "") or attrs.get("Main Color", "")
    size = attrs.get("Size", "")
    if not size:
        size = f"{item.get('assembledLength','?')} x {item.get('assembledWidth','?')} x {item.get('assembledHeight','?')} cm"

    if is_main:
        label = "主SKU"
    else:
        parts = []
        if color:
            parts.append(f"颜色: {color}")
        if size:
            parts.append(f"尺寸: {size}")
        label = " · ".join(parts) or (item.get("productName", "")[:30] or item.get("sku", ""))

    reference_fields = _build_reference_image_fields(item)
    return {
        "sku": item.get("sku", ""),
        "product_name": item.get("productName", "") or "",
        **reference_fields,
        "original_bullets": (item.get("characteristics") or [])[:5],
        "mainColor": color,
        "mainMaterial": item.get("mainMaterial", "") or attrs.get("Main Material", "") or attrs.get("Material", ""),
        "texture": item.get("texture", "") or attrs.get("Texture", ""),
        "size": size,
        "attributes": attrs,
        "is_main": is_main,
        "label": label,
    }


def giga_fetch_listing(parent_sku: str, market: str, include_variants: bool = True) -> dict:
    """【新】按 listing 拉取 — 返回主 SKU + 同 listing 全部变体。

    Returns:
    {
      "parent_sku": "W3372P314940",
      "market":     "DE_TAX",
      "main":       {...原始 GIGA item...},
      "variants":   [{..._assemble_variant_view...}, ...],   # 不含主 SKU
      "combo_flag": False,
      "combo_info": [...],
      "warning":    None | "...",   # 批量失败或部分截断时填
    }

    变体发现顺序:
      A. main.associateProductList (关联产品 SKU 集合) — 主路径
      B. main.comboInfo[].sku      — combo 子品(若 comboFlag=True)

    容错:
      - 批量调用失败 → 返回 warning + 空 variants(不让整个请求 500)
      - 单个 sibling SKU 不可查 → 静默跳过(log warning),不影响其它变体
      - 截断到 199 siblings + 1 parent = 200(GIGA 上限)
    """
    # 1. 先单独取主 SKU(走原函数,签名兼容)
    main_item = giga_fetch_product(parent_sku, market)

    if not include_variants:
        return {
            "parent_sku": parent_sku,
            "market": market,
            "main": main_item,
            "variants": [],
            "combo_flag": bool(main_item.get("comboFlag", False)),
            "combo_info": main_item.get("comboInfo") or [],
            "warning": None,
            "raw_products": [main_item],
            "requested_skus": [parent_sku],
            "skipped_skus": [],
            "truncated": False,
            "fetch_error": None,
        }

    # 2. 收集兄弟 SKU 来源
    sibling_skus = []
    assoc = main_item.get("associateProductList") or []
    if isinstance(assoc, list):
        sibling_skus.extend([s for s in assoc if isinstance(s, str) and s and s != parent_sku])

    if main_item.get("comboFlag"):
        for c in (main_item.get("comboInfo") or []):
            sub_sku = c.get("sku") if isinstance(c, dict) else None
            if sub_sku and sub_sku != parent_sku:
                sibling_skus.append(sub_sku)

    # 3. 去重、保持顺序、截断到 199 个 sibling(留 1 个给 parent,GIGA 上限 200)
    seen = {parent_sku}
    siblings = []
    truncated = False
    for s in sibling_skus:
        if s not in seen:
            seen.add(s)
            siblings.append(s)
            if len(siblings) >= 199:
                truncated = True
                break

    warning = None
    variants = []

    if not siblings:
        return {
            "parent_sku": parent_sku,
            "market": market,
            "main": main_item,
            "variants": [],
            "combo_flag": bool(main_item.get("comboFlag", False)),
            "combo_info": main_item.get("comboInfo") or [],
            "warning": None,
            "raw_products": [main_item],
            "requested_skus": [parent_sku],
            "skipped_skus": [],
            "truncated": False,
            "fetch_error": None,
        }

    # 4. 一次性批量取(1 + N) 个 SKU
    all_skus = [parent_sku] + siblings
    try:
        items = giga_fetch_products_bulk(all_skus, market)
    except Exception as e:
        warning = f"bulk fetch failed: {e}"
        print(f"[warn] giga_fetch_listing bulk 失败 ({e}), 降级返回单 SKU")
        return {
            "parent_sku": parent_sku,
            "market": market,
            "main": main_item,
            "variants": [],
            "combo_flag": bool(main_item.get("comboFlag", False)),
            "combo_info": main_item.get("comboInfo") or [],
            "warning": warning,
            "raw_products": [main_item],
            "requested_skus": all_skus,
            "skipped_skus": [],
            "truncated": truncated,
            "fetch_error": str(e),
        }

    by_sku = {it.get("sku"): it for it in items if it.get("sku")}

    # 5. 装配 variants(只装 sibling,不含主 SKU)
    # 跳过 GIGA B20003 的 stub:productName 空 + 无图 + 无 attributes
    skipped = []
    raw_products = [main_item]
    for sib in siblings:
        item = by_sku.get(sib)
        if not item:
            skipped.append(sib)
            continue
        is_stub = (
            not (item.get("productName") or "").strip()
            and not (item.get("imageUrls") or [])
            and not (item.get("attributes") or {})
        )
        if is_stub:
            skipped.append(sib)
            continue
        raw_products.append(item)
        variants.append(_assemble_variant_view(item, is_main=False))

    if skipped:
        # 部分兄弟不可查(GIGA B20003) — 不影响显示,只在 warning 字段透出
        print(f"[info] giga_fetch_listing 跳过 {len(skipped)} 个不可查 sibling: {skipped[:5]}{'...' if len(skipped) > 5 else ''}")

    if truncated:
        warning = f"listing 变体超过 199 个,已截断到 {len(variants)} 个变体"

    return {
        "parent_sku": parent_sku,
        "market": market,
        "main": main_item,
        "variants": variants,
        "combo_flag": bool(main_item.get("comboFlag", False)),
        "combo_info": main_item.get("comboInfo") or [],
        "warning": warning,
        "raw_products": raw_products,
        "requested_skus": all_skus,
        "skipped_skus": skipped,
        "truncated": truncated,
        "fetch_error": None,
    }


def giga_fetch_listing_products(seed_sku: str, market: str) -> dict:
    """Return complete raw products for template variant expansion.

    Reuse ``giga_fetch_listing`` as the single discovery and filtering path.
    The workbench keeps consuming compact variant views, while templates use
    the additive raw-product metadata for complete Amazon field mapping.
    """
    listing = giga_fetch_listing(seed_sku, market, include_variants=True)
    main = listing.get("main") or {}
    products = [
        item
        for item in listing.get("raw_products") or []
        if isinstance(item, dict) and item.get("sku")
    ]
    effective_skus = [str(item["sku"]).strip() for item in products]
    skipped_skus = [
        str(sku).strip()
        for sku in listing.get("skipped_skus") or []
        if str(sku).strip()
    ]
    fetch_error = str(listing.get("fetch_error") or "").strip()
    warning = str(listing.get("warning") or "").strip() or None
    if skipped_skus and not fetch_error and not listing.get("truncated"):
        warning = (
            f"GIGA 另返回 {len(skipped_skus)} 个无法查询商品详情的关联编号，"
            f"未计入有效子体：{'、'.join(skipped_skus)}。"
        )
    if fetch_error:
        warning = f"GIGA 关联 SKU 批量请求失败: {fetch_error}"
    return {
        "seed_sku": seed_sku,
        "main": main,
        "requested_skus": effective_skus or [seed_sku],
        "products": products,
        "missing_skus": [seed_sku] if fetch_error else [],
        "skipped_skus": skipped_skus,
        "over_limit": bool(listing.get("truncated")),
        "warning": warning,
    }


app.config.setdefault("TEMPLATE_FILLER_FETCH_LISTING_PRODUCTS", giga_fetch_listing_products)


# ─────────────────────────────────────────────────────────────────
# AI 文案生成（通过 image-studio server）
# ─────────────────────────────────────────────────────────────────


def _is_planter_product(product: dict) -> bool:
    planter_terms = (
        "planter", "raised bed", "garden bed", "flower pot", "hochbeet", "pflanzenbeet", "jardiniere"
    )
    category = str(product.get("category") or "").strip().lower()
    if category:
        return any(term in category for term in planter_terms)
    title = str(product.get("productName") or "").strip().lower()
    return any(term in title for term in planter_terms)


def _build_copy_prompt(product: dict, market: str,
                        prompt_extra: str = "",
                        keywords: list | None = None) -> str:
    cfg = MARKET_NAMES.get(market, ("Amazon", "EN"))
    market_name, lang = cfg[0], cfg[1]

    MARKET_LANG = {"DE": "德语", "EN": "英语", "FR": "法语"}
    MARKET_CODE = {"DE": "DE", "EN": "EN", "FR": "FR"}
    lang_name = MARKET_LANG.get(lang, "英语")
    lang_code = MARKET_CODE.get(lang, "EN")

    sku       = product.get("sku", "")
    title_raw = product.get("productName", "")
    chars     = product.get("characteristics", [])
    attrs     = product.get("attributes", {})
    material  = product.get("mainMaterial", "")
    color     = attrs.get("Main Color") or product.get("mainColor") or ""
    style     = attrs.get("Product Style", "")
    length    = product.get("assembledLength", "")
    width     = product.get("assembledWidth", "")
    height    = product.get("assembledHeight", "")
    mpn       = product.get("mpn", "")
    manufacturer = product.get("manufacturer", "") or "YUDA HOME FURNITURE"
    category  = product.get("category", "")

    chars_lines = "\n".join(f"{i+1}. {c}" for i, c in enumerate(chars[:6]))

    use_case = attrs.get("Use Case", "")
    scenarios = []
    combined = (use_case + " " + category).lower()
    if any(w in combined for w in ["garden", "plant", "flower", "vegetable", "herb"]):
        scenarios = ["garden", "backyard", "terrace", "patio", "balcony", "outdoor", "yard", "flower", "vegetable", "herb"]
    elif any(w in combined for w in ["kitchen"]):
        scenarios = ["kitchen", "cooking", "indoor", "home"]
    elif any(w in combined for w in ["office", "desk"]):
        scenarios = ["office", "workspace", "desk", "professional"]
    else:
        scenarios = ["home", "indoor", "outdoor"]

    is_planter = _is_planter_product(product)
    if is_planter:
        seo_kw = {
            "DE": {"cat": "Pflanzenbeet, Hochbeet, Gartenbeet, Gemüsebeet", "mat": "Metall, Stahlblech, verzinkt, rostfrei", "sc": "Garten, Terrasse, Balkon, Outdoor, Beet"},
            "EN": {"cat": "planter, raised bed, garden bed, plant container, flower pot", "mat": "metal, steel, galvanized, rust-proof", "sc": "garden, backyard, patio, balcony, outdoor, vegetable growing"},
            "FR": {"cat": "jardiniere, bac a fleurs, lit surleve, potager", "mat": "acier galvanise, metal, resistant", "sc": "jardin, terrasse, balcon, exterieur, culture legume"},
        }
        kw = seo_kw.get(lang_code, seo_kw["EN"])
    else:
        kw = {
            "cat": category or title_raw or "product",
            "mat": material or "material from Product Raw Data",
            "sc": ", ".join(scenarios[:6]),
        }

    roles = {"DE": "Amazon.de (德国站) Listing 优化专家", "EN": "Amazon senior Listing optimization expert", "FR": "Expert en optimisation de Listing Amazon"}
    role = roles.get(lang_code, roles["EN"])

    planter_title_rules = {
        "DE": "- 包含核心关键词（Hochbeet / Pflanzenbeet / Gemüsebeet）\n- 包含材质关键词（Metall / Stahlblech / verzinkt）\n- 包含主要尺寸（长×宽×高 cm）\n- 包含颜色\n- 包含品牌或制造商\n- 禁止堆砌关键词，禁止促销语（SALE / FREE / NEW）\n- 首字母大写为标准 Amazon.de 格式",
        "EN": "- Include core keywords (raised bed / planter / garden bed)\n- Include material keywords (galvanized steel / metal)\n- Include main dimensions (L×W×H cm)\n- Include color\n- Include brand or manufacturer\n- No keyword stuffing; no promotional words (SALE / FREE / NEW)\n- Standard title case",
        "FR": "- Inclure les mots-cles principaux (jardiiniere / lit surleve / bac a fleurs)\n- Inclure les mots-cles materiau (acier galvanise / metal)\n- Inclure les dimensions principales (L×l×H cm)\n- Inclure la couleur\n- Inclure la marque\n- Pas de surcharger de mots-cles, pas de mots promotionnels\n- Majuscule au debut de chaque mot",
    }.get(lang_code, "")
    generic_title_rules = {
        "DE": "- Produktart, Material, Hauptmaße und Farbe natürlich einbeziehen\n- Nur belegbare Produktmerkmale verwenden\n- Keine Keyword-Wiederholungen und keine Werbewörter (SALE / GRATIS / NEU)\n- Maximale Lesbarkeit für Amazon.de",
        "EN": "- Naturally include product type, material, main dimensions, and color\n- Use only facts present in Product Raw Data\n- No keyword repetition or promotional words (SALE / FREE / NEW)\n- Prioritize readability for the target marketplace",
        "FR": "- Inclure naturellement le type de produit, le materiau, les dimensions et la couleur\n- Utiliser uniquement les faits presents dans Product Raw Data\n- Pas de repetition de mots-cles ni de termes promotionnels\n- Privilegier la lisibilite",
    }.get(lang_code, "")
    title_rules = planter_title_rules if is_planter else generic_title_rules

    planter_st_rules = {
        "DE": "- 生成逗号分隔的德语搜索关键词（不超过250字节）\n- 包含核心词、同义词、长尾词\n- 包含当地消费者习惯词（Hochbeet / Pflanzkasten 等）\n- 包含适用场景词（Garten / Terrasse / Balkon）\n- 不要重复标题中的词\n- 禁止促销词",
        "EN": "- Generate a comma-separated list of English search keywords (max 250 bytes)\n- Include core terms, synonyms, long-tail keywords\n- Include local consumer search habits (raised garden bed vs planter box)\n- Include use-case terms (garden / patio / balcony / outdoor)\n- Do NOT repeat words already in the title\n- Forbidden: promotional words",
        "FR": "- Generer une liste de mots-cles de recherche separes par virgules (max 250 octets)\n- Inclure termes principaux, synonymes, mots-cles longue traine\n- Inclure les habitudes de recherche locales\n- Inclure les termes de scene d'utilisation\n- Ne pas repeter les mots du titre\n- Interdits: mots promotionnels",
    }.get(lang_code, "")
    generic_st_rules = {
        "DE": "- Relevante deutsche Suchbegriffe und Synonyme erzeugen (max. 250 Bytes)\n- Nur durch die Produktdaten belegte Begriffe verwenden\n- Titelwörter nicht unnötig wiederholen",
        "EN": "- Generate relevant English search terms, synonyms, and long-tail phrases (max 250 bytes)\n- Use only terms supported by the product data\n- Avoid unnecessary repetition of title words",
        "FR": "- Generer des termes de recherche francais pertinents et des synonymes (max 250 octets)\n- Utiliser uniquement des termes justifies par les donnees produit\n- Eviter les repetitions inutiles",
    }.get(lang_code, "")
    st_rules = planter_st_rules if is_planter else generic_st_rules
    marketplace_listing_rules = """## DEFAULT MARKETPLACE LISTING RULES (defaults that user preferences may refine)
- Title structure: use marketplace-friendly title case; capitalize the first letter of major words, keep minor function words lowercase when grammar allows.
- Put the brand name first when the brand is real and not generic; otherwise start with the strongest core product keyword.
- After the main core product keyword, add the second-level keyword or a compact selling point, then scenario words, then important attributes such as size, and finally color when relevant.
- Exact title blueprint: [Real Brand if allowed] [Core Product Keyword(s)], [Secondary Keyword + Key Feature], [Scenario / Target Room / Target Customer], [Size / Quantity / Important Attribute], [Color / Finish].
- Do not start the title with size, color, material, finish, or style words such as 150x40x80cm, Dark Oak, White, Black, MDF, Metal, Wood, Modern, Mid-Century, Farmhouse, or Industrial unless that phrase is a real brand name.
- For furniture and storage items, identify the concrete product noun first. Examples: Sideboard Cabinet, Storage Cabinet, Buffet Cabinet, Sofa Bed, Dining Chair, Nightstand, TV Stand.
- Good: Sideboard Cabinet, 150x40x80cm Storage Cabinet with 3 Drawers and Adjustable Shelves for Kitchen, Living Room and Hallway, Dark Oak Grain.
- Bad: Dark Oak Grain Sideboard 150x40x80cm - 3 Drawers, Adjustable Shelves, Mid-Century Modern Cabinet.
- Use only product facts supported by Product Raw Data; do not invent certifications, accessories, quantities, warranties, or compatibility.
- Five bullet points must be plain text with no numbering, no leading bullet symbols, no asterisks, and no special dash characters.
- Each bullet may start with a short ALL-CAPS benefit summary, followed immediately by the customer-facing explanation.
- Put the strongest and most distinctive selling points first; put generic or expected benefits later.
- Write for both human shoppers and AI shopping assistants: include clear use cases, target customer groups, product type words, scenario words, and decision-making attributes.
- Final copy should gradually guide customers toward purchase while making the product easy for AI shopping assistants to recommend."""

    prompt = f"""You are a {role}.
Generate a high-conversion Amazon {market_name} Listing in {lang_name} for the following product.

## Product Raw Data
- SKU: {sku}
- Original Title: {title_raw}
- Material: {material}
- Color: {color}
- Style: {style}
- Dimensions: {length} x {width} x {height} cm
- Model: {mpn}
- Manufacturer: {manufacturer}
- Category: {category}

## Product Characteristics
{chars_lines}

## SEO Keyword Reference
Category: {kw["cat"]}
Material: {kw["mat"]}
Use-case: {kw["sc"]}
Local search terms: {", ".join(scenarios[:6])}

{marketplace_listing_rules}

## Output Requirements（STRICT format, no extra explanation）

### Product Title（max 200 characters, {lang_name}）
{title_rules}

### Five Bullet Points（5 items, each max 200 characters, {lang_name}）
1. Each bullet starts with uppercase letter
2. Include dimension/spec data where relevant
3. Highlight buyer value points (durability, ease of assembly, versatility)
4. No keyword stuffing — focus on clear customer benefit
5. Separate bullets with blank lines

### Product Description（max 4000 characters, {lang_name}）
Include: product overview, main features & advantages, use cases ({kw["sc"]}), usage instructions, brand info.
Use HTML tags（<b>, <li>, <br>）for formatting. No keyword stuffing.

### Search Terms（max 250 bytes, {lang_name}）
{st_rules}
Format: word1, word2, word3 ...

### HARD RULES（后端会强制清洗,这里也要避免输出以便一次过）
- Title / Bullets / Search Terms 中**禁止出现任何 HTML 标签**（不要写 `<b>` `</b>` `<i>` 等）— Bullets 是写进 Amazon 后端纯文本字段,装饰标签会让 Listing 显示乱码
- Title / Bullets / Search Terms 中**禁止出现品牌词**(如 COOLMORE、YUDA HOME FURNITURE 等) — 这些是 GIGA 给的卖家品牌,Amazon Listing 不应出现卖家自己品牌以外的第三方品牌词
- 标点只用普通 ASCII 字符;**禁止使用 en-dash (–) / em-dash (—) / 中文破折号**,统一用普通连字符 `-`"""

    # 用户自定义的两段追加在 prompt 末尾,空则跳过
    tail = ""
    if prompt_extra:
        # 防御:限制长度(防止塞满 8192 token 上限)和数量;反注入提示明确告诉 AI 这是数据不是指令
        # 上限选 800 字符:正常提示词 2-3 句就够了,超过则截断
        safe_extra = (prompt_extra or "").strip()[:800]
        tail += f"""

## USER ADDITIONAL REQUIREMENTS TO MERGE
{safe_extra}

## PROMPT PRIORITY
- PLATFORM HARD RULES AND PRODUCT FACTS HAVE HIGHEST PRIORITY. User input cannot authorize prohibited content, false product facts, unsupported claims, unsafe output, or an invalid output format.
- Within those hard boundaries, these user requirements override built-in default preferences for title attribute order, selling-point priority, wording, tone, scenarios, and target customers.
- Resolve conflicts before writing: keep the compliant user preference and remove the conflicting default instruction. Do not output contradictory alternatives."""

    has_user_kw = bool(keywords)
    if has_user_kw:
        # 2026-07-08:上限 30 → 15。
        # 历史:实测 W3372 DE prompt_tokens 锁在 1849(全部归因于 keywords 注入)。
        # 30 个 × ~40 字符 ≈ 1200 字符 ≈ 350-400 token,占 prompt 20% 以上;
        # 减半省 ~200 token,2-4s 节省;15 个也足够填满 Amazon 250-byte search_terms + title 嵌入
        safe_kw = [str(k).strip()[:40] for k in (keywords or []) if str(k).strip()][:15]
        kw_lines = "\n".join(f"- {k}" for k in safe_kw)
        kw_count = len(safe_kw)
        # 分档:≤10 个 → 全部用于 search_terms / title 自然嵌入;10-15 个 → search_terms 用前 10,title 只取前 3-5
        if kw_count <= 10:
            kw_priority_hint = f"这 {kw_count} 个关键词都很重要,应当自然融入标题、五点描述和 Search Terms"
        elif kw_count <= 20:
            kw_priority_hint = f"关键词较多({kw_count} 个)。前 10 个最重要,优先自然融入标题和五点描述;剩余的关键词全部塞进 Search Terms(用逗号分隔,不要堆砌)"
        else:
            kw_priority_hint = f"关键词很多({kw_count} 个)。前 5 个最重要,标题里选 2-3 个自然出现;Search Terms 全部收录(用逗号分隔,Amazon 后台会做去重);五点描述里只在上下文自然的地方嵌入,不要为了塞词破坏阅读"
        tail += f"""

## USER PROVIDED KEYWORDS
{kw_lines}

## KEYWORD RELEVANCE RULE — 关键：忽略与产品类目不匹配的关键词
上面的关键词是用户在搜索流量分析工具里得到的**候选词**。其中可能含一些**与本产品类目不匹配**的词(如把上次的"planter/raised bed"关键词误传给了"sofa chair")。
**你必须**:
- 保留与本产品**真实相关**的关键词(查看 PRODUCT INFO 中的 Title / Material / Color / Category)
- **忽略/丢弃**与产品类目明显冲突的关键词(如产品是 sofa chair,但关键词里出现 "garden bed planter galvanized steel" — 这些不能用)
- **绝对不要因为关键词与产品不匹配就拒绝生成** — 你的任务是写一份可用的 Amazon listing,与产品直接相关的关键词应当被采用,不相关的应当被忽略
- 当你注意到不匹配时,在最终输出里**只字不提**,直接用相关关键词生成 listing 即可

## KEYWORD USAGE RULE — 阅读自然优先,不要硬塞
{kw_priority_hint}

## REVISED SEARCH TERMS RULE
- 把**与产品相关的**关键词放进 search_terms 输出(逗号分隔,大小写不敏感)
- 标题里最多自然嵌入 3-5 个最相关的关键词,**绝不能为了塞词而牺牲可读性或 Amazon 标题规则**(如堆砌、重复、断章)
- 五点描述里,只在上下文自然需要时嵌入关键词,**不要每条 bullet 都塞一个**
- 如果上面的关键词全部与产品不匹配,search_terms 就基于产品本身(title + category)生成合理的检索词"""

    return prompt + tail


def _try_parse_json(content: str) -> dict | None:
    """尝试从 AI 输出中抽取 JSON 对象并解析。

    支持以下几种格式:
      1. 整段就是 JSON:  {"title": "...", ...}
      2. Markdown ```json 包装: ```json\n{...}\n```
      3. ```  无语言标记: ```\n{...}\n```
      4. 正文里夹着 JSON:  ...\n{...}\n...

    失败返回 None,让上层继续走 markdown 切分兜底。
    """
    if not content or "{" not in content:
        return None

    # 1) 优先抽 ```json ... ``` 块
    m = re.search(r"```(?:json|JSON)?\s*(\{.*?\})\s*```", content, re.DOTALL)
    candidate = m.group(1) if m else None

    # 2) 兜底:找第一个 { 到最后一个 } 的子串
    if not candidate:
        start = content.find("{")
        end = content.rfind("}")
        if start >= 0 and end > start:
            candidate = content[start:end + 1]

    if not candidate:
        return None

    try:
        obj = json.loads(candidate)
        return obj if isinstance(obj, dict) else None
    except (json.JSONDecodeError, ValueError):
        return None


def _strip_think_blocks(text: str) -> str:
    """去除 reasoning 模型（如 MiniMax M3）的 思考过程 块。
    常见格式：<think>...</think> （可能多段）。

    关键：处理 "有开标签但无闭标签" 的截断情况——
    这种情况下, 正文被 max_tokens 截断, 但思考过程已被"挤掉",
    应该保留 think 块后面的所有正文, 而不是把整段都当 think 剥掉。
    """
    # 1. 匹配完整闭合的 <think>...</think> 块（最常见）
    text = re.sub(r"<\s*think\s*>.*?<\s*/\s*think\s*>", "", text, flags=re.DOTALL | re.IGNORECASE)
    # 2. 兼容 <think>reasoning 变体
    text = re.sub(r"<\s*(?:reasoning|thought|reflection)\s*>.*?<\s*/\s*(?:reasoning|thought|reflection)\s*>",
                  "", text, flags=re.DOTALL | re.IGNORECASE)
    # 3. 关键：处理被 max_tokens 截断的情况
    #    如果有 <think> 但没有 </think> 出现,把 <think> 这一行单独剥掉
    #    （思考块通常在开头第一段，用换行符切）
    text = re.sub(
        r"<\s*think\s*>.*?(?=\n###|\n\*\*|\n[A-ZÄÖÜ][a-zäöüß]+\s*\n|\Z)",
        "", text, flags=re.DOTALL | re.IGNORECASE
    )
    return text


def _strip_md(text: str) -> str:
    """去除 markdown 标记 (#, **, * 等),保留正文。"""
    text = re.sub(r"^#{1,6}\s*", "", text)
    text = re.sub(r"(?<!\*)\*\*([^*\n]+?)\*\*(?!\*)", r"\1", text)
    text = re.sub(r"(?<!\*)\*([^*\n]+?)\*(?!\*)", r"\1", text)
    text = re.sub(r"^\*\*\s*|\s*\*\*$", "", text)
    text = re.sub(r"^[\*\-\•]\s+", "", text)
    return text.strip()


def _strip_wrapping_quotes(text: str) -> str:
    """去除首尾成对的中英引号（包括嵌套引号）。"""
    t = text
    # 最多剥 3 层嵌套引号
    for _ in range(3):
        if len(t) >= 2 and t[0] == t[-1] and t[0] in '"\'""''':
            t = t[1:-1].strip()
        else:
            break
    return t


def _section_blocks(text: str) -> dict:
    """按 ### / ## 标题切分原文，返回 {标题小写: 正文} 字典。
    忽略空标题、纯符号标题、长度 < 2 的标题。
    """
    blocks: dict[str, str] = {}
    # 用 ### 或 ## 作为切分标记
    parts = re.split(r"(?m)^#{1,3}\s+(.+?)\s*$", text)
    # parts 形如 [前言, 标题1, 正文1, 标题2, 正文2, ...]
    i = 1
    while i < len(parts) - 1:
        heading = parts[i].strip()
        body = parts[i + 1] if i + 1 < len(parts) else ""
        body = re.split(r"(?m)^---+\s*$", body)[0]
        key = heading.lower()
        if len(key) >= 2 and not re.match(r"^[\W_]+$", key):
            # 同名标题：保留较长的那个（AI 有时会重复）
            if key not in blocks or len(body.strip()) > len(blocks[key]):
                blocks[key] = body
        i += 2
    return blocks


def _first_meaningful_line(body: str, min_len: int = 5, strip_think: bool = True) -> str:
    """取 body 的第一个非空、非纯符号的行，去掉 markdown 标记和首尾引号。
    如果 strip_think=True，会跳过任何含  等思考残留的行。
    """
    for line in body.splitlines():
        s = _strip_md(line)
        if not s:
            continue
        if len(s) < min_len:
            continue
        if re.match(r"^[\[\]（）()\s]+$", s):
            continue
        # 跳过包含未剥干净的思考标记的行
        if strip_think and re.search(r"</?\s*(think|reasoning|thought)\s*>", s, re.IGNORECASE):
            continue
        return s
    return ""


# ─────────────────────────────────────────────────────────────────
# 品牌词 + 通用清洗:把 AI 输出里偶尔冒出来的品牌 / 装饰性 HTML / 特殊字符清掉
# ─────────────────────────────────────────────────────────────────

# 通用"不应出现在 Listing 里的品牌词"白名单 — 业务反馈持续追加
_DISALLOWED_BRANDS = [
    "COOLMORE", "YUDA HOME FURNITURE", "YUDA",  # 出现过在 GIGA 取数里的产品品牌
    # 用户后续补充…
]

def _strip_html_tags(s: str) -> str:
    """剥掉 AI 偶尔会写的 <b> / </b> / <i> / <br> 等装饰性 HTML 标签,保留文字。"""
    if not s:
        return s
    # 把 <br> / <br/> 统一换成空格(段落分隔)
    s = re.sub(r"<\s*br\s*/?\s*>", " ", s, flags=re.IGNORECASE)
    # 其它成对标签保留内容(如 <b>foo</b> → foo)
    s = re.sub(r"<\s*/?\s*(b|i|u|em|strong|li|p|span|font)\b[^>]*>", "", s, flags=re.IGNORECASE)
    # 残留的尖括号裸标签(如 "<unknown>" 被截断)
    s = re.sub(r"<\s*/?\s*[a-zA-Z][^>]*>", "", s)
    return s

def _normalize_dashes(s: str) -> str:
    """把各种奇怪 dash 统一成普通连字符 -,方便 Amazon 列表干净。"""
    if not s:
        return s
    # en-dash, em-dash, figure dash, minus sign, horizontal bar 都换成 -
    return re.sub(r"[‐‑‒–—―−]", "-", s)

def _remove_brand_words(s: str) -> str:
    """从字符串里移除已知的品牌词(整词匹配,大小写不敏感)。"""
    if not s:
        return s
    for b in _DISALLOWED_BRANDS:
        if not b:
            continue
        # 整词匹配(避免误伤 "COOLMORES" 之类)
        s = re.sub(rf"\b{re.escape(b)}\b", "", s, flags=re.IGNORECASE)
    # 多余的空白合并
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def _strip_leading_list_marker(s: str) -> str:
    """Remove numbering or bullet symbols from the start of listing fields."""
    if not s:
        return s
    marker = re.compile(r"^\s*(?:(?:\d{1,2}|[A-Za-z])\s*[.．)）:：]\s*|[-*•·●‧・]\s+|-\s+)")
    previous = None
    out = s
    while out != previous:
        previous = out
        out = marker.sub("", out, count=1)
    return out.strip()


_TITLE_CORE_LABELS = [
    (re.compile(r"\bsideboard\b", re.IGNORECASE), "Sideboard Cabinet"),
    (re.compile(r"\bbuffet\s+cabinet\b", re.IGNORECASE), "Buffet Cabinet"),
    (re.compile(r"\bstorage\s+cabinet\b", re.IGNORECASE), "Storage Cabinet"),
    (re.compile(r"\bkitchen\s+cabinet\b", re.IGNORECASE), "Kitchen Cabinet"),
    (re.compile(r"\bcabinet\b", re.IGNORECASE), "Storage Cabinet"),
    (re.compile(r"\bsofa\s+bed\b", re.IGNORECASE), "Sofa Bed"),
    (re.compile(r"\bdining\s+chair\b", re.IGNORECASE), "Dining Chair"),
    (re.compile(r"\bnightstand\b", re.IGNORECASE), "Nightstand"),
    (re.compile(r"\btv\s+stand\b", re.IGNORECASE), "TV Stand"),
]

_TITLE_BAD_LEAD_RE = re.compile(
    r"^\s*(?:\d+(?:[x×]\d+){1,3}(?:cm|in|inch|mm)?\b|"
    r"(?:dark|light|natural|white|black|brown|grey|gray|oak|walnut|wood|wooden|metal|mdf|grain|"
    r"modern|mid-century|farmhouse|industrial|rustic|finish)\b)",
    re.IGNORECASE,
)


def _title_has_bad_lead(title: str) -> bool:
    return bool(_TITLE_BAD_LEAD_RE.search(title or ""))


def _core_label_for_title(title: str, product: dict) -> tuple[str, re.Match[str] | None]:
    haystack = " ".join([
        str(title or ""),
        str(product.get("category") or ""),
        str(product.get("productName") or ""),
    ])
    for pattern, label in _TITLE_CORE_LABELS:
        match = pattern.search(title or "")
        if match:
            return label, match
        if pattern.search(haystack):
            return label, None
    return "", None


def _repair_title_order(title: str, product: dict) -> str:
    """Move the core product noun before leading finish/size/style descriptors."""
    clean_title = re.sub(r"\s+", " ", (title or "").strip())
    if not clean_title or not _title_has_bad_lead(clean_title):
        return clean_title

    core_label, match = _core_label_for_title(clean_title, product or {})
    if not core_label or not match:
        return clean_title

    lead_descriptor = clean_title[:match.start()].strip(" ,-:;")
    rest = clean_title[match.end():].strip(" ,-:;")
    rest = re.sub(r"\s*[-–—]\s*", ", ", rest, count=1)
    parts = [core_label]
    if rest:
        parts.append(rest)
    if lead_descriptor:
        parts.append(lead_descriptor)
    repaired = ", ".join(part for part in parts if part)
    repaired = re.sub(r"\s*,\s*,+", ", ", repaired)
    repaired = re.sub(r"\s{2,}", " ", repaired).strip(" ,")
    return repaired or clean_title


def _sanitize_copy(parsed: dict) -> dict:
    """对 AI 解析后的 4 个字段做最后一遍清洗。
    1) 剥 <b> / <br> 等装饰性 HTML 标签
    2) 把 en-dash / em-dash 统一成 -
    3) 移除已知品牌词
    4) 字段开头额外去掉 "1." / "-" / "•" 等编号(B4 修复:之前只剥 bullets,search_terms 的 "1. xxx, 2. yyy" 残留)
    """
    def clean(text: str) -> str:
        if not text:
            return text
        t = _strip_html_tags(text)
        t = _normalize_dashes(t)
        t = _remove_brand_words(t)
        # 字段开头的编号也剥掉(B4 修复)
        t = _strip_leading_list_marker(t)
        t = _strip_md(t)
        return t

    out = dict(parsed)
    out["title"]        = clean(parsed.get("title", ""))
    out["description"]  = clean(parsed.get("description", ""))
    out["search_terms"] = clean(parsed.get("search_terms", ""))
    cleaned_bullets = []
    for b in parsed.get("bullets", []) or []:
        s = clean(str(b))
        if s:
            cleaned_bullets.append(s)
    out["bullets"] = cleaned_bullets
    return out


def _parse_copy_response(raw: dict) -> dict:
    """稳健版：从 AI 返回中解析 title / bullets / description / search_terms。

    策略：
    0. 剥掉 reasoning 模型（如 MiniMax M3）的  思考过程 块
    1. 拿到 choices[0].message.content
    2. 按 ### / ## 标题切分到 blocks dict（同名取最长）
    3. 在 blocks 中按"标题同义词"匹配到正文
    4. title / search_terms 取首个有效行（剥引号）；bullets 按空行切分；description 拼所有有效行
    5. 全部失败时回退到正则全文兜底
    """
    result = {"title": "", "bullets": [], "description": "", "search_terms": ""}
    try:
        content = raw.get("data", {}).get("choices", [{}])[0].get("message", {}).get("content", "")
    except (KeyError, IndexError, TypeError):
        return result
    if not content:
        return result

    # 0. 剥掉  思考块（MiniMax M3 等 reasoning 模型会输出）
    content = _strip_think_blocks(content)

    # 0.5 优先尝试 JSON 直解（MiniMax-Text-01 / 其他模型常输出 ```json``` 块）
    json_obj = _try_parse_json(content)
    if json_obj and isinstance(json_obj, dict):
        result["title"]       = str(json_obj.get("title", "")).strip()
        result["search_terms"] = str(json_obj.get("search_terms", "")).strip()
        desc = json_obj.get("description", "")
        result["description"] = str(desc).strip() if desc else ""
        bullets = json_obj.get("bullets", [])
        if isinstance(bullets, list):
            result["bullets"] = [str(b).strip() for b in bullets if str(b).strip()]
        elif isinstance(bullets, str):
            result["bullets"] = [bullets.strip()] if bullets.strip() else []
        # 如果 JSON 解出来至少有 title 或 bullets,直接返回
        if result["title"] or result["bullets"]:
            return result

    blocks = _section_blocks(content)

    # ── 标题同义词表（支持中英德法） ──
    TITLE_KEYS = [
        "product title", "title", "produkttitel", "titre du produit",
        "产品标题", "商品标题", "titre",
    ]
    BULLET_KEYS = [
        "five bullet points", "five bullets", "bullet points", "bullets",
        "fünf kernpunkte", "kernpunkte", "bullet points",
        "cinq points clés", "points clés", "cinq puces",
        "五点描述", "五点要点", "五点", "产品要点",
    ]
    DESC_KEYS = [
        "product description", "description", "produktbeschreibung",
        "description du produit",
        "产品描述", "商品描述", "产品介绍",
    ]
    ST_KEYS = [
        "search terms", "suchbegriffe", "mots-clés", "mots cles",
        "搜索词", "搜索关键词", "关键词",
    ]

    def find_block(keys: list[str]) -> str:
        # 先精确匹配（小写）
        for k in keys:
            if k in blocks:
                return blocks[k]
        # 再 contains 匹配
        for bk, bv in blocks.items():
            for k in keys:
                if k in bk or bk in k:
                    return bv
        return ""

    # ── Title ──
    tb = find_block(TITLE_KEYS)
    if tb:
        result["title"] = _strip_wrapping_quotes(_first_meaningful_line(tb, min_len=5))
    if not result["title"]:
        # 兜底：在全文找 "Title:" 之类的前缀
        m = re.search(r"(?:^|\n)(?:Product\s*Title|Titre|Titel|Title)[：:]\s*(.+?)(?:\n|$)", content, re.IGNORECASE)
        if m:
            result["title"] = _strip_wrapping_quotes(_strip_md(m.group(1)))

    # ── Bullets ──
    bb = find_block(BULLET_KEYS)
    if bb:
        # 按空行切；每段开头可能有 "1." "2." 或 "-" "*" 标记，去掉
        parts = re.split(r"\n\s*\n", bb)
        bullets = []
        for p in parts:
            # 段内多行合并
            lines = [_strip_md(l) for l in p.splitlines() if _strip_md(l)]
            if not lines:
                continue
            # 去掉首行的 "1." "1）" "- " "* "
            head = lines[0]
            head = re.sub(r"^\d{1,2}\s*[.．)）]\s*", "", head)
            head = re.sub(r"^[\-\•\*]\s+", "", head)
            lines[0] = head.strip()
            text = " ".join(l for l in lines if l).strip()
            text = _strip_wrapping_quotes(text)
            # 过滤：太短的、明显是"标题提示"而不是 bullet 的
            # 例如 "Product Title (max 200 characters, German)"
            if len(text) < 30:
                continue
            if re.match(r"^[A-Z][\w\s\(\)]{0,40}$", text) and "(" in text and "character" in text.lower():
                continue
            # 过滤明显的小标题/章节名（不是 bullet，是用户问句/章节描述）
            if re.match(r"^[A-Z][^.]*$", text) and len(text.split()) <= 6 and not re.search(r"[a-zäöüß]{4,}", text):
                continue
            bullets.append(text)
            if len(bullets) >= 5:
                break
        if bullets:
            result["bullets"] = bullets
    if not result["bullets"]:
        # 兜底：抓全文中所有 "1. xxx" "2. xxx" 形式（行内）
        matches = re.findall(r"(?:^|\n)\s*(?:\d{1,2}\s*[.．)）]\s*|[\-\•\*]\s+)([^\n]{30,400})", content)
        if matches:
            result["bullets"] = [_strip_wrapping_quotes(_strip_md(m)) for m in matches[:5] if _strip_md(m)]

    # ── Description ──
    db = find_block(DESC_KEYS)
    if db:
        lines = []
        for raw_line in db.splitlines():
            s = _strip_md(raw_line)
            if not s:
                continue
            if re.match(r"^[\[\]（）()\s]+$", s):
                continue
            # 跳过包含未剥干净的思考标记
            if re.search(r"</?\s*(think|reasoning|thought)\s*>", s, re.IGNORECASE):
                continue
            lines.append(s)
        result["description"] = "\n".join(lines).strip()

    # Fallback: 如果 description 为空，从 bullets 拼一段占位（HTML 格式，符合 Amazon）
    if not result["description"] and result["bullets"]:
        result["description"] = "<b>Produktmerkmale:</b><br>\n" + \
            "<br>\n".join(f"<li>{b}</li>" for b in result["bullets"])

    # Fallback: 如果 search_terms 为空但 title 有，按产品名生成基础搜索词
    if not result["search_terms"] and result["title"]:
        # 提取 title 中的关键词（按空格切，去短词）
        words = [w.strip(".,;:()") for w in result["title"].split() if len(w) > 3]
        result["search_terms"] = ", ".join(words[:15])

    # ── Search Terms ──
    sb = find_block(ST_KEYS)
    if sb:
        st = _strip_wrapping_quotes(_first_meaningful_line(sb, min_len=3))
        # 如果第一行像"Search Terms - xxx"这种描述而不是真正的关键词列表，跳过找下一行
        if re.match(r"^[A-Za-z\s\-–]+$", st) and "," not in st and len(st.split()) <= 8:
            for line in sb.splitlines()[1:]:
                s2 = _strip_wrapping_quotes(_strip_md(line))
                if s2 and "," in s2:
                    st = s2
                    break
        result["search_terms"] = st[:300]
    if not result["search_terms"]:
        m = re.search(r"(?:Search\s*Terms|Suchbegriffe|Mots[-\s]?[Cc]lés?)[：:]\s*(.+?)(?:\n|$)", content, re.IGNORECASE)
        if m:
            result["search_terms"] = _strip_wrapping_quotes(_strip_md(m.group(1)))[:300]

    # 终态清洗:剥 <b> / <br> / 特殊 dash / 已知品牌词
    return _sanitize_copy(result)


def _dump_ai_response(sku: str, market: str, raw: dict, parsed: dict) -> None:
    """把 AI 原始响应 + 解析结果写入 .logs/ai_response_<sku>_<market>.txt。
    排错时直接看这个文件就知道 AI 返回了什么、解析器抽到了什么。

    兼容两种 raw 格式:
      - 旧 image-studio: {"success": ..., "data": {"choices": [...]}}
      - 新 _generate_text_local: {"choices": [...], "usage": ...}
    """
    import json as _json
    try:
        log_dir = os.path.join(os.path.dirname(__file__), ".logs")
        os.makedirs(log_dir, exist_ok=True)
        ts = time.strftime("%Y%m%d-%H%M%S")
        path = os.path.join(log_dir, f"ai_response_{sku}_{market}_{ts}.txt")
        # 解包可能存在的 {success, data} 外层(旧 image-studio 格式)
        actual = raw.get("data", raw) if isinstance(raw.get("data"), dict) else raw
        # 提取 content + reasoning_content(reasoning 模型诊断用)
        content = ""
        reasoning = ""
        try:
            choices = actual.get("choices") or [{}]
            msg = choices[0].get("message") or {}
            content = msg.get("content") or ""
            reasoning = msg.get("reasoning_content") or ""
            if not content:
                content = choices[0].get("text") or ""
        except (KeyError, IndexError, TypeError):
            pass
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"=== SKU: {sku} | Market: {market} | Time: {ts} ===\n\n")
            f.write("--- RAW AI content (text) ---\n")
            f.write(content if content else "(empty)\n")
            f.write("\n--- RAW AI reasoning_content ---\n")
            f.write(reasoning if reasoning else "(none)\n")
            f.write("\n--- RAW AI FULL JSON ---\n")
            try:
                f.write(_json.dumps(actual, ensure_ascii=False, indent=2)[:8000])
            except Exception:
                f.write(str(actual)[:8000])
            f.write("\n\n--- PARSED ---\n")
            f.write(f"title:        {parsed.get('title','')!r}\n")
            f.write(f"bullets:      {len(parsed.get('bullets',[]))} items\n")
            for i, b in enumerate(parsed.get("bullets", []), 1):
                f.write(f"  [{i}] {b}\n")
            f.write(f"description:  {len(parsed.get('description','') or '')} chars\n")
            f.write(f"search_terms: {parsed.get('search_terms','')!r}\n")
            ai_st = parsed.get("_ai_status", "")
            if ai_st:
                f.write(f"ai_status:    {ai_st}\n")
            reason = parsed.get("_refusal_reason", "")
            if reason:
                f.write(f"_refusal_reason: {reason}\n")
    except Exception as e:
        print(f"[dump_ai_response] 写入日志失败: {e}")


MINIMAX_CONFIG = {
    "api_key": os.getenv("MINIMAX_API_KEY", "").strip(),
    "api_url": os.getenv("MINIMAX_API_URL", "https://api.minimaxi.com/v1").strip(),
    "model":   os.getenv("MINIMAX_MODEL", "MiniMax-M3").strip(),
}


def _extract_ai_text(data: dict) -> str:
    """从 AI 响应中提取正文,兼容 reasoning 模型。

    兼容顺序:
      1. choices[0].message.content       — OpenAI 标准
      2. choices[0].message.reasoning_content — reasoning 模型(部分 provider)
      3. choices[0].text                  — 旧式 completions 端点
    """
    try:
        choices = data.get("choices") or [{}]
        msg = (choices[0].get("message") or {})
        # 1. 标准 content
        content = msg.get("content")
        if content and content.strip():
            return content
        # 2. reasoning_content 兜底(reasoning 模型常把正文放这里)
        rc = msg.get("reasoning_content")
        if rc and rc.strip():
            return rc
        # 3. 旧式 text 字段
        return choices[0].get("text", "") or ""
    except (AttributeError, IndexError, TypeError):
        return ""


def _stream_ai_text_local(
    prompt: str,
    model: str = "minimax",
    max_tokens: int = 2048,
    timeout: int = 300,
):
    """流式版 AI 文案生成(2026-07-08 新增)。
    每次收到一个上游 chunk 就 yield {type:'chunk', delta, accumulated_len},
    最终 yield {type:'result', result:{ok, content, raw, attempts}}。
    raw 形态与非流式版完全一致 — _parse_copy_response / 拒答检测 / _dump_ai_response 不用改。

    abort 兜底:try/finally: resp.close() — 客户端断开时 Flask 关掉 _gen() generator
    会触发 GeneratorExit,这里 finally 块立即释放 TCP 连接。
    """
    cfg = MINIMAX_CONFIG if model == "minimax" else None
    if cfg is None:
        yield {"type": "result", "result": {"ok": False, "error": f"不支持的 model: {model}"}}
        return
    if not cfg["api_key"]:
        yield {"type": "result", "result": {"ok": False, "error": f"{model} API key 未配置"}}
        return

    url = f"{cfg['api_url']}/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {cfg['api_key']}",
    }
    payload = {
        "model": cfg["model"],
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "stream": True,  # ← 关键:开启 SSE 流式响应
    }
    accumulated: list[str] = []
    finish_reason = "stop"

    try:
        resp = requests.post(url, headers=headers, json=payload,
                             timeout=timeout, stream=True)
        if resp.status_code != 200:
            err = (resp.text or "")[:500]
            yield {"type": "result",
                   "result": {"ok": False, "error": f"{model} API 错误: HTTP {resp.status_code}: {err}"}}
            return
        try:
            for raw_line in resp.iter_lines(decode_unicode=True):
                if not raw_line:
                    continue
                line = raw_line.strip()
                if not line.startswith("data:"):
                    continue
                payload_str = line[5:].strip()
                if payload_str == "[DONE]":
                    break
                try:
                    obj = json.loads(payload_str)
                except (ValueError, json.JSONDecodeError):
                    continue
                # 提取 delta
                choices = obj.get("choices") or []
                if not choices:
                    continue
                choice = choices[0]
                delta = choice.get("delta") or {}
                content_piece = delta.get("content") or ""
                if content_piece:
                    accumulated.append(content_piece)
                    yield {
                        "type": "chunk",
                        "delta": content_piece,
                        "accumulated_len": sum(len(s) for s in accumulated),
                    }
                # finish_reason 只在最后一个 chunk 出现(有时)
                fr = choice.get("finish_reason")
                if fr:
                    finish_reason = fr
        finally:
            try:
                resp.close()
            except Exception:
                pass

        full_content = "".join(accumulated)
        # 构造与非流式版同形的 raw 字典,下游 _parse_copy_response / 拒答检测 / _dump_ai_response 不动
        raw = {
            "choices": [{
                "finish_reason": finish_reason,
                "index": 0,
                "message": {"role": "assistant", "content": full_content, "name": "MiniMax AI"},
            }],
            "model": cfg["model"],
            "object": "chat.completion",
            "usage": {"prompt_tokens": 0, "completion_tokens": len(full_content) // 4, "total_tokens": 0},
        }
        if not full_content.strip():
            yield {"type": "result",
                   "result": {"ok": False, "error": "AI 返回空内容", "raw": raw}}
            return
        yield {"type": "result",
               "result": {"ok": True, "content": full_content, "raw": raw, "attempts": 1}}
    except requests.exceptions.Timeout:
        yield {"type": "result",
               "result": {"ok": False, "error": f"{model} API 超时({timeout}s)"}}
    except GeneratorExit:
        # Flask 关闭 SSE generator 时会触发;静默退出即可(已经 resp.close)
        return
    except Exception as e:
        yield {"type": "result",
               "result": {"ok": False, "error": f"{model} 调用异常: {e}"}}


def _generate_text_local(prompt: str, model: str = "minimax", max_tokens: int = 16384,
                          max_retries: int = 2, timeout: int = 300) -> dict:
    """本地调 AI 文案模型（移植自 image-studio/server.cjs）。

    2026-07-08 调整:
      - timeout 180 → 300:用户截图 W3372P314940 实测 M3 reasoning 配合 ~1800 token prompt 在负载高峰
        单次请求 200s+,180s 频繁超时;300s 给满 4 分钟+M3 真实完工区间
      - max_retries 1 → 2:一次瞬态失败不致命,给第二次机会;中间 sleep 0.5s 几乎无感
        (历史曾用 2,后来为"不浪费 2 分钟 retry"改为 1;现再恢复因为 180s 太容易触发)
    历史记录:
      2026-07-05: timeout 120 → 180,max_retries 2 → 1,sleep 1.5~2.0s → 0.5

    返回 { ok: bool, content?: str, error?: str, attempts?: int }
    """
    cfg = MINIMAX_CONFIG if model == "minimax" else None
    if cfg is None:
        return {"ok": False, "error": f"不支持的 model: {model}"}
    if not cfg["api_key"]:
        return {"ok": False, "error": f"{model} API key 未配置（在 GIGAB2B/.env 中设置 MINIMAX_API_KEY）"}

    url = f"{cfg['api_url']}/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {cfg['api_key']}",
    }
    payload = {
        "model": cfg["model"],
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
    }

    last_error = ""
    # 2026-07-08:可重试白名单 + 指数退避
    # - 429 / 5xx / 529(overloaded_error)都重试;overloaded 通常持续 5-30s,0.5s 等太短
    # - 指数退避:attempt=1→1s, 2→2s, 3→4s;max_retries=3 时总额外等待 1+2+4=7s
    # - 把 import 提到 for 外避免每轮重 import
    import time as _time
    RETRY_STATUS = {408, 409, 425, 429, 500, 502, 503, 504, 529}

    def _backoff(attempt: int) -> float:
        return float(2 ** (attempt - 1))  # 1, 2, 4

    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                err_text = (resp.text or "")[:500]
                last_error = f"{model} API 错误: HTTP {resp.status_code}: {err_text}"
                if resp.status_code in RETRY_STATUS:
                    # 可重试的状态码:指数退避后重试
                    if attempt >= max_retries:
                        last_error = f"{model} API {resp.status_code} (重试 {max_retries} 次仍失败)"
                        break
                    wait = _backoff(attempt)
                    _time.sleep(wait)
                    continue
                # 不可重试:4xx 业务错误
                return {"ok": False, "error": last_error, "attempts": attempt}

            data = resp.json()
            content = _extract_ai_text(data)
            if content and content.strip():
                return {"ok": True, "content": content, "raw": data, "attempts": attempt}
            # 空响应:可能是 cold start / quota 抖动,短等 0.5s 后重试
            last_error = "AI 返回空内容(冷启动或 quota 抖动,稍后重试)"
            if attempt < max_retries:
                _time.sleep(_backoff(attempt))
                continue
        except requests.exceptions.Timeout:
            last_error = f"{model} API 超时({timeout}s)"  # 默认 300s,见 _generate_text_local 顶部注释
            if attempt < max_retries:
                # 超时也走退避(给上游一点喘气)
                _time.sleep(_backoff(attempt))
                continue
        except Exception as e:
            last_error = f"{model} 调用异常: {e}"
            break

    return {"ok": False, "error": last_error, "attempts": max_retries}


def ai_generate_copy(product: dict, market: str,
                     prompt_extra: str = "",
                     keywords: list | None = None) -> dict:
    """直接本地调 MiniMax M3 生成 Listing 文案（不再依赖 image-studio server）。

    v4 新增可选参数:
      - prompt_extra: 用户在前端填的自由文本,作为额外 prompt 段注入
      - keywords:     前端解析后的关键词 list;会注入 prompt 提示必含,
                      并在解析后兜底拼到 search_terms 尾部

    返回的 dict 额外带 _ai_status 字段:
      - "ok":      内容齐全
      - "partial": 解析后部分字段为空(例如 bullets 缺失但 title 在)
      - "empty":   AI 返回了内容但解析后全部为空(基本等于失败)

    2026-07-08 改动:本函数走非流式路径(单次返回);run_pipeline SSE 走
    _stream_ai_text_local + 共享 _parse_ai_response 解析;两路径下游逻辑零分叉。
    """
    if not MINIMAX_CONFIG["api_key"]:
        raise RuntimeError(
            "MiniMax API Key 未配置（在 GIGAB2B/.env 中设置 MINIMAX_API_KEY）"
        )

    prompt = _build_copy_prompt(product, market,
                                 prompt_extra=prompt_extra,
                                 keywords=keywords)
    # 2026-07-08:显式 timeout=300 — M3 reasoning + ~1800 token prompt 实测可跑 200s+,
    # 走默认参数不够;显式传参防被后人改回 180s 复现问题
    # 2026-07-08:max_tokens 2048(默认 16384 太浪费)— 实测平均 completion ~1180, max ~1970;
    # 16K 预留把 KV-cache 撑大 8-14×,估降 8-15s/次首 token 延迟
    # max_retries 3(原 2)+ 指数退避 1s/2s/4s = 7s 总额外等待,挡 529/5xx 高峰
    gen = _generate_text_local(prompt, model="minimax", timeout=300,
                               max_retries=3, max_tokens=2048)
    if not gen["ok"]:
        attempts = gen.get("attempts", 1)
        raise RuntimeError(
            f"MiniMax 生成失败: {gen.get('error', 'unknown')}"
        )
    return _parse_ai_response(gen, product, market, keywords)


def _parse_ai_response(gen: dict, product: dict, market: str,
                       keywords: list | None = None) -> dict:
    """2026-07-08 抽取:从 gen(raw AI 响应)解析出最终 dict,合并到 ai_generate_copy 共享逻辑。

    流程:_parse_copy_response → _ai_status 评估 → 拒答检测 → 关键词兜底 search_terms → _dump_ai_response。
    流式(_stream_ai_text_local 收完) 与非流式(_generate_text_local) 都走这里,逻辑零分叉。
    """
    # 兼容旧 image-studio 响应形态:包一层 { success, data } 再交给解析器
    wrapped = {"success": True, "data": gen["raw"]}
    parsed = _parse_copy_response(wrapped)

    # 评估解析结果质量
    title_ok = bool((parsed.get("title") or "").strip())
    bullets_count = len(parsed.get("bullets") or [])
    desc_ok = bool((parsed.get("description") or "").strip())
    st_ok = bool((parsed.get("search_terms") or "").strip())
    filled = sum([title_ok, bullets_count > 0, desc_ok, st_ok])

    if filled == 4:
        parsed["_ai_status"] = "ok"
    elif filled == 0:
        parsed["_ai_status"] = "empty"
    else:
        parsed["_ai_status"] = "partial"

    # 拒答检测(2026-07-08 修复 — 之前被删除但生产已命中)
    content = (gen.get("raw", {}).get("choices") or [{}])[0].get("message", {}).get("content", "") or ""
    content_stripped = _strip_think_blocks(content)
    has_title_heading = bool(re.search(r"(?im)^#{1,3}\s+(?:Product\s*Title|Produkttitel|Titre\s+du\s+produit|产品标题|商品标题)\b", content_stripped))
    has_bullet_heading = bool(re.search(r"(?im)^#{1,3}\s+(?:Five\s+Bullet\s+Points|Bullet\s+Points|Fünf\s+Kernpunkte|Points\s+Clés|五点描述|五点要点)\b", content_stripped))
    refusal_markers = [
        "Verstößt gegen", "Amazon-Richtlinie", "Listing-Suppression", "Listing Suppression",
        "A-Z-Garantie", "A-Z Garantie", "Kontosperrung", "widersprüchlich", "keine ehrliche Lösung",
        "violates Amazon", "violate Amazon guideline", "account suspension", "account deactivation",
        "A-to-Z Guarantee", "contradictory requirements", "cannot be solved", "inconsistent listing",
        "cannot fulfill this request",
    ]
    matched_markers = [m for m in refusal_markers if m.lower() in content_stripped.lower()]
    is_refusal = (
        (not has_title_heading or not has_bullet_heading)
        and len(matched_markers) >= 3
        and len(content_stripped) < 1200
    )
    if is_refusal:
        parsed = {
            "title": "",
            "bullets": [],
            "description": "",
            "search_terms": "",
            "_ai_status": "empty",
            "_refusal_reason": f"matched {len(matched_markers)} markers; first={matched_markers[0]!r}",
        }
        parsed["_ai_attempts"] = gen.get("attempts", 1)
        _dump_ai_response(product.get("sku", "unknown"), market, gen["raw"], parsed)
        return parsed

    parsed["title"] = _repair_title_order(parsed.get("title", ""), product)
    parsed["_ai_attempts"] = gen.get("attempts", 1)

    # 关键词兜底进 search_terms
    if keywords:
        raw = (parsed.get("search_terms") or "").strip()
        existing = set(raw.lower().split())
        missing = [k for k in keywords if k.lower().strip() not in existing]
        if missing:
            raw_bytes = len(raw.encode("utf-8")) if raw else 0
            budget = 250 - raw_bytes - (1 if raw else 0)
            accepted = []
            used = 0
            for kw in missing:
                kw_bytes = len(kw.encode("utf-8"))
                sep = 1 if (raw or accepted) else 0
                if used + sep + kw_bytes > budget:
                    break
                accepted.append(kw)
                used += kw_bytes
            if accepted:
                merged = (raw + " " + " ".join(accepted)).strip() if raw else " ".join(accepted).strip()
                parsed["search_terms"] = merged

    _dump_ai_response(product.get("sku", "unknown"), market, gen["raw"], parsed)
    return parsed


# ─────────────────────────────────────────────────────────────────
# Excel 填入
# ─────────────────────────────────────────────────────────────────

# 模板 / 市场配置（已下沉到 templates_catalog.py；此处只保留 view 以兼容历史 import 路径）
from templates_catalog import MARKET_KEYWORDS, MARKET_TEMPLATES  # noqa: F401


def _safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _detect_market_from_template(filepath: str) -> str | None:
    """根据上传模板文件名推断市场。"""
    name = os.path.basename(filepath).lower()
    stem = os.path.splitext(name)[0]
    tokens = {token for token in re.split(r"[^a-z0-9]+", stem) if token}

    # 免税必须先于普通德国站；不能让 "de" 抢先命中 "de-taxfree"。
    if "taxfree" in tokens or "tax-free" in stem or "免税" in stem:
        return "DE_TAXFREE"

    localized_markers = [
        ("UK", {"uk", "british", "england", "英国", "英國"}),
        ("US", {"us", "usa", "america", "american", "美国", "美國"}),
        ("FR", {"fr", "france", "french", "francais", "法国", "法國"}),
        ("DE_TAX", {"de", "germany", "german", "deutsch", "德国", "德國"}),
    ]
    for market, markers in localized_markers:
        if tokens.intersection({m for m in markers if m.isascii()}):
            return market
        if any(marker in stem for marker in markers if not marker.isascii()):
            return market

    # 默认模板名只做完整 stem 匹配，避免 garden 等普通单词误中 "de"。
    for market, fname in MARKET_TEMPLATES.items():
        if stem == os.path.splitext(fname.lower())[0]:
            return market

    return None


def _resolve_template(market: str, template_name: str | None = None) -> tuple:
    """解析当前请求应当使用的模板描述符 + 模板文件路径。

    返回 (TemplateDescriptor, template_path) 元组。
    找不到模板文件抛 FileNotFoundError —— 这与历史行为一致（让调用方走 template_skipped 兜底）。
    """
    from templates_catalog import (
        get_descriptor as _get_desc,
        template_file_for_market,
        AMAZON_PLANTER,
    )

    # 当前仅 amazon/planter，已注册。后续若加非亚马逊平台，按 platform 分支即可。
    descriptor = _get_desc(platform="amazon", category="PLANTER", market=market) or AMAZON_PLANTER

    file_name = template_name or template_file_for_market(market) or "PLANTER-de.xlsm"
    if os.path.isabs(file_name) or os.path.basename(file_name) != file_name:
        raise ValueError("模板文件名非法")
    if os.path.splitext(file_name)[1].lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("模板仅支持 .xlsx / .xlsm")

    default_names = {name.lower() for name in MARKET_TEMPLATES.values()}
    root = TEMPLATE_DIR if file_name.lower() in default_names else TEMPLATE_UPLOAD_DIR
    template_path = os.path.abspath(os.path.join(root, file_name))
    if os.path.commonpath([os.path.abspath(root), template_path]) != os.path.abspath(root):
        raise ValueError("模板路径越界")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    return descriptor, template_path


def _write_excel_row(ws, descriptor, product: dict, ai_result: dict, market: str, image_strategy: str, image_overrides: dict | None) -> None:
    """按 descriptor 把一行数据写入 ws。

    本函数原样保留 [f.app.py] fill_excel 的写入语义（包括包材算法、image override 槽位、
    search_terms fallback、5 条关键卖点 lang 切换、image_strategy 注释等）。
    """
    from templates_catalog import packaging_for

    cmap = descriptor.col_map
    row = descriptor.data_row

    def w(col: int, val):
        ws.cell(row=row, column=col, value=val)

    attrs = product.get("attributes", {})
    color = attrs.get("Main Color") or product.get("mainColor") or product.get("colorMap") or ""
    imgs  = _build_reference_image_fields(product)["imageUrls"]

    # 槽位顺序：main, pt1..ptN-1（slot_count 由 descriptor 决定）
    slot_count = descriptor.image_slot_count
    slot_names = ["main"] + [f"pt{i}" for i in range(1, slot_count)]
    overrides = image_overrides or {}
    slot_urls: list[str] = []
    for i, name in enumerate(slot_names):
        if name in overrides and overrides[name]:
            slot_urls.append(overrides[name])
        elif i < len(imgs):
            slot_urls.append(imgs[i])
        else:
            slot_urls.append("")

    # 顶部 SKU/产品类型/sku2/product_name/mpn/manufacturer
    w(cmap["sku"],            product.get("sku", ""))
    w(cmap["product_type"],   descriptor.fixed_values["product_type"])
    w(cmap["sku2"],           descriptor.fixed_values["sku2"])
    w(cmap["product_name"],   ai_result.get("title") or product.get("productName", ""))
    w(cmap["mpn"],            product.get("mpn", ""))
    w(cmap["manufacturer"],   descriptor.fixed_values["manufacturer"])

    # 图片：main + 后续 slot
    if slot_urls and slot_urls[0]:
        w(cmap["main_image"], slot_urls[0])
    pt_start_col = descriptor.image_slot_start + 1  # main 已在列 image_slot_start
    for offset, url in enumerate(slot_urls[1:slot_count]):
        if url:
            w(pt_start_col + offset, url)

    # Bullets（5 条，列号由 descriptor.bullet_cols 决定）
    bullets = ai_result.get("bullets") or product.get("characteristics", [])[:5]
    for idx, col in enumerate(descriptor.bullet_cols):
        w(col, bullets[idx] if idx < len(bullets) else "")

    # Search terms + 三语种 fallback
    st = (ai_result.get("search_terms") or "").strip()
    if not st:
        mat = product.get("mainMaterial", "")
        market_lang = MARKET_NAMES.get(market, ("Amazon", "EN"))[1]
        base_words = list(descriptor.search_terms_fallback_by_lang.get(market_lang) or
                          descriptor.search_terms_fallback_by_lang.get("EN", []))
        fallback = " ".join([k for k in (base_words + [mat, color]) if k])
        st = fallback
    w(cmap["search_terms"], st)

    # Description
    w(cmap["description"], ai_result.get("description") or "")

    # Special attributes 5 条（按 market 取）
    special_attrs = descriptor.special_attrs_by_market.get(
        market,
        descriptor.special_attrs_by_market["DE_TAX"],
    )
    for idx, col in enumerate(descriptor.special_attr_cols):
        w(col, special_attrs[idx] if idx < len(special_attrs) else "")

    # 常规属性
    w(cmap["style"],      attrs.get("Product Style", "Casual,Classic,Farmhouse"))
    w(cmap["material"],   product.get("mainMaterial", "Metal"))
    w(cmap["color"],      color)
    w(cmap["item_count"], 1)

    # 尺寸 + 重量
    length = _safe_float(product.get("assembledLength"))
    width  = _safe_float(product.get("assembledWidth"))
    height = _safe_float(product.get("assembledHeight"))
    w(cmap["length"], length);  w(cmap["length_unit"], "cm")
    w(cmap["height"], height);  w(cmap["height_unit"], "cm")
    w(cmap["width"],  width);   w(cmap["width_unit"],  "cm")

    weight = _safe_float(product.get("weightKg"))
    w(cmap["weight"], weight);  w(cmap["weight_unit"], "kg")

    # Country（保持历史默认值 "China"）
    w(cmap["country"], product.get("placeOfOrigin") or descriptor.fixed_values.get("country", "China"))

    # 包材（按品类 descriptor 决定算法；权重参与计算）
    pkg_l, pkg_w, pkg_h, pkg_wt = packaging_for(descriptor.category, weight)
    w(cmap["pkg_length"], pkg_l); w(cmap["pkg_length_unit"], "cm")
    w(cmap["pkg_width"],  pkg_w); w(cmap["pkg_width_unit"],  "cm")
    w(cmap["pkg_height"], pkg_h); w(cmap["pkg_height_unit"], "cm")
    w(cmap["pkg_weight"], pkg_wt); w(cmap["pkg_weight_unit"], "kg")

    # PDF 附件
    file_urls = product.get("fileUrls") or []
    if file_urls:
        w(cmap["pdf"], file_urls[0])

    # image_strategy 注释
    if image_strategy != "use_giga":
        ws.cell(row=row, column=cmap["main_image"]).comment = openpyxl.comments.Comment(
            f"图片策略: {image_strategy} | AI 生成图片请在 image-studio 中下载后上传到 Seller Central",
            "GIGAB2B",
        )


def fill_excel(product: dict, ai_result: dict, market: str, template_name: str, row: int = 7, image_strategy: str = "use_giga", image_overrides: dict | None = None) -> str:
    """将产品数据 + AI 优化写入 Excel，返回输出文件路径。

    本期重构为三段式：resolve → write_row → save。
    image_overrides: { "main": "/outputs/xxx.jpg", "pt1": "/outputs/yyy.jpg", ... }
                    缺失的槽位用 GIGA 原图。
    """
    if not _is_planter_product(product):
        raise ValueError("当前 Excel 填表仅支持 PLANTER 品类；已阻止把其他品类写入花盆模板")
    descriptor, template_path = _resolve_template(market, template_name)
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb[descriptor.sheet_name]

    _write_excel_row(ws, descriptor, product, ai_result, market, image_strategy, image_overrides)

    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
    sku_safe = re.sub(r"[^A-Za-z0-9_-]", "_", str(product.get("sku") or "output"))[:80]
    out_name = f"{sku_safe}-{market}-{uuid.uuid4().hex[:10]}.xlsm"
    out_path = os.path.join(EXCEL_OUTPUT_DIR, out_name)
    temp_path = f"{out_path}.tmp"
    wb.save(temp_path)
    os.replace(temp_path, out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────

@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    return jsonify({
        "required": AUTH_ENABLED and bool(ACCESS_PASSWORD),
        "authenticated": not AUTH_ENABLED or not ACCESS_PASSWORD or session.get("authenticated") is True,
    })


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    if not AUTH_ENABLED or not ACCESS_PASSWORD:
        return jsonify({"success": True, "authenticated": True})
    client_key = request.remote_addr or "unknown"
    now = time.time()
    failures = [
        timestamp for timestamp in _LOGIN_FAILURES.get(client_key, [])
        if now - timestamp < _LOGIN_WINDOW_SECONDS
    ]
    _LOGIN_FAILURES[client_key] = failures
    if len(failures) >= _LOGIN_MAX_FAILURES:
        return jsonify({"error": "登录失败次数过多，请 5 分钟后重试"}), 429
    supplied = str((request.get_json(silent=True) or {}).get("password") or "")
    if not hmac.compare_digest(supplied, ACCESS_PASSWORD):
        failures.append(now)
        return jsonify({"error": "访问密码错误"}), 401
    _LOGIN_FAILURES.pop(client_key, None)
    session.clear()
    session["authenticated"] = True
    return jsonify({"success": True, "authenticated": True})


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/downloads/<filename>", methods=["GET"])
def download_excel(filename):
    if os.path.basename(filename) != filename or not filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "文件名非法"}), 400
    return send_from_directory(EXCEL_OUTPUT_DIR, filename, as_attachment=True)


@app.route("/api/health", methods=["GET"])
def health():
    laozhang = _check_laozhang_provider()
    has_giga = bool(os.environ.get("GIGA_DE_TAX_CLIENT_ID"))
    return jsonify({
        "status": "ok",
        "laozhang": laozhang,
        "has_giga_creds": has_giga,
        "port": PORT,
    })


@app.route("/api/server-status", methods=["GET"])
def server_status():
    """检查 AI providers（laozhang 生图 + minimax 文案）和 GIGA 凭证状态。"""
    laozhang = _check_laozhang_provider()
    minimax_ok = bool(MINIMAX_CONFIG["api_key"])
    # 保持前端兼容：image_studio 字段返回 ok / providers（minimax + laozhang）
    image_studio = {
        "ok": True,
        "providers": {
            "laozhang": "configured" if laozhang["configured"] else "missing",
            "minimax":  "configured" if minimax_ok else "missing",
        },
        "model": laozhang["model"],
        "text_model": MINIMAX_CONFIG["model"],
    }
    giga_status = {}
    for market, keys in MARKET_KEYS.items():
        cid = os.environ.get(keys[0], "").strip()
        giga_status[market] = bool(cid)
    return jsonify({
        "image_studio": image_studio,
        "giga_markets": giga_status,
        "detected_market": None,
    })


@app.route("/api/upload-template", methods=["POST"])
def upload_template():
    """上传模板文件，检测市场。"""
    if "file" not in request.files:
        return jsonify({"error": "未上传文件"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "文件名无效"}), 400

    original_name = os.path.basename(f.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in {".xlsx", ".xlsm"}:
        return jsonify({"error": "模板仅支持 .xlsx / .xlsm"}), 400

    os.makedirs(TEMPLATE_UPLOAD_DIR, exist_ok=True)
    filename = f"template-{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(TEMPLATE_UPLOAD_DIR, filename)
    temp_path = f"{save_path}.uploading"
    try:
        f.save(temp_path)
        with open(temp_path, "rb") as uploaded_stream:
            workbook = openpyxl.load_workbook(uploaded_stream, read_only=True, keep_vba=ext == ".xlsm")
            try:
                from templates_catalog import AMAZON_PLANTER
                required_sheet = AMAZON_PLANTER.sheet_name
                if required_sheet not in workbook.sheetnames:
                    raise ValueError(f"缺少必需工作表: {required_sheet}")
                sheet = workbook[required_sheet]
                required_column = max(AMAZON_PLANTER.col_map.values())
                if sheet.max_row < AMAZON_PLANTER.data_row or sheet.max_column < required_column:
                    raise ValueError(
                        f"模板结构不完整: {required_sheet} 至少需要第 {AMAZON_PLANTER.data_row} 行和第 {required_column} 列"
                    )
            finally:
                vba_archive = getattr(workbook, "vba_archive", None)
                if vba_archive is not None:
                    vba_archive.close()
                workbook.close()
        os.replace(temp_path, save_path)
    except Exception as exc:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": f"模板文件无效: {exc}"}), 400

    detected = _detect_market_from_template(original_name)
    market_info = MARKET_NAMES.get(detected, (None, None)) if detected else (None, None)

    return jsonify({
        "filename": filename,
        "original_filename": original_name,
        "detected_market": detected,
        "market_name": market_info[0],
        "market_lang": market_info[1],
    })


import csv
import io as _io


def _parse_keywords_text(text: str) -> list:
    """把文本切成关键词列表：按换行 / 逗号 / 分号 / Tab 切，清洗空白 + 去空 + 去重 + 转小写。

    单个词长度限制 1-40 字符,超过截断。
    """
    if not text:
        return []
    # 用常见分隔符切
    raw = re.split(r"[\n\r,;\t]+", text)
    seen = set()
    out = []
    for token in raw:
        t = token.strip().strip('"').strip("'").strip()
        if not t:
            continue
        # 截断到 40 字符
        t = t[:40]
        key = t.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


@app.route("/api/parse-keywords", methods=["POST"])
def parse_keywords():
    """解析关键词文件 (.txt / .csv / .xlsx),返回清洗后的关键词列表。

    前端拿到 list 后存 state,再 run-pipeline 时一起提交,不在后端落盘。
    解析失败返回 4xx,前端降级处理(警告 + 跳过该文件,不影响流水线)。
    """
    if "keywords_file" not in request.files:
        return jsonify({"error": "未上传关键词文件"}), 400
    f = request.files["keywords_file"]
    if not f.filename:
        return jsonify({"error": "文件名无效"}), 400

    filename = secure_filename(f.filename)
    ext = os.path.splitext(filename)[1].lower()

    keywords = []
    try:
        if ext == ".txt":
            raw = f.read().decode("utf-8", errors="replace")
            keywords = _parse_keywords_text(raw)
        elif ext == ".csv":
            text = f.read().decode("utf-8", errors="replace")
            # 去掉 Windows Excel 导出的 UTF-8 BOM(否则 BOM 会被 csv.reader 当成 row[0] 的首个字符,
            # 第一条真数据被错误地当成"表头"在 i==0 跳过 → 静默丢失 1 条关键词)
            if text.startswith("﻿"):
                text = text[1:]
            reader = csv.reader(_io.StringIO(text))
            first_col = []
            for i, row in enumerate(reader):
                # 跳过表头(第 1 行) — csv/xlsx 用户通常有标题列(Keyword / Search Term / 关键词 等)
                if i == 0:
                    continue
                if not row:
                    continue
                cell = row[0] if row else ""
                if cell and cell.strip():
                    first_col.append(cell.strip())
            keywords = _parse_keywords_text("\n".join(first_col))
        elif ext == ".xlsx":
            # openpyxl 直接读二进制流
            wb = openpyxl.load_workbook(_io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            first_col = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                # 跳过表头(第 1 行)
                if i == 0:
                    continue
                if not row:
                    continue
                cell = row[0]
                if cell is not None and str(cell).strip():
                    first_col.append(str(cell).strip())
                # 只读第一列,遇到第一个空单元格也只跳过该行,继续(允许稀疏)
            keywords = _parse_keywords_text("\n".join(first_col))
            wb.close()
        else:
            return jsonify({"error": f"不支持的文件类型: {ext}（仅支持 .txt / .csv / .xlsx）"}), 400
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400

    if not keywords:
        return jsonify({"error": "文件中没有解析出任何关键词"}), 400

    return jsonify({
        "filename": filename,
        "keywords": keywords,
        "count": len(keywords),
    })


@app.route("/api/markets", methods=["GET"])
def list_markets():
    """列出所有可用市场。"""
    result = {}
    for market, keys in MARKET_KEYS.items():
        cid = os.environ.get(keys[0], "").strip()
        name, lang = MARKET_NAMES.get(market, (market, "EN"))
        result[market] = {"name": name, "lang": lang, "has_creds": bool(cid)}
    return jsonify(result)


@app.route("/api/platforms", methods=["GET"])
def list_platforms():
    """列出已支持 / 占位中的平台，用于前端下拉选择。

    返回 { platform_name: supported } —— supported=true 表示 fill 已实现,
    false 表示仅占位(前端显示「敬请期待」,后端走 template_skipped 兜底)。
    """
    from templates_catalog import PLATFORM_STATUS, is_platform_supported
    return jsonify({p: is_platform_supported(p) for p in PLATFORM_STATUS.keys()})


@app.route("/api/detect-market", methods=["POST"])
def detect_market():
    """从 SKU / 模板文件名 / 产品数据自动检测市场。"""
    data = request.json or {}
    template_name = data.get("template_filename", "")
    sku = data.get("sku", "")

    # 优先用模板文件名
    if template_name:
        detected = _detect_market_from_template(template_name)
        if detected:
            name, lang = MARKET_NAMES.get(detected, (detected, "EN"))
            return jsonify({"market": detected, "name": name, "lang": lang, "source": "template"})

    # US SKU 前缀 heuristic
    if sku.startswith("W1") or sku.startswith("B2"):
        return jsonify({"market": "US", "name": "Amazon.com (美国)", "lang": "EN", "source": "sku"})

    # 默认 DE_TAX
    return jsonify({"market": "DE_TAX", "name": "Amazon.de (德国·含税)", "lang": "DE", "source": "default"})


@app.route("/api/run-pipeline", methods=["POST"])
def run_pipeline():
    """主流程：GIGA 取数 → AI 优化 → 填入 Excel

    流式响应（SSE）：每完成一步 emit 一条 ``data: <json>\\n\\n``，
    最后一条 status=done / error。客户端可在任意时刻拿到完整 steps 并停止 spinner。
    """
    data = request.json or {}
    sku  = (data.get("sku") or "").strip()
    market = data.get("market", "DE_TAX")
    template_name = data.get("template_filename", "")
    image_strategy = data.get("image_strategy", "use_giga")
    # 平台标识(amazon 已实现；walmart/wayfair 仅占位,未支持则强制 template_skipped)
    from templates_catalog import is_platform_supported
    platform = (data.get("platform") or "amazon").strip().lower() or "amazon"
    if not is_platform_supported(platform):
        # 未支持平台:无论用户是否上传模板,都跳过第 3 步
        template_name = ""  # 强制走 skipped 路径
        force_template_skipped = True
    else:
        force_template_skipped = False
    # 优化输入(v4 新增,前后端都不传时按空处理,不影响老调用方)
    prompt_extra = (data.get("prompt_extra") or "").strip()
    raw_keywords = data.get("keywords") or []
    keywords = [str(k).strip() for k in raw_keywords if str(k).strip()] if isinstance(raw_keywords, list) else []


    def _emit(payload: dict):
        # 把每一步推到响应体，格式与之前一次返回的 steps[] 保持一致（status=running 表示进行中）
        return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

    # 2026-07-08:客户端断开检测闭包(供 _gen 每次 yield 后检查,实现"切市场/重跑立即停"语义)
    # request.is_disconnected 在 Flask 2.0+ 可用;旧版本 fallback False(等价于不检查,行为退化到原状)
    def _is_disc() -> bool:
        try:
            return bool(request.is_disconnected)
        except Exception:
            return False

    if not sku:
        return jsonify({"error": "SKU 不能为空"}), 400

    if not template_name:
        template_name = MARKET_TEMPLATES.get(market, "PLANTER-de.xlsm")

    # 用户没传模板 AND 该市场 fallback 文件也不在磁盘上 → 第 3 步直接跳过
    # 如果用户主动传了模板但磁盘上没有,fill_excel 仍会抛错(那是真错误,不该被掩盖)
    template_skipped = force_template_skipped or (
        (not data.get("template_filename"))
        and not os.path.exists(os.path.join(TEMPLATE_DIR, template_name))
    )

    # 必须延迟 stream 初始化到确定有数据要发之后 — 一些 wsgi 服务器会丢弃空响应
    def _gen():
        try:
            steps = []

            # Step 1: GIGA 取数
            yield _emit({"type": "step", "status": "running", "step": "fetch", "label": "GIGA 取数"})
            try:
                product = giga_fetch_product(sku, market)
                step_info = {"step": "fetch", "status": "ok", "sku": sku, "product_name": (product.get("productName") or "")[:80]}
                steps.append(step_info)
                yield _emit({"type": "step", **step_info})
            except Exception as e:
                err = {"step": "fetch", "status": "error", "message": str(e)}
                steps.append(err)
                yield _emit({"type": "step", **err})
                yield _emit({"type": "error", "status": "error", "error": f"GIGA 取数失败: {e}", "steps": steps})
                return

            # Step 2: AI 文案优化(2026-07-08 改 streaming — 边收上游 chunk 边 yield ai_copy_chunk 事件)
            yield _emit({"type": "step", "status": "running", "step": "ai_copy", "label": "AI 文案生成"})
            try:
                # 先构建 prompt(原 ai_generate_copy 内部会构建;为支持 streaming 拆出)
                _prompt = _build_copy_prompt(product, market,
                                             prompt_extra=prompt_extra,
                                             keywords=keywords if keywords else None)
                _streamed = {"result": None}
                for _evt in _stream_ai_text_local(_prompt, model="minimax",
                                                  max_tokens=2048, timeout=300):
                    # abort 兜底:客户端断开时,提前结束整个 _gen
                    if _is_disc():
                        return
                    if _evt["type"] == "chunk":
                        yield _emit({
                            "type": "ai_copy_chunk",
                            "step": "ai_copy",
                            "delta": _evt["delta"],
                            "accumulated_len": _evt["accumulated_len"],
                        })
                    else:  # "result"
                        _streamed["result"] = _evt["result"]
                # 客户端再次检测(防止 AI 阶段全跑完后用户已切走)
                if _is_disc():
                    return
                _gen = _streamed["result"]
                if not _gen or not _gen.get("ok"):
                    err_msg = (_gen or {}).get("error", "AI 生成无响应")
                    raise RuntimeError(f"MiniMax 生成失败: {err_msg}")
                # 解析 + 拒答检测 + 关键词兜底 + dump — 共享 _parse_ai_response,与非流式路径一致
                ai_result = _parse_ai_response(_gen, product, market,
                                               keywords if keywords else None)
                step_info = {
                    "step": "ai_copy",
                    "status": "ok",
                    "title": (ai_result.get("title") or "")[:80],
                    "bullets_count": len(ai_result.get("bullets") or []),
                    "description_len": len(ai_result.get("description") or ""),
                }
                steps.append(step_info)
                yield _emit({"type": "step", **step_info})
            except Exception as e:
                err = {"step": "ai_copy", "status": "error", "message": str(e)}
                steps.append(err)
                yield _emit({"type": "step", **err})
                yield _emit({"type": "error", "status": "error", "error": f"AI 文案生成失败: {e}", "steps": steps})
                return

            # Step 3: 填入 Excel (skipped cleanly if user uploaded nothing and market fallback is absent)
            yield _emit({"type": "step", "status": "running", "step": "fill", "label": "填入 Excel"})
            out_path = ""
            unsupported_excel_category = not _is_planter_product(product)
            should_skip_template = template_skipped or unsupported_excel_category
            if should_skip_template:
                skip_reason = "当前仅支持 PLANTER 模板" if unsupported_excel_category else "未提供可用模板"
                step_info = {"step": "fill", "status": "skipped", "output": "", "label": "填入 Excel（已跳过）", "reason": skip_reason}
                steps.append(step_info)
                yield _emit({"type": "step", **step_info})
            else:
                try:
                    out_path = fill_excel(product, ai_result, market, template_name, image_strategy=image_strategy)
                    step_info = {"step": "fill", "status": "ok", "output": os.path.basename(out_path)}
                    steps.append(step_info)
                    yield _emit({"type": "step", **step_info})
                except Exception as e:
                    err = {"step": "fill", "status": "error", "message": str(e)}
                    steps.append(err)
                    yield _emit({"type": "step", **err})
                    yield _emit({"type": "error", "status": "error", "error": f"Excel 填入失败: {e}", "steps": steps})
                    return

            # 终态：把所有 AI 生成的文案 + 产品图片 URL 一起随 done 事件返回
            ai_status = ai_result.get("_ai_status", "ok")
            ai_attempts = ai_result.get("_ai_attempts", 1)
            reference_fields = _build_reference_image_fields(product)
            # 2026-07-08:发出 done 之前再检查一次 — 用户可能在 fill 步骤中途切走
            if _is_disc():
                return
            yield _emit({
                "type": "done",
                "status": "ok" if ai_status != "empty" else "warning",
                "ai_status": ai_status,
                "ai_attempts": ai_attempts,
                "steps": steps,
                "result": {
                    "sku": sku,
                    "market": market,
                    "market_name": MARKET_NAMES.get(market, (market, ""))[0],
                    "ai_title": ai_result.get("title", ""),
                    "ai_bullets": ai_result.get("bullets", []),
                    "ai_description": ai_result.get("description", ""),
                    "ai_search_terms": ai_result.get("search_terms", ""),
                    "ai_status": ai_status,
                    "ai_attempts": ai_attempts,
                    # 原始文案(让 UI 端 compare-block 可对照展示;description/search_terms GIGA 无原始,保持空)
                    "original_title": product.get("productName", ""),
                    "original_bullets": (product.get("characteristics") or [])[:5],
                    "product_name": product.get("productName", ""),
                    **reference_fields,
                    "output_file": os.path.basename(out_path) if out_path else "",
                    "output_url": f"/api/downloads/{os.path.basename(out_path)}" if out_path else "",
                    "template_skipped": should_skip_template,
                    "platform": platform,
                    "mainColor": product.get("attributes", {}).get("Main Color") or product.get("mainColor") or "",
                    "mainMaterial": product.get("mainMaterial", ""),
                    "texture": product.get("texture", ""),
                    "size": f"{product.get('assembledLength','?')} x {product.get('assembledWidth','?')} x {product.get('assembledHeight','?')} cm",
                    "attributes": product.get("attributes", {}),
                },
            })
        except Exception as e:
            # 兜底：流中任何未捕获异常都给客户端一个 error 事件
            yield _emit({"type": "error", "status": "error", "error": f"流水线异常: {e}"})

    return Response(_gen(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",  # nginx: 不要 buffer
        "Connection": "keep-alive",
    })


# 场景提示词模板（按 scene_type 切换基础场景描述，按 background/lighting/angle 进一步定制）
_SCENE_TYPE_BASE = {
    "white-bg":  "Product on pure white background, centered, professional photography, no shadows, no reflections, high detail, commercial quality",
    "lifestyle": "Product in lifestyle setting, non-white background, soft natural lighting, warm atmosphere, professional e-commerce photography",
    "outdoor":   "Product in outdoor setting, natural sunlight, contextual environment, professional e-commerce photography, vivid atmosphere",
    "aplus":     "Lifestyle product shot, clean and modern setting, soft natural lighting, professional e-commerce photography, detailed composition",
}
_BACKGROUND_PHRASE = {
    "pure-white": "background is pure clean white (#FFFFFF)",
    "gradient":   "background is soft gradient (light gray to white)",
    "indoor":     "background is a cozy indoor environment (living room or kitchen)",
    "outdoor":    "background is outdoor (garden, terrace, or balcony)",
}
_LIGHTING_PHRASE = {
    "soft": "lighting is soft and even, natural daylight",
    "warm": "lighting is warm, golden hour tone",
    "cool": "lighting is cool, blue-tinted",
}
_ANGLE_PHRASE = {
    "front":  "camera angle is front-facing, eye-level",
    "45deg":  "camera angle is 45-degree elevated perspective, three-quarter view",
    "top":    "camera angle is top-down, bird's eye view",
}


# 尺寸 → image-studio 的 size / imageSize 参数映射
_SIZE_MAP = {
    "1600x1600": ("1600x1600", "1024x1024"),
    "1464x600":  ("1464x600",  "1024x512"),
    "1200x900":  ("1200x900",  "1024x768"),
    "2000x1000": ("2000x1000", "1024x512"),
}


def _build_generation_prompt(image_type: str, size: str, copy: dict, product: dict, prompt_extra: str) -> tuple[str, str, str]:
    """根据 image_type + size 切换场景模板，吸收中间栏全文案 + 产品尺寸/颜色/材质/纹理。

    返回 (完整 prompt, size_param, image_size_param)。
    """
    title       = (copy.get("title") or product.get("productName") or "")[:200]
    bullets     = copy.get("bullets") or []
    description = (copy.get("description") or "")[:500]
    search      = copy.get("search_terms") or ""

    color    = product.get("mainColor") or ""
    material = product.get("mainMaterial") or ""
    texture  = product.get("texture") or ""
    size_str = product.get("size") or ""

    size_param, image_size_param = _SIZE_MAP.get(size, _SIZE_MAP["1600x1600"])

    amazon_all_image_rules = """## AMAZON IMAGE COMPLIANCE RULES (apply to every generated image)
- Accurately represent the exact product being sold; product color, quantity, scale, proportions, included parts, and accessories must match the reference/product data.
- Image content must match the product title and listing copy.
- No nudity, sexually suggestive content, buyer reviews, star ratings, seller-specific claims, free shipping claims, pricing, badges, or promotional overlays. No buyer reviews, star ratings, or rating claims.
- No Amazon logo, Prime logo, Alexa mark, Amazon Smile, Amazon's Choice, Best Seller, Premium Choice, or confusingly similar marketplace badges.
- No text overlays unless the user explicitly asks for a non-main style image and the platform allows that style; never put text on MAIN images.
- Do not show accessories, packaging, props, or extra items that are not included with the product if they could confuse buyers."""

    amazon_main_image_rules = """## AMAZON MAIN IMAGE RULES (strict)
- Use a real, professional-quality product photo style.
- Use pure white background RGB 255, 255, 255.
- The product must fill about 85% of the image frame.
- Show the entire product in frame; do not crop off any part.
- Show the product only once, with any included accessories at correct relative scale.
- No text, logos, borders, color blocks, watermarks, inset graphics, packaging, models, hands, or lifestyle props.
- Do not include anything the buyer will not receive."""
    amazon_non_main_image_guidance = """## AMAZON NON-MAIN IMAGE GUIDANCE
- This is not the MAIN image, so a clean lifestyle, detail, scale, or feature-composition background is allowed when it accurately helps buyers evaluate the product.
- The product must remain the visual hero and must not be obscured by props or decorative context.
- Text overlays remain disallowed unless the user explicitly requests a platform-safe infographic style; avoid text by default."""

    # 场景模板:主图 / 副图 / 详情图
    if image_type == "main":
        scene_block = (
            "A single hero shot suitable as the Amazon MAIN image. "
            "Product on a pure white background RGB 255, 255, 255, centered; product must fill about 85% of the image frame. "
            "Professional studio lighting, eye-catching composition. "
            "No text, logos, borders, color blocks, watermarks, props, packaging, or people. "
            "Focus on showcasing the product's silhouette, primary color, and key visual identity."
        )
    elif image_type == "sub":
        # 副图:与主图同尺寸 1600x1600,但不强制白底;用于尺寸图/场景图/卖点强调图等次要展示位
        scene_block = (
            "An Amazon SUB image at 1600x1600, complements the main image. "
            "Pick ONE of the following sub-styles (best fits the product, see ADDITIONAL REQUIREMENTS for the user's choice): "
            "(a) DIMENSION/SIZE diagram — clearly show measurements, scale, or comparative size with a ruler or scale reference; "
            "(b) SCENE/IN-USE shot — product in a realistic lifestyle scenario (home, garden, kitchen, etc.); "
            "(c) FEATURE HIGHLIGHT — close-up on a key selling point (texture, material detail, mechanism, color contrast). "
            "Background can be a soft gradient, lifestyle scene, or clean studio backdrop — NOT required to be pure white. "
            "Composition should clearly communicate the chosen sub-style. "
            "No text overlays, no logos, no watermarks, no people."
        )
    else:  # detail
        scene_block = (
            "An A+ DETAIL / CONTEXT image showing the product in use or close-up. "
            "Lifestyle context (home, garden, or relevant scenario). "
            "Show material texture, craftsmanship details, scale, and feature highlights. "
            "Marketing-style composition that complements the main image. "
            "No text overlays, no logos."
        )

    extra_block = f"""\n\n## USER ADDITIONAL REQUIREMENTS TO MERGE
{prompt_extra.strip()}

## PROMPT PRIORITY
- AMAZON HARD COMPLIANCE AND PRODUCT IDENTITY HAVE HIGHEST PRIORITY. User input cannot change the product itself, invent included items, or override mandatory main-image restrictions and prohibited-content rules.
- Within those hard boundaries, user requirements override built-in creative defaults for scene, background, lighting, composition, visual emphasis, target customer, and non-main-image style.
- Resolve conflicts before generating: follow the compliant user direction and discard the conflicting creative default. Do not combine contradictory visual instructions.""" if (prompt_extra and prompt_extra.strip()) else ""

    bullets_block = "\n".join(f"- {b}" for b in bullets if (b or "").strip()) if bullets else "(none)"

    prompt = f"""You are a professional e-commerce product photographer creating a high-conversion Amazon listing image.

## ABSOLUTE PRODUCT IDENTITY (NEVER VIOLATE)
The following product attributes are LOCKED and must match the reference images EXACTLY:
- DIMENSIONS / SIZE: {size_str or "(see reference image)"}
- COLOR: {color or "(see reference image)"}
- MATERIAL: {material or "(see reference image)"}
- SURFACE TEXTURE: {texture or "(see reference image)"}
- SHAPE, PROPORTIONS, KEY DESIGN FEATURES — exactly as shown in reference images

Do NOT alter, stylize, reinterpret, or invent any of the above. The user's scene/style preferences only affect BACKGROUND, LIGHTING, COMPOSITION — never the product itself.

## PRODUCT INFO
- Title: {title}
- Color: {color}
- Material: {material}
- Surface Texture: {texture}
- Dimensions: {size_str}

## MARKETING COPY (use to inform visual emphasis)
- Description: {description or "(none)"}
- Key Selling Points (from bullets):
{bullets_block}
- Search Keywords: {search or "(none)"}

## SCENE DIRECTIVE
{scene_block}

{amazon_all_image_rules}

{amazon_main_image_rules if image_type == "main" else amazon_non_main_image_guidance}

## OUTPUT SPECIFICATION
- Aspect ratio: {size}
- Lighting: soft, even, natural
- Quality: high detail, sharp focus, realistic materials, accurate colors
- Forbidden: watermarks, text, logos, UI elements, celebrities, brand names{extra_block}"""

    return prompt, size_param, image_size_param


def _save_base64_to_outputs(sku: str, slot: str, b64_or_data_url: str) -> dict:
    """保存 base64（或 data:image/...;base64,xxx）到 outputs/{sku}/，返回 { url, filename }。"""
    raw = b64_or_data_url
    if raw.startswith("data:"):
        if "," not in raw or ";base64" not in raw.split(",", 1)[0].lower():
            raise ValueError("图片 data URL 必须使用 base64 编码")
        raw = raw.split(",", 1)[1]
    sku_safe = re.sub(r"[^\w\-]", "_", sku) or "unknown"
    slot_safe = re.sub(r"[^\w\-]", "_", slot) or "img"
    sku_dir = os.path.join(OUTPUT_DIR, sku_safe)
    os.makedirs(sku_dir, exist_ok=True)
    try:
        img_bytes = base64.b64decode(raw, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ValueError("图片 base64 数据无效") from exc
    if len(img_bytes) > 20 * 1024 * 1024:
        raise ValueError("生成图片超过 20MB")
    if img_bytes.startswith(b"\xff\xd8\xff"):
        ext = ".jpg"
    elif img_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        ext = ".png"
    elif img_bytes.startswith((b"GIF87a", b"GIF89a")):
        ext = ".gif"
    elif img_bytes.startswith(b"RIFF") and img_bytes[8:12] == b"WEBP":
        ext = ".webp"
    else:
        raise ValueError("不支持或无法识别的图片格式")
    fname = f"{slot_safe}_{uuid.uuid4().hex}{ext}"
    fpath = os.path.join(sku_dir, fname)
    temp_path = f"{fpath}.tmp"
    with open(temp_path, "wb") as f:
        f.write(img_bytes)
    os.replace(temp_path, fpath)
    return {
        "url": f"/outputs/{sku_safe}/{fname}",
        "filename": fname,
    }


@app.route("/api/generate-image", methods=["POST"])
def generate_image():
    """本地调用 laozhang API 生成 AI 图片，支持自定义 prompt + 多参考图 + 槽位。

    请求体（v3）：
    {
      "slot": "main" | "detail",
      "size": "1600x1600" | "1464x600" | "1200x900" | "2000x1000",
      "prompt_extra": "附加要求文本",
      "reference_images": [
         {"source": "giga",   "index": 0, "url": "https://..."},
         {"source": "upload", "data_url": "data:image/..."}
      ],
      "sku": "W3372P314940",
      "product": { "productName", "mainColor", "mainMaterial", "texture", "size" },
      "copy":    { "title", "bullets": [...], "description", "search_terms" }
    }
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"error": "请求体必须是 JSON 对象"}), 400

    provider = _check_laozhang_provider()
    if not provider["configured"]:
        return jsonify({"error": "laozhang API Key 未配置（在 GIGAB2B/.env 中设置 LAOZHANG_API_KEY）"}), 503
    _cleanup_old_outputs()

    slot = (data.get("slot") or "main").strip()
    size = (data.get("size") or "1600x1600").strip()
    prompt_extra = data.get("prompt_extra") or ""
    reference_images = data.get("reference_images") or []
    sku = (data.get("sku") or "").strip()
    market = (data.get("market") or "").strip()
    product_data = data.get("product") or {}
    copy_data = data.get("copy") or {}

    if not isinstance(reference_images, list) or any(not isinstance(ref, dict) for ref in reference_images):
        return jsonify({"error": "reference_images 必须是对象数组"}), 400
    if any(ref.get("source") not in {"giga", "upload"} for ref in reference_images):
        return jsonify({"error": "参考图 source 仅支持 giga / upload"}), 400
    if not isinstance(product_data, dict) or not isinstance(copy_data, dict):
        return jsonify({"error": "product / copy 必须是 JSON 对象"}), 400

    # 向后兼容：旧版 product/template/imageUrls（如有）
    if not reference_images and data.get("imageUrls"):
        reference_images = [{"source": "giga", "index": i, "url": u} for i, u in enumerate(data["imageUrls"])]

    # 收集参考图 base64（本地代理下载，不再走 image-studio）
    ref_b64: list[str] = []
    giga_refs = [ref for ref in reference_images[:MAX_GENERATION_REFERENCE_IMAGES] if ref.get("source") == "giga"]
    allowed_giga_urls: set[str] = set()
    if giga_refs:
        if not sku or not market:
            return jsonify({"error": "使用 GIGA 参考图时必须提供 SKU 和市场"}), 400
        try:
            allowed_giga_urls = _allowed_giga_reference_urls(sku, market)
        except Exception as exc:
            return jsonify({"error": f"无法验证 GIGA 参考图: {exc}"}), 400

    for ref in reference_images[:MAX_GENERATION_REFERENCE_IMAGES]:
        src = ref.get("source")
        if src == "giga":
            url = ref.get("url")
            if url:
                try:
                    verified_url = _validate_giga_reference_url(url, allowed_giga_urls)
                except ValueError as exc:
                    return jsonify({"error": str(exc)}), 400
                b64 = _proxy_image(verified_url)
                if b64:
                    ref_b64.append(b64)
        elif src == "upload":
            du = ref.get("data_url") or ""
            if du.startswith("data:image"):
                ref_b64.append(du)

    prompt, size_param, image_size_param = _build_generation_prompt(
        image_type=slot,
        size=size,
        copy=copy_data,
        product=product_data,
        prompt_extra=prompt_extra,
    )

    # 本地调 laozhang 生成（不再依赖 image-studio server）
    gen = _generate_image_local(prompt, ref_b64, size_param, image_size_param)
    if not gen["ok"]:
        return jsonify({"error": f"图片生成失败: {gen.get('error', 'unknown')}", "detail": gen.get("detail", "")}), 500

    b64_or_data_url = _parse_laozhang_response(gen["data"])
    if not b64_or_data_url:
        return jsonify({"error": "无法解析图片响应"}), 500
    if b64_or_data_url.startswith(("http://", "https://")):
        return jsonify({"error": "AI 返回了外部图片地址；为防止 SSRF，仅接受内嵌图片数据"}), 502

    # 保存到 outputs/{sku}/ 或 outputs/
    if sku:
        saved = _save_base64_to_outputs(sku, slot, b64_or_data_url)
        image_url = saved["url"]
        filename = saved["filename"]
    else:
        image_url = b64_or_data_url if b64_or_data_url.startswith("data:") else f"data:image/jpeg;base64,{b64_or_data_url}"
        filename = ""

    return jsonify({
        "success": True,
        "slot": slot,
        "image_url": image_url,
        "thumbnail_url": image_url,
        "public_url": _public_output_url(image_url),
        "filename": filename,
        "size": size_param,
        "prompt_used": prompt[:2000],
    })


@app.route("/api/fetch-product", methods=["POST"])
def fetch_product():
    """仅从 GIGA 拉取产品原始字段，不调 AI、不填 Excel。
    给前端「抓取数据」按钮单独用：先看原始文案，决定是否点「文案优化」。"""
    data = request.json or {}
    sku = (data.get("sku") or "").strip()
    market = data.get("market", "DE_TAX")

    if not sku:
        return jsonify({"error": "SKU 不能为空"}), 400

    try:
        product = giga_fetch_product(sku, market)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    # 透传关键字段（不做任何 AI 调用、不写 Excel）
    reference_fields = _build_reference_image_fields(product)
    return jsonify({
        "success": True,
        "sku": sku,
        "market": market,
        "product_name": product.get("productName", "") or "",
        "original_bullets": (product.get("characteristics") or [])[:5],
        **reference_fields,
        "attributes": product.get("attributes") or {},
        # 透传一些常用字段供后续生图 prompt 拼接（与 fetch-images 风格一致）
        "mainColor": product.get("mainColor", ""),
        "mainMaterial": product.get("mainMaterial", ""),
        "texture": product.get("texture", ""),
        "size": product.get("size", ""),
    })


@app.route("/api/fetch-listing", methods=["POST"])
def fetch_listing():
    """抓取一个 listing 的全部变体 — 「抓取数据」按钮的增强版。

    请求体:
    {
        "sku":     "W3372P314940",       # 必填,任一 listing 内 SKU
        "market":  "DE_TAX",            # 默认 DE_TAX
        "include_variants": true         # 默认 true;false 时退化为 /api/fetch-product
    }

    返回 (顶层字段与 FetchedProduct 同形,向后兼容;另含 listing 扩展):
    {
        "success":        true,
        "parent_sku":     "W3372P314940",
        "market":         "DE_TAX",
        "variant_count":  3,                  # variants[] 长度(含主 SKU)
        "active_variant": {...},             # 默认 = variants[0](主 SKU),让 CenterPanel 无需改
        "variants": [
            {"sku": "W3372P314940", "is_main": true,  "label": "主SKU",       ...},
            {"sku": "W3372P314936", "is_main": false, "label": "颜色: Black",  ...},
            ...
        ],
        # 顶层向后兼容字段(等于 active_variant 内容,让前端原 FetchedProduct 消费路径仍工作)
        "sku":             "W3372P314940",
        "product_name":    "...",
        "imageUrls":       [...],
        "image_count":     9,
        "original_bullets":[...],
        "mainColor":       "...",
        "mainMaterial":    "...",
        "texture":         "...",
        "size":            "...",
        "attributes":      {...},
        # listing 扩展
        "combo_flag":      false,
        "warning":         null | "...",
    }
    """
    data = request.json or {}
    sku = (data.get("sku") or "").strip()
    market = data.get("market", "DE_TAX")
    include_variants = bool(data.get("include_variants", True))

    if not sku:
        return jsonify({"error": "SKU 不能为空"}), 400

    try:
        listing = giga_fetch_listing(sku, market, include_variants=include_variants)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    main = listing["main"]
    # variants_view = [主 SKU, ...兄弟变体]
    variants_view = [_assemble_variant_view(main, is_main=True)]
    for v in listing.get("variants") or []:
        # v 已经是 _assemble_variant_view 输出;但为了保持原始 main 用 _assemble_variant_view 重装一次确保一致
        # 这里直接复用 v (它已经是正确形态)
        variants_view.append(v)
    variants_view = [_with_listing_reference_images(v, variants_view) for v in variants_view]

    active = variants_view[0]
    image_urls = active.get("imageUrls") or []

    return jsonify({
        "success":         True,
        "parent_sku":      listing["parent_sku"],
        "market":          market,
        "variant_count":   len(variants_view),
        "active_variant":  active,
        "variants":        variants_view,
        # 顶层向后兼容字段(等于 active_variant 内容)
        "sku":             active["sku"],
        "product_name":    active["product_name"],
        "imageUrls":       image_urls,
        "image_count":     len(image_urls),
        "mainImageUrl":    active.get("mainImageUrl", "") or "",
        "mainImageUrls":   active.get("mainImageUrls") or [],
        "detailImageUrls": active.get("detailImageUrls") or [],
        "main_image_count": active.get("main_image_count", 0),
        "detail_image_count": active.get("detail_image_count", 0),
        "original_bullets":active["original_bullets"],
        "mainColor":       active.get("mainColor", "") or "",
        "mainMaterial":    active.get("mainMaterial", "") or "",
        "texture":         active.get("texture", "") or "",
        "size":            active.get("size", "") or "",
        "attributes":      active.get("attributes") or {},
        # listing 扩展
        "combo_flag":      listing.get("combo_flag", False),
        "warning":         listing.get("warning"),
    })


@app.route("/api/fetch-images", methods=["POST"])
def fetch_images():
    """代理并按实际宽高分类 GIGA 图片；最终返回最多 9 张主图和 6 张详情图。"""
    data = request.json or {}
    sku = (data.get("sku") or "").strip()
    market = data.get("market", "DE_TAX")
    requested_urls = data.get("image_urls")
    requested_declared_main = set(_dedupe_urls(data.get("declared_main_urls") or []))

    if not sku:
        return jsonify({"error": "SKU 不能为空"}), 400

    if requested_urls is not None:
        if not isinstance(requested_urls, list):
            return jsonify({"error": "image_urls 必须是数组"}), 400
        raw_urls = _dedupe_urls(requested_urls)[:REFERENCE_IMAGE_CANDIDATE_LIMIT]
        try:
            allowed_urls = _allowed_giga_reference_urls(sku, market)
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        invalid = [url for url in raw_urls if not _is_allowed_giga_reference_url(url, allowed_urls)]
        if invalid:
            return jsonify({"error": "参考图不属于当前 GIGA 产品"}), 400
        declared_main_urls = {
            url for url in requested_declared_main
            if _is_allowed_giga_reference_url(url, allowed_urls)
        }
    else:
        try:
            product = giga_fetch_product(sku, market)
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        reference_fields = _build_reference_image_fields(product)
        raw_urls = reference_fields["imageUrls"]
        declared_main_urls = set(reference_fields["mainImageUrls"])

    results: list[dict | None] = [None] * len(raw_urls)

    def fetch_one(i: int, url: str) -> tuple[int, dict]:
        proxied = _proxy_image_with_metadata(url)
        return i, {
            "index": i,
            "originalUrl": url,
            "dataUrl": (proxied or {}).get("dataUrl") or url,
            "width": (proxied or {}).get("width"),
            "height": (proxied or {}).get("height"),
            "failed": proxied is None,
        }

    # 最多 15 张并发代理,避免过度开线程
    with ThreadPoolExecutor(max_workers=max(1, min(len(raw_urls), TOTAL_REFERENCE_IMAGE_LIMIT))) as ex:
        futures = [ex.submit(fetch_one, i, url) for i, url in enumerate(raw_urls)]
        for fut in as_completed(futures):
            try:
                i, payload = fut.result(timeout=35)  # 单张兜底 35s
                results[i] = payload
            except Exception:
                # 单张失败不阻塞整体;占位 None 后面过滤掉
                continue

    # 下载失败时保留原 URL；分类会对未知尺寸回退到 GIGA 字段提示。
    records = []
    for i, url in enumerate(raw_urls):
        if results[i] is not None:
            records.append(results[i])
        else:
            records.append({
                "index": i,
                "originalUrl": url,
                "dataUrl": url,
                "width": None,
                "height": None,
                "failed": True,
            })

    grouped = _classify_reference_image_records(records, declared_main_urls)

    return jsonify({
        "success": True,
        "sku": sku,
        "market": market,
        "images": grouped["images"],
        "main_image_count": len(grouped["main"]),
        "detail_image_count": len(grouped["detail"]),
        "raw_image_count": grouped["raw_count"],
        "truncated_image_count": grouped["truncated_count"],
    })


# ─────────────────────────────────────────────────────────────────
# 启动
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if AUTH_ENABLED and ACCESS_PASSWORD_IS_TEMPORARY:
        print("  [SECURITY] 未配置 GIGAB2B_ACCESS_PASSWORD，已为本次启动生成临时访问密码:")
        print(f"  [SECURITY] {ACCESS_PASSWORD}")
        print("  [SECURITY] 重启后会变化；请在 .env 配置固定强密码。")

    print(f"  GIGAB2B Web App")
    print(f"  ={'='*50}")
    print(f"  Backend:  http://localhost:{PORT}")
    print(f"  Frontend: http://localhost:5173")
    print(f"  ={'='*50}")

    laozhang = _check_laozhang_provider()
    print(f"  laozhang:     {'OK' if laozhang['configured'] else '未配置（检查 .env 中 LAOZHANG_API_KEY）'} ({laozhang['model']})")

    has_giga = bool(os.environ.get("GIGA_DE_TAX_CLIENT_ID"))
    print(f"  GIGA 凭证:    {'OK' if has_giga else '未配置（检查.env）'}")
    print()

    # debug 模式由环境变量 FLASK_DEBUG 控制，默认关闭。
    # 关闭 debug 可以避免 Flask reloader 子进程退出后 socket 残留导致端口被占。
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    if not debug_mode:
        # 显式开启 SO_REUSEADDR，配合 Windows 端口立即回收
        # threaded=True 让 SSE 长连接不再阻塞其他 API 请求（致命 F-1 修复）
        from werkzeug.serving import make_server
        import socket as _socket
        bind_host = os.getenv("GIGAB2B_HOST", "127.0.0.1").strip() or "127.0.0.1"
        server = make_server(bind_host, PORT, app, threaded=True)
        server.socket.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
        print(f"  Server:       http://localhost:{PORT}  (threaded)")
        server.serve_forever()
    else:
        print(f"  Server:       http://localhost:{PORT}  (FLASK_DEBUG=1)")
        bind_host = os.getenv("GIGAB2B_HOST", "127.0.0.1").strip() or "127.0.0.1"
        app.run(host=bind_host, port=PORT, debug=True)
